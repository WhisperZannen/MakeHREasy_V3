import io
import os
import sqlite3
import tempfile
import unittest
import zipfile

from openpyxl import load_workbook


class SocialOAExportTest(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.db_path = os.path.join(self.temp_dir.name, 'hr_oa_test.db')
        os.environ['MAKE_HR_DB_PATH'] = self.db_path

        from database.init_db import init_database
        init_database(self.db_path)

        conn = sqlite3.connect(self.db_path)
        conn.execute("PRAGMA foreign_keys = ON")
        conn.execute(
            "INSERT INTO departments(dept_name, dept_category) VALUES ('测试部门', '生产')"
        )
        dept_id = conn.execute(
            "SELECT dept_id FROM departments WHERE dept_name='测试部门'"
        ).fetchone()[0]
        employees = [
            ('E001', '1001', '普通人员'),
            ('E002', '1002', '挂靠人员'),
            ('E003', None, '无工号新人'),
            ('E004', '1004', '下沉人员'),
        ]
        conn.executemany("""
            INSERT INTO employees(
                emp_id, person_id, employee_no, name, dept_id, status, join_company_date
            ) VALUES (?, 'person-' || ?, ?, ?, ?, '在职', '2026-07-01')
        """, [(emp_id, emp_id, employee_no, name, dept_id) for emp_id, employee_no, name in employees])
        conn.executemany("""
            INSERT INTO employee_profiles(
                emp_id, oa_social_account_no, oa_annuity_account_no
            ) VALUES (?, ?, ?)
        """, [
            ('E001', '10002120618', '10003264948'),
            ('E002', '99999999999', '99999999999'),
            ('E003', '0', '0'),
            ('E004', '0', '0'),
        ])
        conn.execute("""
            INSERT INTO employee_arrangements(
                emp_id, arrangement_type, start_date, status
            ) VALUES ('E002', 'proxy_social', '2026-01-01', 'active')
        """)
        conn.execute("""
            INSERT INTO employee_arrangements(
                emp_id, arrangement_type, start_date, payroll_included,
                labor_cost_included, status
            ) VALUES (
                'E004', 'down_secondment', '2026-07-01', 0, 0, 'active'
            )
        """)

        for emp_id, _, _ in employees:
            conn.execute("""
                INSERT INTO ss_monthly_records(
                    record_id, cost_month, emp_id, close_status
                ) VALUES (?, '2026-07', ?, 'draft')
            """, (f'2026-07_{emp_id}', emp_id))

        item_values = {
            'pension': (5000, 800, 400),
            'medical': (5000, 400, 100),
            'medical_serious': (5000, 0, 7),
            'unemp': (5000, 35, 15),
            'injury': (5000, 10, 0),
            'maternity': (5000, 35, 0),
            'fund': (5000, 600, 600),
            'annuity': (5000, 400, 75),
        }
        for item, (base, company, personal) in item_values.items():
            conn.execute("""
                INSERT INTO social_monthly_items(
                    item_record_id, monthly_record_id, cost_month, emp_id,
                    business_type_snapshot, insurance_item, base_amount,
                    company_amount, personal_amount, close_status
                ) VALUES (?, '2026-07_E001', '2026-07', 'E001',
                          'normal', ?, ?, ?, ?, 'draft')
            """, (f'2026-07_E001_{item}', item, base, company, personal))
        conn.execute("""
            INSERT INTO social_monthly_items(
                item_record_id, monthly_record_id, cost_month, emp_id,
                business_type_snapshot, insurance_item, base_amount,
                company_amount, personal_amount, close_status
            ) VALUES (
                '2026-07_E002_pension', '2026-07_E002', '2026-07', 'E002',
                'proxy_social', 'pension', 5000, 800, 400, 'draft'
            )
        """)
        conn.execute("""
            INSERT INTO social_monthly_items(
                item_record_id, monthly_record_id, cost_month, emp_id,
                business_type_snapshot, insurance_item, base_amount,
                company_amount, personal_amount,
                payment_channel_code,
                payment_export_included, oa_export_included, close_status
            ) VALUES (
                '2026-07_E004_fund', '2026-07_E004', '2026-07', 'E004',
                'down_secondment', 'fund', 5000, 600, 600,
                'province_public:fund',
                1, 0, 'draft'
            )
        """)
        conn.execute("""
            INSERT INTO social_monthly_items(
                item_record_id, monthly_record_id, cost_month, emp_id,
                business_type_snapshot, insurance_item, base_amount,
                company_amount, personal_amount, close_status
            ) VALUES (
                '2026-07_E003_pension', '2026-07_E003', '2026-07', 'E003',
                'normal', 'pension', 4498, 719.68, 359.84, 'draft'
            )
        """)
        conn.execute("""
            INSERT INTO ss_retroactive_records(
                retro_id, process_month, emp_id, retro_type,
                total_comp_retro, total_pers_retro, status
            ) VALUES ('R001', '2026-07', 'E001', '养老保险', 10, 5, '待推送到当期账单')
        """)
        conn.commit()
        conn.close()

        from modules import core_social_oa
        self.oa = core_social_oa

    def tearDown(self):
        os.environ.pop('MAKE_HR_DB_PATH', None)
        self.temp_dir.cleanup()

    def test_schema_seeds_exact_legacy_file_names(self):
        settings = self.oa.get_oa_export_settings()
        self.assertEqual(
            [row['file_name'] for row in settings],
            [
                '2022-工伤保险.xlsx',
                '2022-基本养老保险.xlsx',
                '2022-基本医疗保险.xlsx',
                '2022-年失业保险.xlsx',
                '2022-生育保险.xlsx',
                '2022-住房公积金.xlsx',
                '2023-年企业年金.xlsx',
            ],
        )

    def test_summary_excludes_proxy_and_blocks_missing_employee_number(self):
        summary = self.oa.get_oa_export_summary('2026-07')
        self.assertEqual(summary['excluded_proxy_people'], 1)
        self.assertEqual(summary['excluded_down_people'], 1)
        self.assertEqual(summary['missing_employee_numbers'], ['无工号新人'])
        pension = next(item for item in summary['items'] if item['insurance_item'] == 'pension')
        self.assertEqual(pension['row_count'], 2)

        with self.assertRaises(self.oa.OAExportValidationError):
            self.oa.build_oa_export_package('2026-07')

    def test_generated_workbooks_match_oa_sheet_and_column_contract(self):
        conn = sqlite3.connect(self.db_path)
        conn.execute("UPDATE employees SET employee_no='1003' WHERE emp_id='E003'")
        conn.commit()
        conn.close()

        package = self.oa.build_oa_export_package('2026-07')
        self.assertEqual(len(package['files']), 7)
        with zipfile.ZipFile(io.BytesIO(package['zip_content'])) as archive:
            self.assertEqual(
                archive.namelist(),
                [row['file_name'] for row in self.oa.get_oa_export_settings()],
            )

        pension_file = next(
            item for item in package['files'] if item['insurance_item'] == 'pension'
        )
        pension_book = load_workbook(io.BytesIO(pension_file['content']), data_only=True)
        self.assertEqual(pension_book.sheetnames, ['Sheet0', 'Columns'])
        self.assertEqual(pension_book['Columns'].sheet_state, 'hidden')
        self.assertEqual(
            [cell.value for cell in pension_book['Columns'][1]],
            self.oa.COMMON_COLUMNS,
        )
        pension_rows = list(pension_book['Sheet0'].iter_rows(values_only=True))
        self.assertEqual(list(pension_rows[0]), self.oa.COMMON_HEADERS)
        self.assertEqual(len(pension_rows), 3)
        first = pension_rows[1]
        self.assertEqual(first[1:4], ('1001', '普通人员', '10002120618'))
        self.assertEqual(first[10:14], (10, 5, 405, 810))

        medical_file = next(
            item for item in package['files'] if item['insurance_item'] == 'medical'
        )
        medical_book = load_workbook(io.BytesIO(medical_file['content']), data_only=True)
        medical_row = list(medical_book['Sheet0'].iter_rows(values_only=True))[1]
        self.assertEqual(medical_row[10], 7)
        self.assertEqual(medical_row[14], 107)

        annuity_file = next(
            item for item in package['files'] if item['insurance_item'] == 'annuity'
        )
        annuity_book = load_workbook(io.BytesIO(annuity_file['content']), data_only=True)
        annuity_row = list(annuity_book['Sheet0'].iter_rows(values_only=True))[1]
        self.assertEqual(annuity_row[3], '10003264948')

    def test_payment_scope_keeps_only_down_secondment_fund(self):
        from modules.core_social_security import load_payment_export_rows

        conn = sqlite3.connect(self.db_path)
        rows = load_payment_export_rows(conn, '2026-07', '2026-07')
        conn.close()

        down = rows[rows['emp_id'] == 'E004']
        self.assertEqual(len(down), 1)
        self.assertEqual(
            down.iloc[0]['payment_channel_code'],
            'province_public:fund',
        )
        self.assertEqual(down.iloc[0]['fund_comp'], 600.0)
        self.assertEqual(down.iloc[0]['fund_pers'], 600.0)
        self.assertEqual(down.iloc[0]['pension_comp'], 0.0)


if __name__ == '__main__':
    unittest.main()
