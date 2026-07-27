import io
import sqlite3
import unittest

import pandas as pd
from openpyxl import load_workbook

from modules.core_social_security import (
    INTERNAL_APPROVAL_TYPE_LABELS,
    load_internal_approval_base_snapshots,
    prepare_internal_approval_person_summary,
    write_internal_approval_sheet,
)


class InternalApprovalExportTest(unittest.TestCase):
    def _source(self):
        return pd.DataFrame([
            {
                'emp_id': 'N001', 'employee_no': '1001', '姓名': '普通员工',
                'cost_month': '2026-06', 'cost_center': '本级',
                'business_type_snapshot': 'normal',
                'social_base': 5000.0,
                'pension_comp': 800.0, 'pension_pers': 400.0,
                'unemp_comp': 35.0, 'unemp_pers': 15.0,
                'injury_comp': 10.0,
            },
            {
                'emp_id': 'N001', 'employee_no': '1001', '姓名': '普通员工',
                'cost_month': '2026-07', 'cost_center': '省公众',
                'business_type_snapshot': 'normal',
                'social_base': 6000.0,
                'pension_comp': 960.0, 'pension_pers': 480.0,
                'unemp_comp': 42.0, 'unemp_pers': 18.0,
                'injury_comp': 12.0,
            },
            {
                'emp_id': 'C001', 'employee_no': '1002', '姓名': '魏巍',
                'cost_month': '2026-07', 'cost_center': '省公众',
                'business_type_snapshot': 'city_transfer',
                'social_base': 7000.0,
                'pension_comp': 1120.0, 'pension_pers': 560.0,
                'unemp_comp': 49.0, 'unemp_pers': 21.0,
                'injury_comp': 14.0,
            },
            {
                'emp_id': 'D001', 'employee_no': '1003', '姓名': '下沉员工',
                'cost_month': '2026-07', 'cost_center': '黄石分公司',
                'business_type_snapshot': 'down_secondment',
                'social_base': 8000.0,
                'pension_comp': 1280.0, 'pension_pers': 640.0,
                'unemp_comp': 56.0, 'unemp_pers': 24.0,
                'injury_comp': 16.0,
            },
            {
                'emp_id': 'P001', 'employee_no': '1004', '姓名': '挂靠员工',
                'cost_month': '2026-07', 'cost_center': '襄阳分公司',
                'business_type_snapshot': 'proxy_social',
                'social_base': 9000.0,
                'pension_comp': 1440.0, 'pension_pers': 720.0,
                'unemp_comp': 63.0, 'unemp_pers': 27.0,
                'injury_comp': 18.0,
            },
        ])

    def test_summary_orders_people_and_keeps_latest_base(self):
        money_cols = [
            'pension_comp', 'pension_pers',
            'unemp_comp', 'unemp_pers', 'injury_comp',
        ]
        result = prepare_internal_approval_person_summary(
            self._source(), money_cols, base_cols=['social_base']
        )

        self.assertEqual(
            result['姓名'].tolist(),
            ['普通员工', '魏巍', '下沉员工', '挂靠员工'],
        )
        normal = result.iloc[0]
        self.assertEqual(normal['social_base'], 6000.0)
        self.assertEqual(normal['pension_comp'], 1760.0)
        self.assertEqual(normal['cost_center'], '省公众')

    def test_workbook_has_requested_headers_groups_and_formulas(self):
        money_cols = [
            'pension_comp', 'pension_pers',
            'unemp_comp', 'unemp_pers', 'injury_comp',
        ]
        summary = prepare_internal_approval_person_summary(
            self._source(), money_cols, base_cols=['social_base']
        )
        rename_map = {
            'employee_no': '工号',
            'cost_center': '财务归属',
            'business_type_snapshot': '人员类别',
            'social_base': '社保基数',
            'pension_comp': '养老企业',
            'pension_pers': '养老个人',
            'unemp_comp': '失业企业',
            'unemp_pers': '失业个人',
            'injury_comp': '工伤',
        }
        export = summary.rename(columns=rename_map)
        export['人员类别'] = (
            export['人员类别']
            .map(INTERNAL_APPROVAL_TYPE_LABELS)
            .fillna('特殊人员')
        )
        export = export[[
            '工号', '姓名', '人员类别', '财务归属', '社保基数',
            '养老企业', '养老个人', '失业企业', '失业个人', '工伤',
        ]]

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            write_internal_approval_sheet(
                writer, export, '2026-07', '测试提款单', '2026-07',
                money_columns=['养老企业', '养老个人', '失业企业', '失业个人', '工伤'],
                base_columns=['社保基数'],
            )

        workbook = load_workbook(io.BytesIO(output.getvalue()), data_only=False)
        sheet = workbook['2026-07']
        headers = [cell.value for cell in sheet[3]]
        self.assertEqual(headers, [
            '序号', '工号', '姓名', '人员类别', '财务归属', '社保基数',
            '养老企业', '养老个人', '失业企业', '失业个人', '工伤',
        ])

        names = [sheet.cell(row, 3).value for row in range(4, sheet.max_row + 1)]
        self.assertEqual(names, [
            '普通员工', '魏巍', '下沉员工',
            '本单位及特殊人员小计',
            '挂靠员工', '挂靠代缴小计', '合计',
        ])
        self.assertEqual(sheet['G7'].value, '=SUM(G4:G6)')
        self.assertEqual(sheet['G9'].value, '=SUM(G8:G8)')
        self.assertEqual(sheet['G10'].value, '=G7+G9')

    def test_base_snapshot_uses_capped_execution_base(self):
        connection = sqlite3.connect(':memory:')
        connection.execute("""
            CREATE TABLE social_monthly_items (
                cost_month TEXT, emp_id TEXT, insurance_item TEXT,
                base_amount REAL, company_amount REAL, personal_amount REAL,
                calculation_policy_entity TEXT
            )
        """)
        connection.execute("""
            CREATE TABLE ss_policy_versions (
                effective_from_month TEXT, manage_entity TEXT,
                pension_comp_rate REAL, pension_pers_rate REAL,
                medical_comp_rate REAL, medical_pers_rate REAL,
                unemp_comp_rate REAL, unemp_pers_rate REAL,
                injury_comp_rate REAL, maternity_comp_rate REAL,
                fund_comp_rate REAL, fund_pers_rate REAL,
                annuity_comp_rate REAL, annuity_pers_rate REAL
            )
        """)
        connection.execute("""
            INSERT INTO ss_policy_versions VALUES (
                '2026-01', '省公众',
                0.16, 0.08, 0.07, 0.02, 0.007, 0.003,
                0.002, 0.007, 0.12, 0.12, 0.08, 0.04
            )
        """)
        connection.execute("""
            INSERT INTO social_monthly_items VALUES (
                '2026-01', 'E001', 'pension',
                50920, 3598.08, 1799.04, '省公众'
            )
        """)

        result = load_internal_approval_base_snapshots(
            connection, '2026-01', '2026-01'
        )
        connection.close()

        self.assertEqual(len(result), 1)
        self.assertAlmostEqual(result.iloc[0]['social_base'], 22488.0, places=2)


if __name__ == '__main__':
    unittest.main()
