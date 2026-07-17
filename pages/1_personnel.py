# ==============================================================================
# 文件路径: pages/1_personnel.py
# 功能描述: 人员与组织架构管理中枢 (V3.20 终极防爆与高岗级排序修复版)
# 实现了什么具体逻辑:
#   1. [核心防御] 修复 Pandas `NaT` 判定歧义导致的 ValueError 页面崩溃。
#   2. [核心修复] 修正岗级倒挂：取负数强制高岗级领导排前，彻底修复领导排序失效。
#   3. [状态对齐] 全面接入“挂靠人员”状态，支持以社保编号作为虚拟工号录入。
# ==============================================================================

import streamlit as st
import pandas as pd
import io
from datetime import datetime, date
from dateutil.relativedelta import relativedelta

# 导入底层接口
from modules.core_dept import get_all_departments, add_department, update_department, soft_delete_department
from modules.core_position import get_all_positions, add_position, update_position
from modules.core_personnel import (
    get_all_employees, add_employee, update_employee, update_employee_status,
    get_all_history, rollback_history, batch_transfer_department_members,
)
from modules.core_arrangements import (
    ACTIVE_LABELS,
    ARRANGEMENT_CLOSE_RESULT_LABELS,
    ARRANGEMENT_LABELS,
    ARRANGEMENT_STATUS_LABELS,
    SETTLEMENT_CYCLE_LABELS,
    SETTLEMENT_MODE_LABELS,
    SPECIAL_ARRANGEMENT_TYPES,
    close_arrangement,
    create_business_entity,
    create_arrangement,
    end_person_social_override,
    get_effective_arrangement,
    get_arrangements_dataframe,
    get_arrangement_route_defaults,
    get_entities_dataframe,
    get_people_management_dataframe,
    get_person_treatment_dataframe,
    get_social_overrides_dataframe,
    INSURANCE_LABELS,
    PERSON_TREATMENT_ITEMS,
    save_arrangement_route_default,
    save_person_social_override,
    save_simple_arrangement,
    set_business_entity_active,
)
from modules.core_payroll import (
    end_payroll_identity,
    get_payroll_identities,
    save_payroll_identity,
)

st.set_page_config(page_title="组织人事中枢", layout="wide")

# ==============================================================================
# 自动初始化“公共挂靠池”
# ==============================================================================
def init_virtual_pools():
    ok_d, d_list = get_all_departments(include_inactive=True)
    if ok_d:
        d_names = [d['dept_name'] for d in d_list]
        if "离退休公共池" not in d_names: add_department("离退休公共池", "其他", None, 9999)

    ok_p, p_list = get_all_positions(include_inactive=True)
    if ok_p:
        p_names = [p['pos_name'] for p in p_list]
        if "无岗位" not in p_names: add_position("无岗位", "通用序列", 9999)

init_virtual_pools()

# ==============================================================================
# 消息状态保持与状态同步锁
# ==============================================================================
if 'ui_msg' in st.session_state:
    if st.session_state.ui_msg_type == 'success': st.success(st.session_state.ui_msg)
    else: st.error(st.session_state.ui_msg)
    del st.session_state.ui_msg, st.session_state.ui_msg_type

def set_msg_and_rerun(msg, type='success'):
    st.session_state.ui_msg = msg
    st.session_state.ui_msg_type = type
    st.session_state.editor_key = str(datetime.now())
    st.rerun()

if 'editor_key' not in st.session_state:
    st.session_state.editor_key = "init_v318"

# ==============================================================================
# 全局中文化映射
# ==============================================================================
DEPT_COL_MAP = {'dept_id': '部门ID', 'dept_name': '部门名称', 'dept_category': '性质', 'parent_dept_id': '上级ID', 'sort_order': '权重', 'status': '状态'}
POS_COL_MAP = {'pos_id': '岗位ID', 'pos_name': '岗位名称', 'pos_category': '序列', 'sort_order': '权重', 'status': '状态'}
EMP_COL_MAP = {
    'emp_id': '_internal_emp_id', 'employee_no': '工号/编号', 'name': '姓名', 'id_card': '身份证号', 'dept_id': '部门ID', 'pos_id': '岗位ID',
    'post_rank': '岗级', 'post_grade': '档次', 'join_company_date': '入职日期', 'status': '状态',
    'pos_name': '岗位', 'tech_grade': 'T级', 'dept_name': '部门'
}

def clean_str(val): return "" if pd.isna(val) or val is None or str(val).lower()=='nan' else str(val).strip()
def clean_date(d_val):
    try: return pd.to_datetime(d_val).strftime('%Y-%m-%d') if pd.notna(pd.to_datetime(d_val)) else None
    except: return None

def refresh_data():
    ok_d, d = get_all_departments(include_inactive=True)
    ok_p, p = get_all_positions(include_inactive=True)
    ok_e, e = get_all_employees(include_resigned=True)
    return (pd.DataFrame(d) if ok_d and d else pd.DataFrame(columns=list(DEPT_COL_MAP.keys())),
            pd.DataFrame(p) if ok_p and p else pd.DataFrame(columns=list(POS_COL_MAP.keys())),
            pd.DataFrame(e) if ok_e and e else pd.DataFrame(columns=list(EMP_COL_MAP.keys())))

df_depts, df_positions, df_emps = refresh_data()

def build_dept_tree(df, parent_id=None, level=0):
    tree = []
    if df.empty: return tree
    children = df[df['parent_dept_id'].isna() | (df['parent_dept_id'] == 0)] if (pd.isna(parent_id) or parent_id == 0) else df[df['parent_dept_id'] == parent_id]
    for _, row in children.iterrows():
        r_d = row.to_dict()
        r_d['层级展示名'] = "　　" * level + ("└─ " if level > 0 else "") + str(row['dept_name'])
        p_name = "无"
        if pd.notna(row['parent_dept_id']) and row['parent_dept_id'] != 0:
            pm = df[df['dept_id'] == row['parent_dept_id']]
            if not pm.empty: p_name = pm.iloc[0]['dept_name']
        r_d['上级名称'] = p_name
        tree.append(r_d)
        tree.extend(build_dept_tree(df, row['dept_id'], level + 1))
    return tree

# ==============================================================================
# 侧边栏导航
# ==============================================================================
st.sidebar.title("🎛️ 组织人事中枢")
# 严格按照你的要求调整了次序：人员 -> 流水 -> 部门 -> 岗位
current_page = st.sidebar.radio(
    "请选择操作模块:",
    ["👥 人员档案", "🧭 特殊人员与待遇", "🕰️ 历史变动流水", "🏢 部门管理", "🎯 岗位字典"]
)
st.title(current_page)

# ==============================================================================
# 模块 A: 🏢 部门管理
# ==============================================================================
if current_page == "🏢 部门管理":
    col_d1, col_d2 = st.columns([1, 2])
    with col_d1:
        st.subheader("📝 部门信息维护")
        edit_mode = st.radio("模式", ["新增部门", "修改现有部门"], horizontal=True)
        target_dept = None
        if edit_mode == "修改现有部门":
            if df_depts.empty: st.warning("无部门数据")
            else:
                d_dict = {r['dept_id']: r['dept_name'] for _, r in df_depts.iterrows()}
                sel_d_id = st.selectbox("选择要修改的部门", options=list(d_dict.keys()), format_func=lambda x: d_dict[x])
                target_dept = df_depts[df_depts['dept_id'] == sel_d_id].iloc[0]

        with st.form("dept_form"):
            d_name = st.text_input("部门名称*", value=target_dept['dept_name'] if target_dept is not None else "")
            cat_opts = ["公司领导", "管控", "生产", "其他"]
            def_cat = target_dept['dept_category'] if target_dept is not None else "管控"
            d_cat = st.selectbox("性质*", cat_opts, index=cat_opts.index(def_cat) if def_cat in cat_opts else 1)
            d_sort = st.number_input("权重*", value=int(target_dept['sort_order']) if target_dept is not None else 999)

            v_p = df_depts[df_depts['status'] == 1]
            if target_dept is not None: v_p = v_p[v_p['dept_id'] != target_dept['dept_id']]
            p_opts = {0: "无(顶级)"}
            for _, r in v_p.iterrows(): p_opts[r['dept_id']] = r['dept_name']

            def_p = int(target_dept['parent_dept_id']) if target_dept is not None and pd.notna(target_dept['parent_dept_id']) else 0
            d_parent = st.selectbox("上级部门", options=list(p_opts.keys()), format_func=lambda x: p_opts[x], index=list(p_opts.keys()).index(def_p) if def_p in p_opts else 0)

            def_stat = "正常" if target_dept is None or target_dept['status'] == 1 else "已撤销"
            d_status = st.selectbox("状态", ["正常", "已撤销"], index=["正常", "已撤销"].index(def_stat))

            if st.form_submit_button("保存部门"):
                if not d_name: st.error("名称必填")
                else:
                    s_val = 1 if d_status == "正常" else 0
                    p_val = int(d_parent) if d_parent != 0 else None
                    if edit_mode == "新增部门":
                        success, msg = add_department(d_name, d_cat, p_val, d_sort)
                    else:
                        success, msg = update_department(int(target_dept['dept_id']), d_name, d_cat, p_val, d_sort, s_val)
                    if success: set_msg_and_rerun(msg)
                    else: st.error(msg)

        with st.expander("📥 批量导入部门"):
            tmp = pd.DataFrame(columns=['部门名称', '性质', '上级名称', '权重'])
            out = io.BytesIO()
            with pd.ExcelWriter(out, engine='openpyxl') as w: tmp.to_excel(w, index=False)
            st.download_button("下载部门模板", data=out.getvalue(), file_name="部门导入模板.xlsx")

            df_file = st.file_uploader("上传 Excel", type=["xlsx"], key="d_up")
            if df_file and st.button("开始执行部门导入"):
                in_df = pd.read_excel(df_file)
                for _, r in in_df.iterrows():
                    nm = clean_str(r.get('部门名称'))
                    if nm: add_department(nm, clean_str(r.get('性质')) or "其他", None, int(r.get('权重', 999)) if pd.notna(r.get('权重')) else 999)

                _, rd = get_all_departments(include_inactive=True); fdf = pd.DataFrame(rd)
                for _, r in in_df.iterrows():
                    nm = clean_str(r.get('部门名称')); pnm = clean_str(r.get('上级名称'))
                    if nm and pnm:
                        tc = fdf[fdf['dept_name'] == nm]; tp = fdf[fdf['dept_name'] == pnm]
                        if not tc.empty and not tp.empty:
                            update_department(
                                int(tc.iloc[0]['dept_id']), str(tc.iloc[0]['dept_name']), str(tc.iloc[0]['dept_category']),
                                int(tp.iloc[0]['dept_id']), int(tc.iloc[0]['sort_order']), int(tc.iloc[0]['status'])
                            )
                set_msg_and_rerun("部门架构生成完毕")

    with col_d2:
        st.subheader("📊 组织架构树")
        s1, s2, s3 = st.columns([2, 1, 1])
        with s1: d_s = st.text_input("🔍 搜索部门")
        with s2: st.write(""); show_i = st.checkbox("含已撤销")
        with s3: st.write(""); show_pools = st.checkbox("显示系统人员池")

        fdf = df_depts.copy()
        if not fdf.empty:
            if not show_i: fdf = fdf[fdf['status'] == 1]
            if not show_pools and 'is_virtual_pool' in fdf.columns:
                fdf = fdf[fdf['is_virtual_pool'] != 1]
            if d_s:
                m_ids = set()
                dh = fdf[fdf['dept_name'].str.contains(d_s, na=False)]['dept_id'].tolist()
                def trace(did, all_df):
                    m_ids.add(did); pid = all_df[all_df['dept_id']==did].iloc[0]['parent_dept_id']
                    if pd.notna(pid) and pid != 0: trace(pid, all_df)
                for h in dh: trace(h, df_depts)
                fdf = df_depts[df_depts['dept_id'].isin(m_ids)]
                if not show_i: fdf = fdf[fdf['status'] == 1]
                if not show_pools and 'is_virtual_pool' in fdf.columns:
                    fdf = fdf[fdf['is_virtual_pool'] != 1]

            t_data = build_dept_tree(fdf)
            if t_data:
                tdf = pd.DataFrame(t_data)
                tdf['status'] = tdf['status'].apply(lambda x: "正常" if x == 1 else "已撤销")
                tdf = tdf.rename(columns=DEPT_COL_MAP)
                st.dataframe(tdf[['部门ID', '层级展示名', '性质', '权重', '状态']], use_container_width=True, hide_index=True)

    st.divider()
    with st.expander("组织调整：部门合并、撤销承接或拆分", expanded=False):
        st.caption(
            "选择原部门、人员和承接部门，系统会一次性生成人员调动流水。"
            "15日及以前生效的当月归新部门，16日以后当月仍归原部门。"
            "部门拆分时，按不同承接部门分几次操作即可。"
        )
        formal_active_depts = df_depts[
            (df_depts['status'] == 1)
            & (df_depts.get('is_virtual_pool', 0) != 1)
        ].copy()
        formal_dept_ids = formal_active_depts['dept_id'].astype(int).tolist()
        formal_dept_names = dict(zip(
            formal_active_depts['dept_id'].astype(int), formal_active_depts['dept_name']
        ))
        if len(formal_dept_ids) < 2:
            st.info("至少需要两个有效正式部门才能进行组织调整。")
        else:
            source_dept = st.selectbox(
                "原部门", formal_dept_ids,
                format_func=lambda value: formal_dept_names[value],
                key="org_source_dept",
            )
            source_people = df_emps[df_emps['dept_id'] == source_dept].copy()
            source_people_ids = source_people['emp_id'].astype(str).tolist()
            source_people_names = {
                str(row['emp_id']): f"{row['name']}（{row.get('employee_no') or '待分配'}）"
                for _, row in source_people.iterrows()
            }
            target_options = [dept_id for dept_id in formal_dept_ids if dept_id != source_dept]
            with st.form("organization_transfer_form"):
                target_dept = st.selectbox(
                    "承接部门", target_options,
                    format_func=lambda value: formal_dept_names[value],
                )
                selected_people = st.multiselect(
                    "需要调整的人员", source_people_ids,
                    default=source_people_ids,
                    format_func=lambda value: source_people_names[value],
                )
                transfer_date = st.date_input("生效日期", value=date.today())
                transfer_reason = st.text_input(
                    "调整说明*", placeholder="例如：研发运营五中心并入研发运营四中心"
                )
                deactivate_source = st.checkbox(
                    "人员全部转出后撤销原部门",
                    value=len(selected_people) == len(source_people_ids),
                )
                if st.form_submit_button("预览无误，执行组织调整", type="primary"):
                    ok, msg = batch_transfer_department_members(
                        selected_people, target_dept, transfer_date.isoformat(),
                        transfer_reason, source_dept, deactivate_source,
                    )
                    if ok:
                        set_msg_and_rerun(msg)
                    else:
                        st.error(msg)

# ==============================================================================
# 模块 B: 🎯 岗位字典
# ==============================================================================
elif current_page == "🎯 岗位字典":
    col_p1, col_p2 = st.columns([1, 2])
    with col_p1:
        st.subheader("📝 岗位维护")
        p_mode = st.radio("模式", ["新增岗位", "修改现有"], horizontal=True)
        t_pos = None
        if p_mode == "修改现有":
            if df_positions.empty: st.warning("无岗位")
            else:
                p_dict = {r['pos_id']: r['pos_name'] for _, r in df_positions.iterrows()}
                s_pid = st.selectbox("选择岗位", options=list(p_dict.keys()), format_func=lambda x: p_dict[x])
                t_pos = df_positions[df_positions['pos_id'] == s_pid].iloc[0]

        with st.form("pos_form"):
            p_name = st.text_input("岗位名称*", value=t_pos['pos_name'] if t_pos is not None else "")
            p_cat = st.text_input("岗位序列", value=t_pos['pos_category'] if t_pos is not None else "通用序列")
            p_sort = st.number_input("权重", value=int(t_pos['sort_order']) if t_pos is not None else 999)
            p_stat = "正常" if t_pos is None or t_pos['status'] == 1 else "停用"
            p_s_val = st.selectbox("状态", ["正常", "停用"], index=["正常", "停用"].index(p_stat))

            if st.form_submit_button("保存岗位"):
                if not p_name: st.error("必填")
                else:
                    sv = 1 if p_s_val == "正常" else 0
                    if p_mode == "新增岗位": ok, msg = add_position(p_name, p_cat, p_sort)
                    else: ok, msg = update_position(int(t_pos['pos_id']), p_name, p_cat, p_sort, sv)
                    if ok: set_msg_and_rerun(msg)
                    else: st.error(msg)

        with st.expander("📥 批量导入岗位字典"):
            tp = pd.DataFrame(columns=['岗位名称', '序列', '权重'])
            op = io.BytesIO()
            with pd.ExcelWriter(op, engine='openpyxl') as w: tp.to_excel(w, index=False)
            st.download_button("下载岗位模板", data=op.getvalue(), file_name="岗位导入模板.xlsx")

            pf = st.file_uploader("上传 Excel", type=["xlsx"], key="p_up")
            if pf and st.button("开始执行岗位导入"):
                idf = pd.read_excel(pf)
                sc = 0
                for _, r in idf.iterrows():
                    nm = clean_str(r.get('岗位名称'))
                    if nm:
                        ok, _ = add_position(nm, clean_str(r.get('序列')) or "通用序列", int(r.get('权重', 999)) if pd.notna(r.get('权重')) else 999)
                        if ok: sc += 1
                set_msg_and_rerun(f"导入完成，新增 {sc} 个岗位")

    with col_p2:
        st.subheader("📊 岗位清单")
        if not df_positions.empty:
            sh_p = st.checkbox("含已停用")
            fp = df_positions.copy()
            if not sh_p: fp = fp[fp['status']==1]
            fp['status'] = fp['status'].apply(lambda x: "正常" if x==1 else "停用")
            fp = fp.rename(columns=POS_COL_MAP)
            st.dataframe(fp[['岗位ID', '岗位名称', '序列', '权重', '状态']], use_container_width=True, hide_index=True)

# ==============================================================================
# 模块 C: 👥 人员档案
# ==============================================================================
elif current_page == "👥 人员档案":

    with st.expander("⏳ 实习生转正预警看板"):
        if not df_emps.empty and 'pos_name' in df_emps.columns:
            interns_df = df_emps[(df_emps['pos_name'] == '实习岗') & (df_emps['status'] == '在职')].copy()
            if not interns_df.empty:
                intern_list = []
                for _, intern in interns_df.iterrows():
                    start_p = intern['intern_start_date'] if pd.notna(intern['intern_start_date']) else intern['join_company_date']
                    start_d = pd.to_datetime(start_p)
                    # [核心防爆修复] 强制使用 pd.notna() 判断，防止出现 The truth value of a NaT is ambiguous 的致命崩溃！
                    if pd.notna(start_d):
                        months = 3 if str(intern.get('education_level')) == '硕士' else 6
                        expected = start_d + relativedelta(months=months)
                        days_left = (expected.date() - date.today()).days
                        st_str = f"🔴 已超期 {-days_left} 天" if days_left < 0 else (f"🟡 临近 ({days_left} 天)" if days_left <= 15 else f"🟢 正常 ({days_left} 天)")
                        intern_list.append({'工号': intern.get('employee_no') or '待分配', '姓名': intern['name'], '部门': intern['dept_name'], '预计转正': expected.strftime('%Y-%m-%d'), '当前状态': st_str})
                if intern_list: st.dataframe(pd.DataFrame(intern_list), use_container_width=True, hide_index=True)
            else: st.success("🎉 无待转正人员")

    st.subheader("🔍 人员检索与档案控制")
    c1, c2, c3, c4 = st.columns(4)
    with c1: q_name = st.text_input("工号/姓名/社保编号")
    with c2: q_dept = st.multiselect("部门", options=df_depts['dept_name'].tolist())
    with c3: q_status = st.multiselect("状态", options=["在职", "离职", "退休","挂靠人员"], default=["在职", "挂靠人员"])

    f_df = df_emps.copy()
    if q_name:
        f_df = f_df[
            f_df['name'].str.contains(q_name, na=False)
            | f_df['employee_no'].fillna('').astype(str).str.contains(q_name, na=False)
        ]
    if q_dept: f_df = f_df[f_df['dept_name'].isin(q_dept)]
    if q_status: f_df = f_df[f_df['status'].isin(q_status)]

    # ==========================================================================
    # [核心修复] 岗级倒挂 Bug 终结者！
    # ==========================================================================
    if not f_df.empty:
        # 提取部门与岗位权重
        dept_weight_map = {r['dept_name']: r['sort_order'] for _, r in df_depts.iterrows()} if not df_depts.empty else {}
        pos_weight_map = {r['pos_name']: r['sort_order'] for _, r in df_positions.iterrows()} if not df_positions.empty else {}

        def get_emp_sort_keys(row):
            d_name = str(row.get('dept_name', ''))
            p_name = str(row.get('pos_name', ''))
            status = str(row.get('status', ''))

            # 1. 部门大权重
            if status == '退休' or '离退休' in d_name: d_weight = 9999
            elif status == '挂靠人员': d_weight = 9000
            elif status == '公共账目' or '统筹' in d_name or '公共' in d_name: d_weight = 9998
            else: d_weight = dept_weight_map.get(d_name, 999)

            # 2. 岗位大权重
            p_weight = pos_weight_map.get(p_name, 999)

            # 3. 个人锚点小权重
            try: r_weight = float(row.get('post_rank', 9999.0)) if pd.notna(row.get('post_rank')) else 9999.0
            except: r_weight = 9999.0

            # [终极修复] 解决岗级倒挂：取负数让大岗级(如23)变成-23，在升序下排在-11前面。无岗级9999依然垫底！
            final_r_weight = -r_weight if r_weight != 9999.0 else 9999.0

            return (d_weight, p_weight, final_r_weight, str(row.get('emp_id', '')))

        # 执行强力排序
        f_df['__sort_tuple__'] = f_df.apply(get_emp_sort_keys, axis=1)
        f_df[['__dw__', '__pw__', '__rw__', '__id__']] = pd.DataFrame(f_df['__sort_tuple__'].tolist(), index=f_df.index)
        f_df = f_df.sort_values(by=['__dw__', '__pw__', '__rw__', '__id__'], ascending=[True, True, True, True])
        f_df = f_df.drop(columns=['__sort_tuple__', '__dw__', '__pw__', '__rw__', '__id__'])

    # ==========================================================================
    # 满血版全量美化导出引擎
    # ==========================================================================
    with c4:
        st.write("")
        if not f_df.empty:
            import sqlite3, os
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            def format_roster_sheet(worksheet, df_columns):
                worksheet.freeze_panes = 'A2'
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                for col_idx, col_name in enumerate(df_columns, 1):
                    col_letter = get_column_letter(col_idx)
                    worksheet.column_dimensions[col_letter].width = 20 if col_name in ['身份证号', '毕业院校', '专业'] else 13
                    for row_idx in range(1, worksheet.max_row + 1):
                        cell = worksheet[f"{col_letter}{row_idx}"]
                        cell.border = thin_border
                        if row_idx == 1:
                            cell.font = Font(bold=True)
                            cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')

            # 击穿底层：拉取扩展表数据实现“绝对全量”
            db_path = os.path.join('database', 'hr_core.db')
            conn = sqlite3.connect(db_path)
            profiles_df = pd.read_sql_query("SELECT * FROM employee_profiles", conn)
            conn.close()

            # 拼接并重命名
            out_df = pd.merge(f_df, profiles_df, on='emp_id', how='left')
            out_df = out_df.rename(columns={
                'employee_no': '工号/编号', 'name': '姓名', 'id_card': '身份证号', 'dept_name': '部门', 'pos_name': '岗位',
                'post_rank': '岗级', 'post_grade': '档次', 'tech_grade': 'T级', 'join_company_date': '入职日期',
                'status': '状态', 'education_level': '学历', 'degree': '学位', 'school_name': '毕业院校',
                'major': '专业', 'graduation_date': '毕业日期', 'first_job_date': '参加工作日期'
            })

            eo = ['工号/编号', '姓名', '部门', '岗位', '岗级', '档次', 'T级', '状态', '入职日期', '参加工作日期', '身份证号', '学历', '学位', '毕业院校', '专业', '毕业日期']
            out_df = out_df[[c for c in eo if c in out_df.columns]]

            ob = io.BytesIO()
            with pd.ExcelWriter(ob, engine='openpyxl') as w:
                out_df.to_excel(w, index=False, sheet_name='员工花名册')
                format_roster_sheet(w.sheets['员工花名册'], out_df.columns)

            st.download_button("📥 导出全量美化名单", data=ob.getvalue(), file_name=f"人员档案_{date.today()}.xlsx", type="primary")
        else:
            st.button("📥 导出全量美化名单", disabled=True)

    # 渲染到前端界面的展示列
    disp = f_df.rename(columns=EMP_COL_MAP)
    t_emp_sel = None
    if not disp.empty:
        target_cols = ['_internal_emp_id', '工号/编号', '姓名', '部门', '岗位', 'T级', '岗级', '档次', '状态']
        ui_cols = [c for c in target_cols if c in disp.columns]

        disp_ui = disp[ui_cols].copy()
        disp_ui.insert(0, "✅勾选修改", False)
        st.caption("上方名单只用于勾选人员，工号和档案内容请在下方“单条维护区”修改并保存。")
        edited_df = st.data_editor(
            disp_ui, hide_index=True, use_container_width=True,
            key=st.session_state.editor_key,
            column_config={
                "✅勾选修改": st.column_config.CheckboxColumn(required=True),
                "_internal_emp_id": None,
            },
            disabled=ui_cols,
        )
        selected_rows = edited_df[edited_df["✅勾选修改"] == True]
        if not selected_rows.empty:
            sel_id = selected_rows.iloc[0]['_internal_emp_id']

            import sqlite3, os
            db_path = os.path.join('database', 'hr_core.db')
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            full_emp_df = pd.read_sql_query("""
                SELECT e.*, p.* FROM employees e
                LEFT JOIN employee_profiles p ON e.emp_id = p.emp_id
                WHERE e.emp_id = ?
            """, conn, params=[sel_id])
            conn.close()

            t_emp_sel = full_emp_df.iloc[0].to_dict() if not full_emp_df.empty else df_emps[df_emps['emp_id'] == sel_id].iloc[0]
            st.info(f"已锁定: {t_emp_sel['name']}")

    st.divider()

    with st.expander("📥 批量智能导入 (支持离退免检与存在即更新)"):
        st.info("💡 若状态填为【离职/退休】，部门岗位可留空，系统将自动编入【离退休公共池】；若状态为【挂靠人员】，工号可填入社保编号。")
        t1, t2 = st.columns(2)
        with t1:
            icols = ['工号', '姓名', '状态', '所属部门', '岗位', '技术等级(T级)', '身份证号', '岗级', '档次', '入职日期', '参加工作日期', '首次就业(是/否)', '学历', '学位', '毕业院校', '专业', '毕业日期']
            tmp = pd.DataFrame(columns=icols); tout = io.BytesIO()
            with pd.ExcelWriter(tout) as w: tmp.to_excel(w, index=False)
            st.download_button("下载人员模板", data=tout.getvalue(), file_name="人员导入模板.xlsx")
        with t2:
            uf = st.file_uploader("上传 Excel", type=["xlsx"], key="emp_up")
            if uf and st.button("开始执行智能导入"):
                idf = pd.read_excel(uf); sc = 0; uc = 0; errs = []
                existing_by_no = {
                    clean_str(row.get('employee_no')): str(row['emp_id'])
                    for _, row in df_emps.iterrows() if clean_str(row.get('employee_no'))
                }
                existing_by_idcard = {
                    clean_str(row.get('id_card')): str(row['emp_id'])
                    for _, row in df_emps.iterrows() if clean_str(row.get('id_card'))
                }
                _, c_d = get_all_departments(True); d_df = pd.DataFrame(c_d)
                _, c_p = get_all_positions(True); p_df = pd.DataFrame(c_p)
                for idx, row in idf.iterrows():
                    eid = clean_str(row.get('工号')) or None
                    idc_raw = clean_str(row.get('身份证号'))
                    internal_id = existing_by_no.get(eid) if eid else None
                    if not internal_id and idc_raw:
                        internal_id = existing_by_idcard.get(idc_raw)
                    if not clean_str(row.get('姓名')):
                        continue
                    status_val = clean_str(row.get('状态')) or '在职'
                    if status_val in ['离职', '退休']:
                        idn = "离退休公共池"; ipn = "无岗位"; rank_val = 0; grade_val = "-"
                    else:
                        idn = clean_str(row.get('所属部门')) or clean_str(row.get('部门'))
                        ipn = clean_str(row.get('岗位')) or clean_str(row.get('岗位名称'))
                        rank_val = int(row.get('岗级', 11)) if pd.notna(row.get('岗级')) else 11
                        grade_val = clean_str(row.get('档次')) or 'E'

                    td = None
                    if idn:
                        md = d_df[d_df['dept_name']==idn]
                        if not md.empty: td = int(md.iloc[0]['dept_id'])
                        else:
                            add_department(idn, "其他", None, 999); _, rd = get_all_departments(True); d_df = pd.DataFrame(rd)
                            td = int(d_df[d_df['dept_name']==idn].iloc[0]['dept_id'])
                    tp = None
                    if ipn:
                        mp = p_df[p_df['pos_name']==ipn]
                        if not mp.empty: tp = int(mp.iloc[0]['pos_id'])
                        else:
                            add_position(ipn, "通用", 999); _, rp = get_all_positions(True); p_df = pd.DataFrame(rp)
                            tp = int(p_df[p_df['pos_name']==ipn].iloc[0]['pos_id'])

                    ed = {'employee_no': eid, 'name': clean_str(row.get('姓名')), 'id_card': idc_raw if idc_raw else None, 'dept_id': td, 'post_rank': rank_val, 'post_grade': grade_val, 'status': status_val, 'join_company_date': clean_date(row.get('入职日期'))}
                    pd_info = {
                        'pos_id': tp,
                        'tech_grade': clean_str(row.get('技术等级(T级)')),
                        'title_order': 999,
                        'education_level': clean_str(row.get('学历')),
                        'degree': clean_str(row.get('学位')),
                        'school_name': clean_str(row.get('毕业院校')),
                        'major': clean_str(row.get('专业')),
                        'graduation_date': clean_date(row.get('毕业日期')),
                        'first_job_date': clean_date(row.get('参加工作日期')),
                        'employment_stage': 'intern' if ipn == '实习岗' else 'regular',
                        'first_employment': 1 if clean_str(row.get('首次就业(是/否)')) in {'是', '1', 'true', 'True'} else 0,
                    }

                    if internal_id:
                        ok, msg = update_employee(internal_id, ed, pd_info, reason="Excel批量覆盖更新")
                        if ok: uc += 1
                        else: errs.append(f"行{idx+2}更新失败: {msg}")
                    else:
                        ok, msg = add_employee(ed, pd_info, "初始批量导入")
                        if ok: sc += 1
                        else: errs.append(f"行{idx+2}新增失败: {msg}")
                if errs: st.error("部分记录异常:\n" + "\n".join(errs))
                set_msg_and_rerun(f"完成！新增 {sc} 人, 更新 {uc} 人")

    with st.expander("📝 单条维护区", expanded=True):
        if t_emp_sel is None: st.info("新增人员时工号可以暂缺，系统会自动建立隐藏内部编号；正式工号下发后再回到这里补录。")
        with st.form("emp_form", clear_on_submit=True):
            f1, f2, f3 = st.columns(3)
            with f1:
                fid = st.text_input(
                    "工号/人力编码（可暂缺）",
                    value=(str(t_emp_sel.get('employee_no') or '') if t_emp_sel is not None else ""),
                    help="工号必须唯一；暂未下发时可以留空，后续可直接补充或修改。",
                    key=f"employee_no_input_{str(t_emp_sel.get('emp_id')) if t_emp_sel is not None else 'new'}",
                )
                fname = st.text_input("姓名*", value=str(t_emp_sel.get('name', "")) if t_emp_sel is not None else "")
                fidc = st.text_input("身份证", value=str(t_emp_sel.get('id_card', "")) if t_emp_sel is not None and pd.notna(t_emp_sel.get('id_card')) else "")
            with f2:
                def_d = t_emp_sel.get('dept_id') if t_emp_sel is not None else None
                vd = df_depts[df_depts['status']==1].copy()
                # 编辑仍挂在已撤销部门的人员时，保留其当前部门供展示，避免下拉框
                # 静默跳到第一个有效部门并制造一次误调动。
                if def_d is not None and def_d not in vd['dept_id'].tolist():
                    current_inactive = df_depts[df_depts['dept_id'] == def_d]
                    if not current_inactive.empty:
                        vd = pd.concat([current_inactive, vd], ignore_index=True)
                dm = {
                    r['dept_id']: (
                        f"{r['dept_name']}（已撤销，请迁出）" if r['status'] != 1 else r['dept_name']
                    )
                    for _, r in vd.iterrows()
                }
                fdept = st.selectbox("部门*", options=list(dm.keys()), format_func=lambda x: dm[x], index=list(dm.keys()).index(def_d) if def_d in dm else 0) if dm else None

                frank = st.number_input("岗级*", 0.0, 28.0,
                                        float(t_emp_sel.get('post_rank', 11.0)) if t_emp_sel is not None and pd.notna(
                                            t_emp_sel.get('post_rank')) else 11.0, step=0.1, format="%.1f")

                g_opts = ["-","A","B","C","D","E","F","G","H","I","J"]
                cur_g = str(t_emp_sel.get('post_grade', 'E')) if t_emp_sel is not None and pd.notna(t_emp_sel.get('post_grade')) else 'E'
                fgrade = st.selectbox("档次*", g_opts, index=g_opts.index(cur_g) if cur_g in g_opts else 5)
            with f3:
                def_p = t_emp_sel.get('pos_id') if t_emp_sel is not None else None
                vp = df_positions[df_positions['status']==1].copy()
                # 与部门处理保持一致：当前岗位即使刚被停用，也要原样展示，
                # 避免编辑其他字段时下拉框静默跳到第一个有效岗位。
                if def_p is not None and def_p not in vp['pos_id'].tolist():
                    current_inactive_pos = df_positions[df_positions['pos_id'] == def_p]
                    if not current_inactive_pos.empty:
                        vp = pd.concat([current_inactive_pos, vp], ignore_index=True)
                pos_name_map = {
                    r['pos_id']: r['pos_name'] for _, r in vp.iterrows()
                }
                pm = {
                    r['pos_id']: (
                        f"{r['pos_name']}（已停用，请调整）"
                        if r['status'] != 1 else r['pos_name']
                    )
                    for _, r in vp.iterrows()
                }
                fpos = st.selectbox("岗位*", options=list(pm.keys()), format_func=lambda x: pm[x], index=list(pm.keys()).index(def_p) if def_p in pm else 0) if pm else None

                ftg = st.text_input("T级", value=str(t_emp_sel.get('tech_grade', "")) if t_emp_sel is not None and pd.notna(t_emp_sel.get('tech_grade')) else "")

                fjoin_val = pd.to_datetime(t_emp_sel.get('join_company_date')) if t_emp_sel is not None and pd.notna(
                    t_emp_sel.get('join_company_date')) else date.today()
                fjoin = st.date_input("入职日", value=fjoin_val, min_value=date(1950, 1, 1))

                st.write("**--- 附加学籍与工作时间 ---**")
                fp1, fp2, fp3 = st.columns(3)
                with fp1:
                    e_opts = ["", "中专", "大专", "本科", "硕士", "博士"]
                    cur_edu = str(t_emp_sel.get('education_level', '')) if t_emp_sel is not None and pd.notna(
                        t_emp_sel.get('education_level')) else ""
                    f_edu = st.selectbox("学历", e_opts, index=e_opts.index(cur_edu) if cur_edu in e_opts else 0)
                    f_degree = st.text_input("学位", value=str(
                        t_emp_sel.get('degree', "")) if t_emp_sel is not None and pd.notna(
                        t_emp_sel.get('degree')) else "")
                with fp2:
                    f_school = st.text_input("毕业院校", value=str(
                        t_emp_sel.get('school_name', "")) if t_emp_sel is not None and pd.notna(
                        t_emp_sel.get('school_name')) else "")
                    f_major = st.text_input("所学专业",
                                            value=str(t_emp_sel.get('major', "")) if t_emp_sel is not None and pd.notna(
                                                t_emp_sel.get('major')) else "")
                with fp3:
                    grad_raw = t_emp_sel.get('graduation_date') if t_emp_sel is not None else None
                    f_grad_date_val = pd.to_datetime(grad_raw).date() if pd.notna(grad_raw) and str(grad_raw).strip() != '' else None
                    f_grad_date = st.date_input("毕业日期", value=f_grad_date_val, min_value=date(1950, 1, 1))

                    first_work_raw = t_emp_sel.get('first_job_date') if t_emp_sel is not None else None
                    f_first_work_val = pd.to_datetime(first_work_raw).date() if pd.notna(first_work_raw) and str(first_work_raw).strip() != '' else None
                    f_first_work = st.date_input("参加工作时间*", value=f_first_work_val, min_value=date(1950, 1, 1))

                is_intern_position = pos_name_map.get(fpos) == '实习岗' if fpos else False
                existing_first_employment = bool(
                    int(t_emp_sel.get('first_employment', 0) or 0)
                ) if t_emp_sel is not None else False
                f_first_employment = st.checkbox(
                    "首次参加工作（公积金从入职次月开始）",
                    value=existing_first_employment,
                    help="只针对首次就业的新入职人员；有过工作经历的人员不要勾选。",
                )
                if is_intern_position:
                    months = 3 if f_edu in {'硕士', '研究生'} else 6
                    expected_regular = fjoin + relativedelta(months=months)
                    st.info(
                        f"系统识别为实习期：预计 {expected_regular:%Y-%m-%d} 转正。"
                        "到期只提醒，不会自动转正；将岗位改为正式岗位并保存后，才记录实际转正。"
                    )

            st.write("**--- 状态与快照控制 ---**")
            cs1, cs2, cs3 = st.columns(3)
            with cs1:
                cur_s = str(t_emp_sel.get('status', '在职')) if t_emp_sel is not None and pd.notna(t_emp_sel.get('status')) else "在职"
                s_opts = ["在职", "挂靠人员", "离职", "退休"]
                fst = st.selectbox("当前状态", s_opts, index=s_opts.index(cur_s) if cur_s in s_opts else 0)
            with cs2: fcd = st.date_input("生效日期*", value=date.today())
            with cs3: frsn = st.text_input("变动说明*", placeholder="必填")

            if st.form_submit_button("保存并生成快照"):
                if not fname or fdept is None: st.error("姓名和部门不能为空")
                elif t_emp_sel is not None and not frsn: st.error("必填说明")
                else:
                    idc_val = fidc.strip() if fidc.strip() else None
                    ed = {'employee_no': fid.strip() or None, 'name': fname, 'id_card': idc_val, 'dept_id': int(fdept),
                          'post_rank': float(frank), 'post_grade': fgrade, 'status': fst,
                          'join_company_date': fjoin.strftime('%Y-%m-%d')}
                    pd_i = {
                        'pos_id': int(fpos) if fpos else None,
                        'tech_grade': ftg,
                        'title_order': 999,
                        'education_level': f_edu,
                        'degree': f_degree,
                        'school_name': f_school,
                        'major': f_major,
                        'graduation_date': f_grad_date.strftime('%Y-%m-%d') if f_grad_date else None,
                        'first_job_date': f_first_work.strftime('%Y-%m-%d') if f_first_work else None,
                        'employment_stage': 'intern' if is_intern_position else 'regular',
                        'first_employment': 1 if f_first_employment else 0,
                        'expected_regularization_date': (
                            t_emp_sel.get('expected_regularization_date')
                            if t_emp_sel is not None else None
                        ),
                        'actual_regularization_date': (
                            t_emp_sel.get('actual_regularization_date')
                            if t_emp_sel is not None else None
                        ),
                    }
                    ad_str = fcd.strftime('%Y-%m-%d %H:%M:%S')
                    if t_emp_sel is not None: ok, msg = update_employee(str(t_emp_sel['emp_id']), ed, pd_i, reason=frsn, change_date=ad_str)
                    else: ok, msg = add_employee(ed, pd_i, reason=frsn or "手工录入", change_date=ad_str)
                    if ok: set_msg_and_rerun(msg)
                    else: st.error(msg)

# ==============================================================================
# 模块 D: 特殊人员与待遇（业务人员简化入口）
# ==============================================================================
elif current_page == "🧭 特殊人员与待遇":
    person_settings_tab, default_rules_tab, payroll_identity_tab = st.tabs([
        "人员情形与个人待遇例外",
        "下沉/地市转入默认规则",
        "薪酬身份与聘期",
    ])
    with person_settings_tab:
        st.subheader("🧭 特殊人员与待遇设置")
        st.caption(
            "先确定这个人属于哪种情形，再只维护与普通人员不同的项目。"
            "办理单位和人工成本归属可以不同，系统会自动生成结算处理。"
        )

        target_month = st.text_input(
            "查看和设置月份", value=date.today().strftime('%Y-%m'), max_chars=7,
            help="所有人员情形和单项例外都按生效期保存，修改未来月份不会改写历史账。",
        )
        people_df = get_people_management_dataframe(target_month)
        if people_df.empty:
            st.info("当前没有可设置的在职或挂靠人员。")
            st.stop()

        metric_cols = st.columns(5)
        metric_cols[0].metric("普通人员", int((people_df['arrangement_type'] == 'normal').sum()))
        metric_cols[1].metric("挂靠代缴", int((people_df['arrangement_type'] == 'proxy_social').sum()))
        metric_cols[2].metric("下沉人员", int((people_df['arrangement_type'] == 'down_secondment').sum()))
        metric_cols[3].metric("地市转入", int((people_df['arrangement_type'] == 'city_transfer').sum()))
        metric_cols[4].metric("有单项例外", int((people_df['个人例外数'] > 0).sum()))

        f1, f2 = st.columns([1, 2])
        with f1:
            situation_filter = st.selectbox(
                "人员筛选", ["全部"] + list(ARRANGEMENT_LABELS.values()) + ["有单项例外"]
            )
        filtered_people = people_df.copy()
        if situation_filter == "有单项例外":
            filtered_people = filtered_people[filtered_people['个人例外数'] > 0]
        elif situation_filter != "全部":
            filtered_people = filtered_people[filtered_people['人员情形'] == situation_filter]
        with f2:
            person_options = filtered_people['emp_id'].astype(str).tolist()
            person_labels = {
                str(row['emp_id']): (
                    f"{row['name']}（{row.get('employee_no') or '待分配'}）｜{row['人员情形']}｜{row['dept_name'] or '未分配部门'}"
                )
                for _, row in filtered_people.iterrows()
            }
            selected_emp_id = st.selectbox(
                "选择人员", person_options,
                format_func=lambda value: person_labels[value],
                placeholder="没有符合筛选条件的人员",
            )
        if not selected_emp_id:
            st.stop()

        selected_person = people_df[
            people_df['emp_id'].astype(str) == str(selected_emp_id)
        ].iloc[0]
        arrangement = get_effective_arrangement(str(selected_emp_id), target_month)
        relation_type = arrangement.get('arrangement_type', 'normal')

        st.divider()
        s1, s2, s3, s4 = st.columns(4)
        s1.markdown(f"**{selected_person['name']}**\n\n{selected_person.get('employee_no') or '工号待分配'}")
        s2.markdown(f"**内部部门**\n\n{selected_person['dept_name'] or '未分配部门'}")
        s3.markdown(f"**人员情形**\n\n{ARRANGEMENT_LABELS.get(relation_type, relation_type)}")
        s4.markdown(
            f"**系统结论**\n\n{selected_person['工资处理']}；{selected_person['人工成本处理']}"
        )

        if relation_type == 'city_transfer':
            st.success("正式转入人员：属于本单位人员，工资和人工成本默认计入本单位；来源地市仅用于年度成本划转。")
        elif relation_type == 'down_secondment':
            st.info("下沉人员：社保中仍可保留由本单位办理的项目，但费用不进入本单位人工成本，按期间导出结算。")
        elif relation_type == 'proxy_social':
            st.info("挂靠代缴：只记录实际代缴项目，不发本单位工资，也不进入本单位人工成本。")
        elif int(selected_person['个人例外数']) > 0:
            st.warning("该人员仍是普通人员，但存在个别待遇项目例外。李峰林应按这种方式管理。")

        entity_df = get_entities_dataframe(active_only=True)
        entity_names = dict(zip(entity_df['entity_code'], entity_df['entity_name']))
        external_entities = entity_df[entity_df['entity_code'] != 'province_public']
        external_options = external_entities['entity_code'].tolist()

        with st.expander("① 设置人员情形", expanded=relation_type != 'normal'):
            relation_options = list(ARRANGEMENT_LABELS)
            selected_relation_type = st.radio(
                "这个人属于哪种情形？",
                relation_options,
                format_func=lambda value: ARRANGEMENT_LABELS[value],
                index=relation_options.index(relation_type) if relation_type in relation_options else 0,
                horizontal=True,
                key=f"arrangement_type_{selected_emp_id}",
            )
            with st.form("simple_arrangement_form"):
                is_switching_situation = selected_relation_type != relation_type
                existing_start = None if is_switching_situation else arrangement.get('start_date')
                try:
                    start_default = pd.to_datetime(existing_start).date() if existing_start else date.today()
                except Exception:
                    start_default = date.today()
                start_value = st.date_input(
                    "生效日期", value=start_default,
                    key=f"arrangement_start_{selected_emp_id}_{selected_relation_type}",
                )

                related_unit = None
                actual_work_unit = None
                payroll_included = True
                labor_cost_included = True
                planned_end = None
                document_no = ""
                relation_remarks = ""
                if selected_relation_type != 'normal':
                    c1, c2 = st.columns(2)
                    current_related = (
                        None if is_switching_situation else arrangement.get('related_branch_code')
                    )
                    current_actual = (
                        None if is_switching_situation else arrangement.get('actual_work_unit_code')
                    )
                    with c1:
                        related_unit = st.selectbox(
                            "来源/关联地市或单位*", external_options,
                            format_func=lambda value: entity_names.get(value, value),
                            index=(
                                external_options.index(current_related)
                                if current_related in external_options else None
                            ),
                            placeholder="请选择关联单位",
                            key=f"related_unit_{selected_emp_id}_{selected_relation_type}",
                        ) if external_options else None
                        actual_work_unit = st.selectbox(
                            "实际工作单位（不填则同关联单位）", external_options,
                            format_func=lambda value: entity_names.get(value, value),
                            index=external_options.index(current_actual) if current_actual in external_options else (
                                external_options.index(related_unit) if related_unit in external_options else None
                            ),
                            placeholder="默认与关联单位一致",
                            key=f"actual_unit_{selected_emp_id}_{selected_relation_type}",
                        ) if external_options else None
                    with c2:
                        if selected_relation_type == relation_type:
                            default_payroll = int(arrangement.get('payroll_included', 1))
                            default_labor = int(arrangement.get('labor_cost_included', 1))
                        else:
                            default_payroll = 0 if selected_relation_type in {'proxy_social', 'down_secondment'} else 1
                            default_labor = 0 if selected_relation_type in {'proxy_social', 'down_secondment'} else 1
                        payroll_included = st.checkbox(
                            "工资由本系统发放", value=bool(default_payroll),
                            key=f"payroll_scope_{selected_emp_id}_{selected_relation_type}",
                        )
                        labor_cost_included = st.checkbox(
                            "计入本单位人工成本", value=bool(default_labor),
                            key=f"labor_scope_{selected_emp_id}_{selected_relation_type}",
                        )
                        if selected_relation_type in {'proxy_social', 'down_secondment'}:
                            st.caption("该类人员默认两项都不勾选；只有政策明确改变时才手工开启。")
                        if selected_relation_type == 'down_secondment':
                            try:
                                planned_default = pd.to_datetime(
                                    arrangement.get('planned_end_date')
                                ).date() if arrangement.get('planned_end_date') else start_value + relativedelta(years=2)
                            except Exception:
                                planned_default = start_value + relativedelta(years=2)
                            planned_end = st.date_input(
                                "计划结束日期", value=planned_default,
                                key=f"planned_end_{selected_emp_id}_{selected_relation_type}",
                            )
                    document_no = st.text_input(
                        "文件或协议编号",
                        value=(
                            "" if is_switching_situation
                            else str(arrangement.get('source_document_no') or '')
                        ),
                        key=f"arrangement_document_{selected_emp_id}_{selected_relation_type}",
                    )
                    relation_remarks = st.text_area(
                        "说明", value=(
                            "" if is_switching_situation else str(arrangement.get('remarks') or '')
                        ),
                        placeholder="只写业务上需要记住的说明即可。",
                        key=f"arrangement_remarks_{selected_emp_id}_{selected_relation_type}",
                    )
                elif selected_person['status'] == '挂靠人员':
                    st.warning("该人员档案状态仍是“挂靠人员”。若要恢复普通人员，请先在人员档案中改为“在职”。")

                if st.form_submit_button("保存人员情形", type="primary"):
                    ok, msg = save_simple_arrangement({
                        'emp_id': str(selected_emp_id),
                        'arrangement_type': selected_relation_type,
                        'related_branch_code': related_unit,
                        'actual_work_unit_code': actual_work_unit,
                        'start_date': start_value.isoformat(),
                        'planned_end_date': planned_end.isoformat() if planned_end else None,
                        'payroll_included': payroll_included,
                        'labor_cost_included': labor_cost_included,
                        'source_document_no': document_no,
                        'remarks': relation_remarks,
                    })
                    if ok:
                        set_msg_and_rerun(msg)
                    else:
                        st.error(msg)

        st.write("### ② 当前待遇办理结果")
        treatment_df = get_person_treatment_dataframe(str(selected_emp_id), target_month)
        st.dataframe(
            treatment_df[[
                '项目', '是否缴纳', '办理单位', '成本归属', '系统处理结果', '规则来源'
            ]],
            use_container_width=True,
            hide_index=True,
        )

        with st.expander("③ 个人待遇例外（只改与默认规则不同的项目）", expanded=False):
            st.caption(
                "只有和普通规则不同的项目才在这里设置。例如李峰林只需要设置“工伤保险”。"
                "没有设置的项目自动沿用普通人员或该类人员规则。"
            )
            with st.form("simple_person_override_form"):
                o1, o2, o3 = st.columns(3)
                with o1:
                    override_item = st.selectbox(
                        "待遇项目", PERSON_TREATMENT_ITEMS,
                        format_func=lambda value: INSURANCE_LABELS[value],
                    )
                    override_enabled = st.checkbox("本项目需要缴纳", value=True)
                with o2:
                    payer_options = entity_df['entity_code'].tolist()
                    default_payer_index = payer_options.index('province_public') if 'province_public' in payer_options else 0
                    override_payer = st.selectbox(
                        "由谁实际办理缴费", payer_options,
                        format_func=lambda value: entity_names.get(value, value),
                        index=default_payer_index,
                    )
                    include_company_cost = st.checkbox("计入本单位人工成本", value=True)
                with o3:
                    external_cost_bearer = st.selectbox(
                        "不计入本单位时，由谁承担", external_options,
                        format_func=lambda value: entity_names.get(value, value),
                        disabled=include_company_cost,
                    ) if external_options else None
                    override_from = st.text_input("生效月份", value=target_month, max_chars=7)
                    has_override_end = st.checkbox("设置结束月份")
                    override_to = st.text_input(
                        "结束月份", value=target_month, max_chars=7,
                        disabled=not has_override_end,
                    )
                override_reason = st.text_input(
                    "特殊原因*", placeholder="例如：一建资质要求工资、合同、工伤缴费主体一致"
                )
                override_document = st.text_input("依据文件（可选）")
                if st.form_submit_button("保存这个项目的特殊设置", type="primary"):
                    ok, msg = save_person_social_override(
                        str(selected_emp_id), override_item, override_from.strip(),
                        override_enabled, override_payer, include_company_cost,
                        external_cost_bearer, override_reason,
                        override_to.strip() if has_override_end else None,
                        override_document,
                    )
                    if ok:
                        set_msg_and_rerun(msg)
                    else:
                        st.error(msg)

        all_overrides = get_social_overrides_dataframe(active_only=True)
        person_overrides = all_overrides[
            all_overrides['emp_id'].astype(str) == str(selected_emp_id)
        ] if not all_overrides.empty else pd.DataFrame()
        if not person_overrides.empty:
            with st.expander("查看或结束已有单项例外", expanded=False):
                person_overrides = person_overrides.copy()
                person_overrides['项目'] = person_overrides['insurance_item'].map(INSURANCE_LABELS)
                person_overrides['办理单位'] = person_overrides['payer_entity_name']
                person_overrides['成本归属'] = person_overrides['cost_bearer_name']
                st.dataframe(
                    person_overrides[[
                        'override_id', '项目', 'effective_from_month', 'effective_to_month',
                        '办理单位', '成本归属', 'special_reason'
                    ]].rename(columns={
                        'override_id': '编号', 'effective_from_month': '开始月',
                        'effective_to_month': '结束月', 'special_reason': '原因',
                    }),
                    use_container_width=True,
                    hide_index=True,
                )
                with st.form("end_person_override_form"):
                    override_ids = person_overrides['override_id'].astype(int).tolist()
                    override_labels = {
                        int(row['override_id']): f"{row['项目']}（{row['effective_from_month']}起）"
                        for _, row in person_overrides.iterrows()
                    }
                    end_override_id = st.selectbox(
                        "选择要结束的例外", override_ids,
                        format_func=lambda value: override_labels[value],
                    )
                    end_override_month = st.text_input("结束月份", value=target_month, max_chars=7)
                    if st.form_submit_button("保存结束月份"):
                        ok, msg = end_person_social_override(end_override_id, end_override_month.strip())
                        if ok:
                            set_msg_and_rerun(msg)
                        else:
                            st.error(msg)

        with st.expander("业务单位维护（新增地市或其他单位）", expanded=False):
            st.caption("地市和承接单位可以随时新增；停用只影响以后选择，不删除历史记录。")
            st.dataframe(
                entity_df[['entity_name', 'entity_type', 'parent_entity_name']].rename(columns={
                    'entity_name': '单位名称', 'entity_type': '单位类型',
                    'parent_entity_name': '上级单位',
                }),
                use_container_width=True,
                hide_index=True,
            )
            with st.form("simple_entity_form"):
                ec1, ec2 = st.columns(2)
                with ec1:
                    new_entity_name = st.text_input("新单位名称")
                with ec2:
                    new_entity_type = st.selectbox("单位类型", ["地市分公司", "其他承接单位"])
                if st.form_submit_button("新增单位"):
                    ok, msg = create_business_entity(new_entity_name, new_entity_type, 'province_company')
                    if ok:
                        set_msg_and_rerun(msg)
                    else:
                        st.error(msg)

    with default_rules_tab:
        st.subheader("下沉/地市转入默认待遇规则")
        st.info(
            "这里维护一类人员的统一默认值。以后新增或调整为该情形的人员会自动套用；"
            "已经生效的同类人员也会从所选月份开始采用新版本。个人特殊设置优先级最高，不会被覆盖。"
        )
        rule_month = st.text_input(
            "查看和设置月份", value=date.today().strftime('%Y-%m'),
            max_chars=7, key="special_default_rule_month",
        )
        rule_type = st.radio(
            "人员情形", ['down_secondment', 'city_transfer'], horizontal=True,
            format_func=lambda value: ARRANGEMENT_LABELS[value],
            key="special_default_rule_type",
        )
        if rule_type == 'down_secondment':
            st.caption(
                "省公司文件默认：养老、年金由省公司集中办理；失业、工伤由下沉地市办理；"
                "医疗、生育、公积金由省公众办理。全部费用由下沉地市承担。"
            )
        else:
            st.caption(
                "当前默认：养老、医疗、失业、工伤、生育由省公众办理；"
                "住房公积金、企业年金由原单位办理；全部计入本单位人工成本。"
            )

        default_rule_df = get_arrangement_route_defaults(rule_type, rule_month)
        if default_rule_df.empty:
            st.warning("当前月份没有可用的默认规则，请在下方建立第一条规则。")
        else:
            st.dataframe(
                default_rule_df[[
                    '项目', '默认办理', '办理单位', '成本归属',
                    '系统处理', '生效月份', '说明',
                ]],
                use_container_width=True,
                hide_index=True,
            )

        edit_rule_item = st.selectbox(
            "选择要调整的项目", PERSON_TREATMENT_ITEMS,
            format_func=lambda value: INSURANCE_LABELS[value],
            key=f"default_rule_item_{rule_type}",
        )
        current_rule_rows = (
            default_rule_df[default_rule_df['insurance_item'] == edit_rule_item]
            if not default_rule_df.empty else pd.DataFrame()
        )
        current_rule = current_rule_rows.iloc[0] if not current_rule_rows.empty else None
        payer_choices = ['province_public', 'province_company', 'related_branch']
        payer_labels = {
            'province_public': '省公众',
            'province_company': '省公司',
            'related_branch': '关联地市/原单位',
        }
        current_payer = (
            current_rule['payer_choice']
            if current_rule is not None and current_rule['payer_choice'] in payer_choices
            else ('province_public' if rule_type == 'city_transfer' else 'related_branch')
        )
        with st.form(f"special_default_rule_form_{rule_type}_{edit_rule_item}"):
            rc1, rc2, rc3 = st.columns(3)
            with rc1:
                rule_enabled = st.checkbox(
                    "该项目默认需要办理",
                    value=(
                        current_rule is None or current_rule['默认办理'] == '是'
                    ),
                )
            with rc2:
                rule_payer = st.selectbox(
                    "默认由谁办理", payer_choices,
                    format_func=lambda value: payer_labels[value],
                    index=payer_choices.index(current_payer),
                )
            with rc3:
                rule_company_cost = st.checkbox(
                    "计入本单位人工成本",
                    value=(
                        bool(current_rule['include_company_cost'])
                        if current_rule is not None else rule_type == 'city_transfer'
                    ),
                )
            rule_effective_month = st.text_input(
                "新规则生效月份", value=rule_month, max_chars=7
            )
            rule_remarks = st.text_input(
                "依据或说明（可选）",
                value=str(current_rule['说明']) if current_rule is not None else "",
            )
            if st.form_submit_button("保存这一项默认规则", type="primary"):
                ok, msg = save_arrangement_route_default(
                    rule_type, edit_rule_item, rule_enabled, rule_payer,
                    rule_company_cost, rule_effective_month.strip(), rule_remarks,
                )
                if ok:
                    set_msg_and_rerun(msg)
                else:
                    st.error(msg)

    with payroll_identity_tab:
        st.subheader("优才、技术精英与专家身份")
        st.info(
            "这里只给人员登记身份和有效期，不在人员档案里增加永久标签。"
            "工资生成时会按发薪月份自动判断是否有效；多重优才身份只采用绩效倍数最高的一项，"
            "不会叠加发放。"
        )

        identity_options = {
            "集团优才": ("talent", "group"),
            "省级优才": ("talent", "province"),
            "技术精英": ("technical_elite", "elite"),
            "首席技术精英": ("technical_elite", "chief"),
            "一级专家（历史调整口径）": ("province_expert", "level_1"),
            "二级专家（历史调整口径）": ("province_expert", "level_2"),
        }
        identity_labels = {value: label for label, value in identity_options.items()}

        selectable_employees = df_emps[
            df_emps["status"].isin(["在职", "实习", "挂靠人员"])
        ].copy()
        if selectable_employees.empty:
            st.warning("当前没有可登记薪酬身份的人员。")
        else:
            employee_ids = selectable_employees["emp_id"].astype(str).tolist()
            employee_labels = {
                str(row["emp_id"]): (
                    f"{row['name']}（{row.get('employee_no') or '工号待分配'}）"
                    f"｜{row.get('dept_name') or '未分配部门'}"
                )
                for _, row in selectable_employees.iterrows()
            }

            with st.form("payroll_identity_form"):
                identity_person = st.selectbox(
                    "人员*", employee_ids,
                    format_func=lambda value: employee_labels.get(value, value),
                )
                identity_name = st.selectbox("身份*", list(identity_options))
                idc1, idc2 = st.columns(2)
                with idc1:
                    identity_start = st.date_input("开始日期*", value=date.today())
                with idc2:
                    identity_has_end = st.checkbox("已明确结束日期")
                    identity_end = st.date_input(
                        "结束日期", value=date.today(), disabled=not identity_has_end,
                    )
                identity_document = st.text_input(
                    "依据文件（可选）", placeholder="例如：中电信鄂公众〔2024〕53号"
                )
                identity_remarks = st.text_input(
                    "说明（可选）", placeholder="仅记录这次聘任需要记住的特殊情况"
                )
                if st.form_submit_button("保存身份与聘期", type="primary"):
                    identity_type, identity_level = identity_options[identity_name]
                    ok, msg = save_payroll_identity(
                        identity_person, identity_type, identity_level,
                        identity_start.isoformat(),
                        identity_end.isoformat() if identity_has_end else None,
                        identity_document, identity_remarks,
                    )
                    if ok:
                        set_msg_and_rerun(msg)
                    else:
                        st.error(msg)

        identity_rows = get_payroll_identities()
        if not identity_rows:
            st.caption("尚未登记优才、技术精英或专家身份。")
        else:
            identity_df = pd.DataFrame(identity_rows)
            identity_df["身份"] = identity_df.apply(
                lambda row: identity_labels.get(
                    (row["identity_type"], row["identity_level"]),
                    f"{row['identity_type']} / {row['identity_level']}",
                ), axis=1,
            )
            identity_df["状态"] = identity_df["status"].map({
                "active": "有效/待生效", "ended": "已结束",
            }).fillna(identity_df["status"])
            st.write("### 已登记身份")
            st.dataframe(
                identity_df[[
                    "employee_no", "employee_name", "身份", "start_date", "end_date",
                    "状态", "source_document", "remarks",
                ]].rename(columns={
                    "employee_no": "工号", "employee_name": "姓名",
                    "start_date": "开始日期", "end_date": "结束日期",
                    "source_document": "依据文件", "remarks": "说明",
                }),
                use_container_width=True, hide_index=True,
            )

            active_identity_df = identity_df[identity_df["status"] == "active"]
            if not active_identity_df.empty:
                with st.expander("提前结束一项身份", expanded=False):
                    active_ids = active_identity_df["identity_id"].astype(int).tolist()
                    active_labels = {
                        int(row["identity_id"]): (
                            f"{row['employee_name']}｜{row['身份']}｜{row['start_date']}起"
                        )
                        for _, row in active_identity_df.iterrows()
                    }
                    with st.form("end_payroll_identity_form"):
                        ending_identity = st.selectbox(
                            "选择身份", active_ids,
                            format_func=lambda value: active_labels[value],
                        )
                        ending_date = st.date_input("实际结束日期", value=date.today())
                        if st.form_submit_button("保存结束日期"):
                            ok, msg = end_payroll_identity(
                                ending_identity, ending_date.isoformat()
                            )
                            if ok:
                                set_msg_and_rerun(msg)
                            else:
                                st.error(msg)

# 旧版复杂入口保留在代码中用于兼容，但不再出现在导航中。
elif current_page == "🔄 特殊用工与结算关系":
    st.subheader("🔄 特殊用工、实际工作与成本结算关系")
    st.info(
        "这里只维护挂靠代缴、地市工作转入和下沉人员，不维护普通员工。"
        "关系按生效日期留痕，用于决定是否由本系统发工资、各险种由谁缴费，以及费用最终由谁承担。"
    )
    st.caption(
        "魏巍属于“地市工作转入”；三名下沉人员属于“下沉人员”；"
        "李峰林如果只有工伤特殊，不在这里建关系，而是在社保页面建立个人险种例外。"
    )

    relation_df = get_arrangements_dataframe(include_closed=True)
    entity_df = get_entities_dataframe(active_only=True)
    entity_options = [None] + entity_df['entity_code'].tolist()
    entity_names = dict(zip(entity_df['entity_code'], entity_df['entity_name']))

    tab_list, tab_add, tab_close, tab_entities = st.tabs([
        "📋 关系台账", "➕ 新建特殊关系", "✅ 到期/结束关系", "🏢 业务单位管理"
    ])

    with tab_list:
        if relation_df.empty:
            st.caption("暂无显式关系；未配置人员继续按原有普通/挂靠逻辑兼容运行。")
        else:
            show_df = relation_df.copy()
            show_df['关系类型'] = show_df['arrangement_type'].map(ARRANGEMENT_LABELS).fillna(show_df['arrangement_type'])
            show_df['工资范围'] = show_df['payroll_included'].map({1: '本系统发薪', 0: '本系统不发薪'})
            show_df['settlement_mode'] = show_df['settlement_mode'].map(SETTLEMENT_MODE_LABELS).fillna(show_df['settlement_mode'])
            show_df['settlement_cycle'] = show_df['settlement_cycle'].map(SETTLEMENT_CYCLE_LABELS).fillna(show_df['settlement_cycle'])
            show_df['计划结束'] = pd.to_datetime(show_df['planned_end_date'], errors='coerce')

            today_ts = pd.Timestamp(date.today())
            expiring = show_df[
                show_df['status'].eq('active')
                & show_df['计划结束'].notna()
                & (show_df['计划结束'] >= today_ts)
                & (show_df['计划结束'] <= today_ts + pd.Timedelta(days=90))
            ]
            overdue = show_df[
                show_df['status'].eq('active')
                & show_df['计划结束'].notna()
                & (show_df['计划结束'] < today_ts)
            ]
            if not overdue.empty:
                st.error(f"有 {len(overdue)} 条关系已超过计划结束日期，必须确认延期、返回或转入。")
            if not expiring.empty:
                st.warning(f"有 {len(expiring)} 条关系将在90天内到期。")

            show_df['status'] = show_df['status'].map(ARRANGEMENT_STATUS_LABELS).fillna(show_df['status'])

            display_columns = {
                'arrangement_id': '关系ID', 'emp_name': '姓名', 'employee_no': '工号',
                '关系类型': '关系类型', 'contract_entity_name': '劳动合同主体',
                'payroll_entity_name': '工资主体', 'actual_work_unit_name': '实际工作单位',
                'related_branch_name': '关联地市', 'accounting_entity_name': '当前记账单位',
                'ultimate_cost_bearer_name': '最终成本承担',
                'start_date': '开始日期', 'planned_end_date': '计划结束',
                'actual_end_date': '实际结束', '工资范围': '工资范围',
                'settlement_mode': '结算方式', 'settlement_cycle': '结算周期',
                'status': '状态', 'remarks': '备注'
            }
            st.dataframe(
                show_df[[c for c in display_columns if c in show_df.columns]].rename(columns=display_columns),
                use_container_width=True,
                hide_index=True,
            )

    with tab_add:
        relation_type = st.selectbox(
            "关系类型",
            list(SPECIAL_ARRANGEMENT_TYPES),
            format_func=lambda value: SPECIAL_ARRANGEMENT_TYPES[value],
        )
        default_payroll = relation_type not in {'proxy_social', 'down_secondment'}
        default_mode = {
            'proxy_social': 'proxy_social',
            'city_transfer': 'annual_labor_cost_reallocation',
            'down_secondment': 'mixed_by_item',
        }[relation_type]
        default_cycle = {
            'proxy_social': 'quarterly',
            'city_transfer': 'annual', 'down_secondment': 'mixed'
        }[relation_type]

        if relation_type == 'proxy_social':
            active_people = df_emps[df_emps['status'] == '挂靠人员'].copy()
            employee_field_label = "挂靠人员*"
            st.caption("挂靠代缴只允许选择人员档案中状态为“挂靠人员”的人员。")
        else:
            active_people = df_emps[df_emps['status'] == '在职'].copy()
            employee_field_label = "人员*"
            st.caption("地市工作转入和下沉人员本来就是本单位在职人员，因此这里会显示本单位人员供选择。")
        emp_options = active_people['emp_id'].astype(str).tolist()
        emp_label = dict(zip(
            active_people['emp_id'].astype(str),
            active_people.apply(lambda row: f"{row['name']}（{row.get('employee_no') or '待分配'}）", axis=1)
        ))

        with st.form("arrangement_create_form"):
            target_emp_id = st.selectbox(
                employee_field_label, emp_options,
                format_func=lambda value: emp_label[value],
                index=0 if emp_options else None,
                placeholder="没有符合当前关系类型的候选人员",
            )
            c1, c2, c3 = st.columns(3)
            with c1:
                contract_entity = st.selectbox(
                    "劳动合同主体", entity_options,
                    format_func=lambda value: "未指定" if value is None else entity_names[value]
                )
                payroll_entity = st.selectbox(
                    "工资发放主体", entity_options,
                    format_func=lambda value: "未指定/外部发放" if value is None else entity_names[value]
                )
            with c2:
                actual_work_unit = st.selectbox(
                    "实际工作单位", entity_options,
                    format_func=lambda value: "未指定" if value is None else entity_names[value]
                )
                related_branch = st.selectbox(
                    "关联地市", entity_options,
                    format_func=lambda value: "未指定" if value is None else entity_names[value]
                )
            with c3:
                accounting_entity = st.selectbox(
                    "当前记账单位", entity_options,
                    format_func=lambda value: "未指定" if value is None else entity_names[value]
                )
                cost_bearer = st.selectbox(
                    "最终成本承担单位", entity_options,
                    format_func=lambda value: "未指定" if value is None else entity_names[value]
                )

            d1, d2 = st.columns(2)
            with d1:
                start_date_value = st.date_input("开始日期*", value=date.today())
                has_planned_end = st.checkbox("设置计划结束日期", value=relation_type == 'down_secondment')
            with d2:
                planned_default = date.today() + relativedelta(years=2)
                planned_end_value = st.date_input("计划结束日期", value=planned_default)
                payroll_included = st.checkbox("纳入本系统工资发放", value=default_payroll)

            s1, s2 = st.columns(2)
            with s1:
                settlement_mode_options = list(SETTLEMENT_MODE_LABELS)
                settlement_mode = st.selectbox(
                    "结算方式", settlement_mode_options,
                    index=settlement_mode_options.index(default_mode),
                    format_func=lambda value: SETTLEMENT_MODE_LABELS[value],
                )
                settlement_cycle = st.selectbox(
                    "结算周期", list(SETTLEMENT_CYCLE_LABELS),
                    index=list(SETTLEMENT_CYCLE_LABELS).index(default_cycle),
                    format_func=lambda value: SETTLEMENT_CYCLE_LABELS[value],
                )
            with s2:
                source_document_no = st.text_input("政策/协议文号")
                relation_remarks = st.text_area("备注")

            submitted = st.form_submit_button("保存关系", type="primary")
            if submitted:
                if target_emp_id is None:
                    st.error("当前没有可建立该类关系的候选人员")
                    st.stop()
                employee_row = active_people[active_people['emp_id'].astype(str) == str(target_emp_id)].iloc[0]
                ok, msg = create_arrangement({
                    'emp_id': str(target_emp_id),
                    'arrangement_type': relation_type,
                    'contract_entity_code': contract_entity,
                    'payroll_entity_code': payroll_entity,
                    'home_dept_id': int(employee_row['dept_id']) if pd.notna(employee_row['dept_id']) else None,
                    'actual_work_unit_code': actual_work_unit,
                    'related_branch_code': related_branch,
                    'accounting_entity_code': accounting_entity,
                    'ultimate_cost_bearer_code': cost_bearer,
                    'start_date': start_date_value.isoformat(),
                    'planned_end_date': planned_end_value.isoformat() if has_planned_end else None,
                    'actual_end_date': None,
                    'payroll_included': 1 if payroll_included else 0,
                    'settlement_mode': settlement_mode,
                    'settlement_cycle': settlement_cycle,
                    'status': 'active',
                    'source_document_no': source_document_no,
                    'remarks': relation_remarks,
                })
                if ok:
                    set_msg_and_rerun(msg)
                else:
                    st.error(msg)

    with tab_close:
        active_relations = relation_df[relation_df['status'] == 'active'] if not relation_df.empty else pd.DataFrame()
        if active_relations.empty:
            st.caption("没有可结束的活动关系。")
        else:
            relation_options = active_relations['arrangement_id'].astype(int).tolist()
            relation_labels = {
                int(row['arrangement_id']): (
                    f"#{int(row['arrangement_id'])} {row['emp_name']} - "
                    f"{ARRANGEMENT_LABELS.get(row['arrangement_type'], row['arrangement_type'])}"
                )
                for _, row in active_relations.iterrows()
            }
            with st.form("arrangement_close_form"):
                close_id = st.selectbox("选择关系", relation_options, format_func=lambda value: relation_labels[value])
                close_date_value = st.date_input("实际结束日期", value=date.today())
                close_status = st.selectbox(
                    "结束结果", list(ARRANGEMENT_CLOSE_RESULT_LABELS),
                    format_func=lambda value: ARRANGEMENT_CLOSE_RESULT_LABELS[value],
                )
                close_remarks = st.text_area("处理说明")
                if st.form_submit_button("确认结束关系", type="primary"):
                    ok, msg = close_arrangement(close_id, close_date_value, close_status, close_remarks)
                    if ok:
                        set_msg_and_rerun(msg)
                    else:
                        st.error(msg)

    with tab_entities:
        st.caption(
            "这里维护所有可能参与劳动合同、工资发放、实际工作、社保缴费和费用结算的单位。"
            "新增地市或其他承接单位后，会自动出现在关系和缴费路由的单位下拉框中。"
        )
        all_entities = get_entities_dataframe(active_only=False)
        if not all_entities.empty:
            entity_display = all_entities.copy()
            entity_display['active'] = entity_display['active'].map(ACTIVE_LABELS).fillna(entity_display['active'])
            st.dataframe(
                entity_display[[
                    'entity_name', 'entity_type', 'parent_entity_name', 'active'
                ]].rename(columns={
                    'entity_name': '单位名称', 'entity_type': '单位类型',
                    'parent_entity_name': '上级单位', 'active': '状态',
                }),
                use_container_width=True,
                hide_index=True,
            )

        with st.expander("➕ 新增业务单位", expanded=False):
            with st.form("business_entity_create_form"):
                ec1, ec2, ec3 = st.columns(3)
                with ec1:
                    new_entity_name = st.text_input("单位名称*")
                with ec2:
                    new_entity_type = st.selectbox(
                        "单位类型*", ["地市分公司", "其他承接单位", "法人", "上级单位"]
                    )
                with ec3:
                    parent_options = [None] + entity_df['entity_code'].tolist()
                    new_parent_entity = st.selectbox(
                        "上级单位（可选）", parent_options,
                        format_func=lambda value: "不设置" if value is None else entity_names[value],
                    )
                if st.form_submit_button("保存业务单位", type="primary"):
                    ok, msg = create_business_entity(
                        new_entity_name, new_entity_type, new_parent_entity
                    )
                    if ok:
                        set_msg_and_rerun(msg)
                    else:
                        st.error(msg)

        if not all_entities.empty:
            with st.expander("启用或停用业务单位", expanded=False):
                all_entity_options = all_entities['entity_code'].tolist()
                all_entity_names = dict(zip(all_entities['entity_code'], all_entities['entity_name']))
                with st.form("business_entity_status_form"):
                    target_entity_code = st.selectbox(
                        "选择单位", all_entity_options,
                        format_func=lambda value: all_entity_names[value],
                    )
                    entity_action = st.radio("操作", ["启用", "停用"], horizontal=True)
                    if st.form_submit_button("确认操作"):
                        ok, msg = set_business_entity_active(
                            target_entity_code, entity_action == "启用"
                        )
                        if ok:
                            set_msg_and_rerun(msg)
                        else:
                            st.error(msg)

# ==============================================================================
# 模块 E: 🕰️ 历史变动流水 (全量归位侧边栏审计)
# ==============================================================================
elif current_page == "🕰️ 历史变动流水":
    st.subheader("🕰️ 全生命周期时空审计")
    ok, h_list = get_all_history()
    if ok and h_list:
        hdf = pd.DataFrame(h_list); hdf['dt_obj'] = pd.to_datetime(hdf['change_date'])
        hdf = hdf.sort_values(by='dt_obj', ascending=False)

        with st.sidebar.expander("🔍 高级筛选与审计导出", expanded=True):
            hs = st.text_input("搜姓名/工号")
            ht = st.multiselect("变动类型", options=sorted(hdf['change_type'].unique().tolist()))
            dept_opts = sorted([str(d) for d in hdf['old_dept_name'].unique().tolist() if d])
            hd = st.multiselect("原归属部门", options=dept_opts)
            d_range = st.date_input("生效时段", value=(date.today() - relativedelta(months=3), date.today()))

        f_h = hdf.copy()
        if hs: f_h = f_h[f_h['emp_name'].str.contains(hs, na=False) | f_h['employee_no'].fillna('').astype(str).str.contains(hs, na=False)]
        if ht: f_h = f_h[f_h['change_type'].isin(ht)]
        if hd: f_h = f_h[f_h['old_dept_name'].isin(hd)]
        if len(d_range) == 2:
            f_h = f_h.dropna(subset=['dt_obj'])
            start_dt = pd.to_datetime(d_range[0])
            end_dt = pd.to_datetime(d_range[1]) + pd.Timedelta(days=1, seconds=-1)
            f_h = f_h[(f_h['dt_obj'] >= start_dt) & (f_h['dt_obj'] <= end_dt)]

        with st.sidebar:
            if not f_h.empty:
                audit_cols = {'change_date': '生效日期', 'emp_name': '姓名', 'employee_no': '工号', 'change_type': '变动类型', 'old_dept_name': '原部门', 'new_dept_name': '新部门', 'old_pos_name': '原岗位', 'new_pos_name': '新岗位', 'old_tech_grade': '原T级', 'new_tech_grade': '新T级', 'old_post_rank': '原岗级', 'new_post_rank': '新岗级', 'old_post_grade': '原档次', 'new_post_grade': '新档次', 'change_reason': '说明'}
                export_final = f_h[list(audit_cols.keys())].rename(columns=audit_cols)
                ob = io.BytesIO()
                with pd.ExcelWriter(ob, engine='openpyxl') as w: export_final.to_excel(w, index=False)
                st.download_button("📥 导出筛选流水", data=ob.getvalue(), file_name=f"变动审计_{date.today()}.xlsx", type="primary", use_container_width=True)

        if not f_h.empty:
            for _, row in f_h.iterrows():
                with st.container(border=True):
                    h1, h2 = st.columns([1, 5])
                    with h1:
                        st.write(f"**{row['emp_name']}**")
                        st.caption(f"工号: {row.get('employee_no') or '待分配'}")
                        type_str = row['change_type']
                        if '离职' in type_str or '退休' in type_str or '变为' in type_str: st.error(f"🛑 {type_str}")
                        elif '实习转正' in type_str: st.warning(f"🔥 {type_str}")
                        else: st.info(type_str)
                    with h2:
                        c1, c2, c3, c4 = st.columns(4)
                        with c1: st.caption("🏢 部门变动"); st.write(f"{row['old_dept_name'] or '-'}\n➡️\n{row['new_dept_name'] or '-'}")
                        with c2: st.caption("💼 岗位变动"); st.write(f"{row['old_pos_name'] or '-'}\n➡️\n{row['new_pos_name'] or '-'}")
                        with c3: st.caption("🏅 T级变动"); st.write(f"{row['old_tech_grade'] or '-'}\n➡️\n{row['new_tech_grade'] or '-'}")
                        with c4: st.caption("💰 待遇对比"); st.write(f"{row['old_post_rank'] or '-'}岗{row['old_post_grade'] or '-'}\n➡️\n{row['new_post_rank'] or '-'}岗{row['new_post_grade'] or '-'}")
                        st.divider()
                        st.caption(f"🕰️ 生效: {row['change_date']} | 说明: {row['change_reason']}")
                        if st.button("⏪ 撤销", key=f"rb_{row['change_id']}"):
                            ro, rm = rollback_history(row['change_id'])
                            if ro: set_msg_and_rerun(rm)
                            else: st.error(rm)
        else: st.warning("暂无匹配流水")
    else: st.info("尚未产生记录")
