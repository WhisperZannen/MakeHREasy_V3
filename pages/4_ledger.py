# ==============================================================================
# 文件路径: pages/4_ledger.py
# 功能描述: 人工成本台账管理中心 (财务合规与领导审阅终极版)
# 核心修正说明:
#   1. 彻底解决社保模块与薪酬模块的“边界隔离”冲突。
#   2. [核心缝合] 对公(社保)保持物理分离，对私(台账)保持物理合并。
#   3. [防覆盖拦截] 修复导入时“股票增值权/纯实发统筹款”被公式强制归零的 Bug。
#   4. [UI防跳跃] 剥离多余的 st.rerun()，彻底解决点击导入后跳回 Tab1 的反人类体验。
# ==============================================================================

import streamlit as st
import pandas as pd
import sqlite3
import os
import io
import uuid

# 用于 Excel 报表精装修
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 导入业务分离后的核心层模块
from modules.core_labor_cost import (
    LEDGER_MAP, DB_TO_CN_MAP, NUMERIC_COLS,
    _get_db_connection, cleanse_db_timestamps,
    sort_flat_ledger_df, add_subtotals_and_totals, get_ledger_data,
    localize_labor_cost_codes, read_labor_ledger_workbook,
    read_finance_account_workbook, prepare_finance_labor_precheck,
    recalculate_labor_cost_columns, get_company_social_snapshot,
)
from modules.core_arrangements import get_effective_arrangement, is_labor_cost_included
from modules.core_personnel import get_effective_department_snapshot
from modules.core_identity import (
    get_employee_no,
    resolve_employee_reference,
)

st.set_page_config(page_title="人工成本台账", layout="wide")

cleanse_db_timestamps()

# ==============================================================================
# Excel 财务级排版渲染引擎
# ==============================================================================
def format_excel_sheet(worksheet, df_columns):
    worksheet.freeze_panes = 'A2'
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    subtotal_fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
    total_fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
    key_col_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    key_columns = ['工资应发合计', '其他人工成本合计', '人工成本合计']

    df_cols_list = list(df_columns)
    name_col_idx = df_cols_list.index('姓名') + 1 if '姓名' in df_cols_list else -1

    for row_idx in range(1, worksheet.max_row + 1):
        is_subtotal = False
        is_total = False
        if name_col_idx != -1 and row_idx > 1:
            cell_val = str(worksheet.cell(row=row_idx, column=name_col_idx).value or "")
            if "【小计】" == cell_val: is_subtotal = True
            elif "【实际成本总计】" == cell_val: is_total = True

        for col_idx, col_name in enumerate(df_columns, 1):
            col_letter = get_column_letter(col_idx)
            cell = worksheet[f"{col_letter}{row_idx}"]
            cell.border = thin_border

            if row_idx == 1:
                worksheet.column_dimensions[col_letter].width = 8 if col_name == '序号' else (15 if col_name in NUMERIC_COLS else 12)
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                if col_name in NUMERIC_COLS:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                if is_total:
                    cell.fill = total_fill
                    cell.font = Font(bold=True, color="000000")
                elif is_subtotal:
                    cell.fill = subtotal_fill
                    cell.font = Font(bold=True, color="000000")
                elif col_name in key_columns:
                    cell.fill = key_col_fill
                    cell.font = Font(bold=True)

# ==============================================================================
# 人工成本台账部门归属刷新工具
# ==============================================================================
def build_effective_dept_snapshot(conn, target_month):
    """
    根据人员模块和人员变动流水，计算某个月每个人“应当归属”的部门。

    这个函数是干什么的？
    ------------------------------------------------------------
    它不是算钱的，也不会改人工成本金额。
    它只负责回答一个问题：

    某个人在 target_month 这个月，人工成本应该归属哪个部门？

    核心业务规则：
    ------------------------------------------------------------
    以每月 15 日为切片点。

    1. 如果员工在当月 15 日及以前调动：
       当月人工成本归新部门。

       例如：
       A 员工 2026-05-05 从研一调到研二。
       因为 5月5日 <= 5月15日，
       所以 2026-05 的人工成本归研二。

    2. 如果员工在当月 15 日之后调动：
       当月人工成本仍归原部门。

       例如：
       B 员工 2026-05-20 从研一调到研二。
       因为 5月20日 > 5月15日，
       所以 2026-05 的人工成本仍归研一。

    为什么要从“当前人员表”往回推？
    ------------------------------------------------------------
    因为 employees 表保存的是员工当前最新部门。
    如果我们要算历史月份，就需要根据 personnel_changes 里的调动记录，
    把目标月 15 日之后发生的调动“倒回去”。

    参数说明：
    ------------------------------------------------------------
    conn:
        SQLite 数据库连接。

    target_month:
        目标月份，格式必须是 YYYY-MM，例如 2026-05。

    返回值：
    ------------------------------------------------------------
    返回一个字典：
    {
        "工号": "应归属部门名称"
    }
    """

    snapshot = get_effective_department_snapshot(target_month, conn)
    return {emp_id: values['dept_name'] for emp_id, values in snapshot.items()}


def resolve_hr_director_tail_carrier(conn, ledger_df):
    """从组织和岗位档案识别人力资源部主任，并确认其存在于本月底表。"""
    director_df = pd.read_sql_query(
        '''
        SELECT e.emp_id, e.employee_no, e.name
        FROM employees e
        JOIN departments d ON e.dept_id = d.dept_id
        JOIN employee_profiles ep ON e.emp_id = ep.emp_id
        JOIN positions p ON ep.pos_id = p.pos_id
        WHERE e.status = '在职'
          AND trim(d.dept_name) = '人力资源部'
          AND trim(p.pos_name) = '主任'
        ORDER BY e.emp_id
        ''',
        conn,
    )
    if len(director_df) != 1:
        raise ValueError(
            f'系统应识别到1名在职人力资源部主任，当前识别到{len(director_df)}名。'
        )

    director_id = str(director_df.iloc[0]['employee_no'] or '').strip()
    director_name = str(director_df.iloc[0]['name']).strip()
    ledger_ids = (
        ledger_df['工号'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
    )
    if not ledger_ids.eq(director_id).any():
        raise ValueError(
            f'本月底表中没有人力资源部主任{director_name}（{director_id}），不能自动承接经费尾差。'
        )
    return director_id, director_name


def upsert_labor_cost_dataframe(in_df):
    """把已经通过预核对的中文人员台账写入数据库。"""
    normalized_df = recalculate_labor_cost_columns(in_df)
    conn = _get_db_connection()
    cursor = conn.cursor()
    success_count = 0
    department_snapshots = {}
    try:
        for _, row in normalized_df.iterrows():
            employee_name = str(row.get('姓名', '')).strip()
            department_name = str(row.get('归属部门', '')).strip()
            if employee_name in ['【小计】', '【实际成本总计】']:
                continue
            if department_name in ['【在职及统筹部分】']:
                continue

            raw_month = str(row.get('核算月份', '')).strip()
            raw_id = row.get('工号', '')
            if pd.isna(raw_id):
                employee_id = ''
            elif isinstance(raw_id, float):
                employee_id = str(int(raw_id))
            else:
                employee_id = str(raw_id).replace('.0', '').strip()
            if not raw_month or raw_month == 'nan':
                continue
            internal_emp_id = resolve_employee_reference(
                employee_no=employee_id,
                id_card=row.get('身份证号'),
                name=employee_name,
                conn=conn,
            )
            if not internal_emp_id:
                raise ValueError(
                    f'无法识别人员：{employee_name or "未填写姓名"}'
                    f'（工号：{employee_id or "待分配"}）'
                )
            cost_month = raw_month[:7].replace('/', '-') if len(raw_month) >= 7 else raw_month

            arrangement = get_effective_arrangement(internal_emp_id, cost_month, conn)
            if not int(arrangement.get('labor_cost_included', 1)):
                # 下沉和挂靠的社保仍留在社保账及结算账，但绝不进入本单位人工成本。
                cursor.execute(
                    "DELETE FROM labor_cost_ledger WHERE cost_month = ? AND emp_id = ?",
                    (cost_month, internal_emp_id),
                )
                continue

            db_data = {}
            for cn_column, db_column in LEDGER_MAP.items():
                if db_column == 'cost_month':
                    db_data[db_column] = cost_month
                    continue
                value = row.get(cn_column, None)
                if cn_column in NUMERIC_COLS:
                    try:
                        clean_value = str(value).replace(',', '').strip()
                        db_data[db_column] = (
                            float(clean_value)
                            if pd.notna(value) and clean_value != ''
                            else 0.0
                        )
                    except (TypeError, ValueError):
                        db_data[db_column] = 0.0
                else:
                    db_data[db_column] = str(value).strip() if pd.notna(value) else ''

            # 人员关系和部门归属以系统有效期为准，不相信Excel里可能过期的技术快照。
            db_data['emp_id'] = internal_emp_id
            relation_type = arrangement.get('arrangement_type', 'normal')
            db_data['arrangement_id'] = arrangement.get('arrangement_id')
            db_data['business_type_snapshot'] = relation_type
            db_data['labor_cost_included_snapshot'] = 1
            db_data['actual_work_unit_code'] = arrangement.get('actual_work_unit_code')
            db_data['accounting_entity_code'] = arrangement.get('accounting_entity_code') or 'province_public'
            db_data['ultimate_cost_bearer_code'] = arrangement.get('ultimate_cost_bearer_code') or 'province_public'
            if cost_month not in department_snapshots:
                department_snapshots[cost_month] = get_effective_department_snapshot(cost_month, conn)
            department = department_snapshots[cost_month].get(internal_emp_id)
            if department:
                db_data['dept_id'] = department['dept_id']
                db_data['dept_name'] = department['dept_name']
            if relation_type == 'city_transfer':
                db_data['reallocation_mode'] = 'annual_labor_cost_reallocation'
                db_data['reallocation_status'] = 'pending'
            else:
                db_data['reallocation_mode'] = 'none'
                db_data['reallocation_status'] = 'not_required'

            columns = list(db_data.keys())
            placeholders = ','.join(['?'] * len(columns))
            updates = ','.join(
                f'{column}=excluded.{column}'
                for column in columns
                if column not in ['cost_month', 'emp_id']
            )
            cursor.execute(
                f'''
                    INSERT INTO labor_cost_ledger ({','.join(columns)})
                    VALUES ({placeholders})
                    ON CONFLICT(cost_month, emp_id) DO UPDATE SET {updates}
                ''',
                tuple(db_data.values()),
            )
            success_count += 1
        conn.commit()
        return success_count
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def build_finance_precheck_workbook(precheck_result):
    """生成可继续导入的人员台账，并附带自动处理与双口径核对证据。"""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        processed = precheck_result['processed_ledger']
        processed.to_excel(writer, index=False, sheet_name='待导入台账')
        if not precheck_result['monthly_reconciliation'].empty:
            precheck_result['monthly_reconciliation'].to_excel(
                writer, index=False, sheet_name='当月核对'
            )
        if not precheck_result['ytd_reconciliation'].empty:
            precheck_result['ytd_reconciliation'].to_excel(
                writer, index=False, sheet_name='累计核对'
            )
        precheck_result['auto_actions'].to_excel(
            writer, index=False, sheet_name='自动处理说明'
        )
        precheck_result['business_checks'].to_excel(
            writer, index=False, sheet_name='业务公式核对'
        )
        precheck_result['pending_accounts'].to_excel(
            writer, index=False, sheet_name='待确认科目'
        )

        format_excel_sheet(writer.sheets['待导入台账'], processed.columns)
        for sheet_name, worksheet in writer.sheets.items():
            if sheet_name == '待导入台账':
                continue
            worksheet.freeze_panes = 'A2'
            for column_index in range(1, worksheet.max_column + 1):
                worksheet.column_dimensions[get_column_letter(column_index)].width = 22
    return output.getvalue()


def save_finance_precheck_audit(
    cost_month, file_names, precheck_result, imported_records
):
    """保存财务源表核对结果；不保存源文件内容，避免数据库膨胀。"""
    batch_id = f"LC-{cost_month}-{uuid.uuid4().hex[:12]}"
    conn = _get_db_connection()
    try:
        conn.execute(
            '''
            INSERT INTO finance_labor_import_batches(
                batch_id, cost_month, ledger_file_name, monthly_file_name,
                ytd_file_name, status, imported_records, confirmed_at
            ) VALUES (?, ?, ?, ?, ?, 'imported', ?, CURRENT_TIMESTAMP)
            ''',
            (
                batch_id,
                cost_month,
                file_names.get('ledger'),
                file_names.get('monthly'),
                file_names.get('ytd'),
                imported_records,
            ),
        )
        reconciliation = pd.concat(
            [
                precheck_result['monthly_reconciliation'],
                precheck_result['ytd_reconciliation'],
            ],
            ignore_index=True,
        )
        for _, row in reconciliation.iterrows():
            conn.execute(
                '''
                INSERT INTO finance_labor_reconciliation(
                    batch_id, reconciliation_scope, control_item, account_codes,
                    finance_amount, ledger_amount, difference_amount,
                    processing_mode, reconciliation_status, remarks
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''',
                (
                    batch_id,
                    row.get('核对范围', ''),
                    row.get('核对项目', ''),
                    row.get('财务科目', ''),
                    float(row.get('财务金额', 0.0)),
                    float(row.get('台账金额', 0.0)),
                    float(row.get('差额（台账-财务）', 0.0)),
                    row.get('处理方式', ''),
                    row.get('核对状态', ''),
                    row.get('说明', ''),
                ),
            )
        conn.commit()
        return batch_id
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

# ==============================================================================
# 页面主框架
# ==============================================================================
st.title("💰 人工成本台账管理中心")
st.caption("🔒 财务数据合规要求：台账一旦生成不可在系统内手动篡改。如需修正，请在 Excel 中修改后重新导入，系统将自动覆盖重置原账目。")
st.info(
    "📌 女工劳保费特殊口径："
    "女工劳保费继续单独记录，并计入员工实际发放；"
    "但不计入人工成本模块的【工资应发合计】、"
    "【其他人工成本合计】和【人工成本合计】。"
)


tab1, tab2, tab3 = st.tabs(["📊 台账多维看板", "📤 领导审阅导出 (范围框选)", "📥 财务底表导入"])

# ------------------------------------------------------------------------------
# Tab 1: 台账多维看板
# ------------------------------------------------------------------------------
with tab1:
    conn = _get_db_connection()
    available_months = pd.read_sql_query("SELECT DISTINCT cost_month FROM labor_cost_ledger ORDER BY cost_month DESC", conn)['cost_month'].tolist()
    available_depts = pd.read_sql_query("SELECT DISTINCT dept_name FROM labor_cost_ledger", conn)['dept_name'].tolist()
    conn.close()

    sc1, sc2, sc3 = st.columns([1, 1, 1])
    with sc1: f_month = st.selectbox("📅 核算月份筛选", ["全部月份"] + available_months)
    with sc2: f_dept = st.multiselect("🏢 归属部门筛选", options=available_depts)
    with sc3: q_search = st.text_input("🔍 搜姓名 / 工号")

    raw_df = get_ledger_data(month_filter=None if f_month == "全部月份" else f_month, dept_filter=f_dept if f_dept else None)

    if not raw_df.empty:
        if q_search:
            raw_df = raw_df[raw_df['emp_name'].str.contains(q_search, na=False) | raw_df['emp_id'].str.contains(q_search, na=False)]

    if not raw_df.empty:
        active_metric_df = raw_df[~( (raw_df['emp_status'] == '退休') | (raw_df['dept_name'].str.contains('离退休', na=False)) )]

        # 人工成本口径总成本。
        # 女工劳保费不计入。
        total_cost = active_metric_df['total_labor_cost'].sum()

        # 人工成本口径工资应发。
        # 女工劳保费不计入。
        total_gross = active_metric_df['gross_salary_total'].sum()

        # 其他人工成本。
        # 女工劳保费也不计入。
        total_other = active_metric_df['other_cost_total'].sum()

        # 女工劳保费继续单独记录和展示，
        # 但不进入上面三个省公司人工成本指标。
        women_labor_total = active_metric_df['allowance_women'].sum()

        total_headcount = len(active_metric_df)

        m1, m2, m3, m4, m5 = st.columns(5)

        m1.metric(
            "在职及统筹总成本 (元)",
            f"{total_cost:,.2f}"
        )

        m2.metric(
            "工资应发合计 (元)",
            f"{total_gross:,.2f}"
        )

        m3.metric(
            "其他人工成本合计 (元)",
            f"{total_other:,.2f}"
        )

        m4.metric(
            "女工劳保费·仅记录 (元)",
            f"{women_labor_total:,.2f}",
            help="随工资发放，但不计入省公司人工成本口径。"
        )

        m5.metric(
            "在职及统筹核算人次",
            f"{total_headcount} 人次"
        )

        if 'reallocation_mode' in active_metric_df.columns:
            reallocation_df = active_metric_df[
                active_metric_df['reallocation_mode'].fillna('none') != 'none'
            ]
            if not reallocation_df.empty:
                st.warning(
                    f"🔁 当前范围有 {len(reallocation_df)} 人次需要地市结算或人工成本划转，"
                    f"涉及台账人工成本 {reallocation_df['total_labor_cost'].sum():,.2f} 元。"
                )

        disp_df = localize_labor_cost_codes(raw_df.rename(columns=DB_TO_CN_MAP))
        disp_cols = [col for col in LEDGER_MAP.keys() if col in disp_df.columns]

        disp_final = sort_flat_ledger_df(disp_df[disp_cols].copy())
        disp_final.insert(0, '序号', range(1, len(disp_final) + 1))
        st.dataframe(disp_final, use_container_width=True, hide_index=True)
    else:
        st.info("💡 当前筛选条件下暂无台账数据。")

# ------------------------------------------------------------------------------
# Tab 2: 领导审阅导出 (范围框选与智能汇集)
# ------------------------------------------------------------------------------
with tab2:
    st.subheader("📤 生成向领导汇报的标准台账")
    st.info("💡 操作提示：请选择导出的【起始】和【结束】月份，系统会自动提取该时间段内的所有明细，并可生成累计总账。")

    tc1, tc2 = st.columns(2)
    with tc1:
        start_month = st.selectbox("📅 导出起始月份", options=available_months if available_months else ["无数据"], index=len(available_months)-1 if available_months else 0)
    with tc2:
        end_month = st.selectbox("📅 导出结束月份", options=available_months if available_months else ["无数据"], index=0 if available_months else 0)

    need_summary = st.checkbox("📊 同时生成【选中范围的累计汇总】Sheet (勾选后，系统会自动把这几个月加起来算个总账)", value=True)

    if st.button("🚀 一键生成并下载 Excel 报表", type="primary"):
        if start_month == "无数据" or end_month == "无数据":
            st.warning("⚠️ 暂无可导出的数据！")
        else:
            s_m, e_m = min(start_month, end_month), max(start_month, end_month)
            selected_months = [m for m in available_months if s_m <= m <= e_m]
            selected_months.sort()

            conn = _get_db_connection()
            placeholders = ",".join(["?"] * len(selected_months))
            query = f"SELECT * FROM labor_cost_ledger WHERE cost_month IN ({placeholders}) ORDER BY dept_name ASC"
            raw_export_df = pd.read_sql_query(query, conn, params=selected_months)
            employee_no_map = dict(conn.execute(
                "SELECT emp_id, COALESCE(employee_no, '待分配') FROM employees"
            ).fetchall())
            conn.close()

            if not raw_export_df.empty:
                ob = io.BytesIO()
                with pd.ExcelWriter(ob, engine='openpyxl') as writer:

                    # ==============================================================================
                    # 修正后的 Tab 2 汇总逻辑：捍卫数据库字段 + 智能增强备注
                    # ==============================================================================
                    if need_summary:
                        # 1. 基础配置：依然保留 emp_status 字段
                        db_num_cols = [LEDGER_MAP[c] for c in NUMERIC_COLS if c in LEDGER_MAP]
                        agg_dict = {col: 'sum' for col in db_num_cols}
                        agg_dict.update({'emp_name': 'first', 'emp_status': 'last'})

                        # 2. 执行双键联合分组（捍卫部门归属）
                        summary_df = raw_export_df.groupby(['emp_id', 'dept_name']).agg(agg_dict).reset_index()

                        # 3. 【核心加固】实时探针：抓取此时此刻的档案信息
                        conn_live = _get_db_connection()
                        live_info = pd.read_sql_query("""
                                                      SELECT e.emp_id, e.status as live_st, d.dept_name as live_dept
                                                      FROM employees e
                                                               LEFT JOIN departments d ON e.dept_id = d.dept_id
                                                      """, conn_live).set_index('emp_id').to_dict('index')
                        conn_live.close()


                        # 4. 【极简状态引擎】状态列绝对纯净，变化信息全部压入备注！
                        def process_display_info(row):
                            eid = row['emp_id']
                            hist_dept = row['dept_name']
                            hist_status = row['emp_status']
                            current = live_info.get(eid, {})

                            live_st = current.get('live_st', '未知')
                            live_dept = current.get('live_dept', '')

                            # 状态列原封不动，当年怎样就怎样！
                            final_status = hist_status
                            remarks = []

                            # A. 状态变化提醒 (比如以前在职，现在离职了，才备注)
                            if live_st != hist_status:
                                remarks.append(f"现已{live_st}")

                            # B. 跨部门提醒 (如果现在还在职，且部门变了，才备注)
                            if live_st not in ['离职', '退休'] and live_dept and live_dept != hist_dept:
                                remarks.append(f"现调至[{live_dept}]")

                            return final_status, " | ".join(remarks)


                        # 5. 应用转换：同时更新状态列和备注列
                        status_and_remarks = summary_df.apply(process_display_info, axis=1)
                        summary_df['emp_status'] = [x[0] for x in status_and_remarks]
                        summary_df['备注'] = [x[1] for x in status_and_remarks]

                        # 6. 翻译回中文表头并排序
                        summary_cn = summary_df.rename(columns=DB_TO_CN_MAP)
                        summary_cn['工号'] = summary_cn['工号'].map(employee_no_map).fillna('待分配')
                        report_cols = [c for c in LEDGER_MAP.keys() if c in summary_cn.columns and c != '核算月份']
                        if '备注' not in report_cols: report_cols.append('备注')

                        summary_cn = summary_cn[report_cols]
                        summary_final = add_subtotals_and_totals(summary_cn, NUMERIC_COLS)

                        sum_sheet_name = f"{len(selected_months)}个月累计汇总"
                        summary_final.to_excel(writer, index=False, sheet_name=sum_sheet_name)
                        format_excel_sheet(writer.sheets[sum_sheet_name], summary_final.columns)

                    for month in sorted(selected_months):
                        month_df = raw_export_df[raw_export_df['cost_month'] == month].copy()
                        if not month_df.empty:
                            month_cn = localize_labor_cost_codes(month_df.rename(columns=DB_TO_CN_MAP))
                            month_cn['工号'] = month_cn['工号'].map(employee_no_map).fillna('待分配')
                            month_cols = [c for c in LEDGER_MAP.keys() if c in month_cn.columns]
                            month_cn = month_cn[month_cols]

                            month_final = add_subtotals_and_totals(month_cn, NUMERIC_COLS)
                            safe_month = str(month).replace(':', '-').replace('/', '-').replace('\\', '-').replace('?', '').replace('*', '')
                            safe_sheet_name = f"{safe_month[:28]}明细"

                            month_final.to_excel(writer, index=False, sheet_name=safe_sheet_name)
                            format_excel_sheet(writer.sheets[safe_sheet_name], month_final.columns)

                file_name = f"人工成本台账汇报_{len(selected_months)}个月数据.xlsx"
                st.download_button("📥 点击下载财务报表", data=ob.getvalue(), file_name=file_name, type="secondary")
            else:
                st.warning("所选范围内无数据。")

    st.write("### 🔁 地市人工成本划转专项表")
    st.caption(
        "用于“正式转入但仍在地市工作”等人员的年度全口径人工成本上报；"
        "只提取已标记需要划转的台账，不混入普通本级人员。"
    )

    reallocation_months = sorted(available_months)
    reallocation_start = None
    reallocation_end = None
    if reallocation_months:
        period_mode = st.radio(
            "专项表统计时间",
            options=["单月", "自定义区间", "年度累计"],
            horizontal=True,
            key="reallocation_period_mode",
        )
        if period_mode == "单月":
            selected_month = st.selectbox(
                "选择核算月份",
                options=reallocation_months,
                index=len(reallocation_months) - 1,
                key="reallocation_single_month",
            )
            reallocation_start = reallocation_end = selected_month
        elif period_mode == "自定义区间":
            rc1, rc2 = st.columns(2)
            with rc1:
                reallocation_start = st.selectbox(
                    "划转表起始月份",
                    options=reallocation_months,
                    index=len(reallocation_months) - 1,
                    key="reallocation_start_month",
                )
            with rc2:
                reallocation_end = st.selectbox(
                    "划转表结束月份",
                    options=reallocation_months,
                    index=len(reallocation_months) - 1,
                    key="reallocation_end_month",
                )
        else:
            available_years = sorted({month[:4] for month in reallocation_months}, reverse=True)
            selected_year = st.selectbox(
                "选择核算年度",
                options=available_years,
                key="reallocation_year",
            )
            year_months = [month for month in reallocation_months if month.startswith(selected_year)]
            reallocation_start, reallocation_end = min(year_months), max(year_months)
            st.caption(f"将导出 {reallocation_start} 至 {reallocation_end} 的现有台账数据。")
    else:
        st.info("暂无可供选择的人工成本月份。")

    if st.button("生成所选期间地市人工成本划转表", type="secondary"):
        if not reallocation_start or not reallocation_end:
            st.warning("暂无可导出的数据！")
            st.stop()
        s_m, e_m = min(reallocation_start, reallocation_end), max(reallocation_start, reallocation_end)
        conn_reallocation = _get_db_connection()
        reallocation_export = pd.read_sql_query(
            """
            SELECT l.cost_month AS 核算月份,
                   COALESCE(e.employee_no, '待分配') AS 工号,
                   l.emp_name AS 姓名,
                   l.dept_name AS 归属部门,
                   l.business_type_snapshot AS 业务关系类型,
                   COALESCE(work.entity_name, l.actual_work_unit_code, '') AS 实际工作单位,
                   COALESCE(bearer.entity_name, l.ultimate_cost_bearer_code, '') AS 最终成本承担单位,
                   l.gross_salary_total AS 工资应发合计,
                   l.other_cost_total AS 其他人工成本合计,
                   l.total_labor_cost AS 人工成本合计,
                   l.reallocation_mode AS 划转方式,
                   l.reallocation_status AS 划转状态
            FROM labor_cost_ledger l
            LEFT JOIN employees e ON l.emp_id = e.emp_id
            LEFT JOIN business_entities work ON l.actual_work_unit_code = work.entity_code
            LEFT JOIN business_entities bearer ON l.ultimate_cost_bearer_code = bearer.entity_code
            WHERE l.cost_month BETWEEN ? AND ?
              AND COALESCE(l.reallocation_mode, 'none') != 'none'
            ORDER BY 最终成本承担单位, 姓名, 核算月份
            """,
            conn_reallocation,
            params=[s_m, e_m]
        )
        conn_reallocation.close()
        reallocation_export = localize_labor_cost_codes(reallocation_export)
        if reallocation_export.empty:
            st.info("所选期间没有已标记的地市人工成本划转记录。")
        else:
            reallocation_summary = (
                reallocation_export.groupby(
                    ['最终成本承担单位', '业务关系类型'], dropna=False, as_index=False
                )
                .agg(
                    核算人次=('工号', 'count'),
                    工资应发合计=('工资应发合计', 'sum'),
                    其他人工成本合计=('其他人工成本合计', 'sum'),
                    人工成本合计=('人工成本合计', 'sum')
                )
            )
            reallocation_io = io.BytesIO()
            with pd.ExcelWriter(reallocation_io, engine='openpyxl') as writer:
                reallocation_export.to_excel(writer, index=False, sheet_name='划转明细')
                reallocation_summary.to_excel(writer, index=False, sheet_name='地市汇总')
                format_excel_sheet(writer.sheets['划转明细'], reallocation_export.columns)
                format_excel_sheet(writer.sheets['地市汇总'], reallocation_summary.columns)
            st.download_button(
                "📥 下载地市人工成本划转表",
                data=reallocation_io.getvalue(),
                file_name=f"地市人工成本划转_{s_m}至{e_m}.xlsx",
            )

    st.divider()

    # ==========================================================================
    # 下月数据初始化引擎 (纯净底表繁衍)
    # ==========================================================================
    st.subheader("🆕 生成新月份初始化底表 (融合社保版)")
    st.info("💡 痛点解决：系统将提取【基准月】人员架构，追加新入职员工。系统去社保模块抓取【目标生成月】扣款数据时，会自动将(基本医疗+大病)完美合并，并填入本表的“医疗保险-个人(含大病)”中！")

    col_t1, col_t2, col_t3 = st.columns(3)
    with col_t1: base_month = st.selectbox("参照基准月 (提取其人员框架)", options=available_months)
    with col_t2: target_month = st.text_input("目标生成月 (格式 YYYY-MM)", value="2026-02" if base_month == "2026-01" else "")
    with col_t3: st.write(""); clear_nums = st.checkbox("清空薪酬类变动金额 (但保留自动抓取的社保公积金)", value=True)

    if st.button("🚀 生成融合社保的录入底表", type="primary"):
        if not target_month:
            st.warning("请填写目标月份！")
        else:
            conn = _get_db_connection()
            try:
                base_df = pd.read_sql_query("SELECT * FROM labor_cost_ledger WHERE cost_month = ?", conn, params=[base_month])
                all_emps = pd.read_sql_query("SELECT emp_id, employee_no, name, dept_id, status FROM employees", conn)
                active_emps = all_emps[all_emps['status'] == '在职']
                active_emps = active_emps[
                    active_emps['emp_id'].astype(str).map(
                        lambda emp_id: is_labor_cost_included(emp_id, target_month, conn)
                    )
                ]

                dept_df = pd.read_sql_query("SELECT dept_id, dept_name FROM departments", conn)
                dept_dict = dict(zip(dept_df['dept_id'], dept_df['dept_name']))

                ss_df = get_company_social_snapshot(target_month, conn)

                if not base_df.empty:
                    emp_status_dict = dict(zip(all_emps['emp_id'], all_emps['status']))

                    effective_dept_snapshot = get_effective_department_snapshot(target_month, conn)

                    keep_mask = []
                    for idx, row in base_df.iterrows():
                        eid = str(row['emp_id'])
                        cost = row.get('total_labor_cost', 0.0)
                        if pd.isna(cost): cost = 0.0

                        curr_status = emp_status_dict.get(eid, row.get('emp_status', '在职'))

                        department = effective_dept_snapshot.get(eid)
                        if department:
                            base_df.at[idx, 'dept_id'] = department['dept_id']
                            base_df.at[idx, 'dept_name'] = department['dept_name']

                        base_df.at[idx, 'emp_status'] = curr_status

                        if not is_labor_cost_included(eid, target_month, conn):
                            keep_mask.append(False)
                        elif '离职' in curr_status and cost == 0.0:
                            keep_mask.append(False)
                        else:
                            keep_mask.append(True)

                    base_df = base_df[keep_mask]
                    base_df['cost_month'] = target_month
                    base_emp_ids = set(base_df['emp_id'].tolist())

                    new_emps = active_emps[~active_emps['emp_id'].isin(base_emp_ids)]
                    new_rows = []
                    for _, r in new_emps.iterrows():
                        new_row = {col: 0.0 if DB_TO_CN_MAP.get(col) in NUMERIC_COLS else None for col in base_df.columns}
                        new_row['cost_month'] = target_month
                        new_row['emp_id'] = r['emp_id']
                        new_row['emp_name'] = r['name']
                        department = effective_dept_snapshot.get(str(r['emp_id']))
                        new_row['dept_id'] = department['dept_id'] if department else r['dept_id']
                        new_row['dept_name'] = department['dept_name'] if department else dept_dict.get(r['dept_id'], '未分配部门')
                        new_row['emp_status'] = r['status']
                        new_rows.append(new_row)

                    if new_rows:
                        base_df = pd.concat([base_df, pd.DataFrame(new_rows)], ignore_index=True)

                    # 把当月有效的用工与成本关系固化进初始化底表。
                    for idx, relation_row in base_df.iterrows():
                        relation_emp_id = str(relation_row.get('emp_id', '')).replace('.0', '').strip()
                        if not relation_emp_id:
                            continue
                        arrangement = get_effective_arrangement(relation_emp_id, target_month, conn)
                        relation_type = arrangement.get('arrangement_type', 'normal')
                        if relation_type == 'city_transfer':
                            reallocation_mode = 'annual_labor_cost_reallocation'
                            reallocation_status = 'pending'
                        else:
                            reallocation_mode = 'none'
                            reallocation_status = 'not_required'

                        base_df.at[idx, 'arrangement_id'] = arrangement.get('arrangement_id')
                        base_df.at[idx, 'business_type_snapshot'] = relation_type
                        base_df.at[idx, 'labor_cost_included_snapshot'] = 1
                        base_df.at[idx, 'actual_work_unit_code'] = arrangement.get('actual_work_unit_code')
                        base_df.at[idx, 'accounting_entity_code'] = arrangement.get('accounting_entity_code')
                        base_df.at[idx, 'ultimate_cost_bearer_code'] = arrangement.get('ultimate_cost_bearer_code')
                        base_df.at[idx, 'reallocation_mode'] = reallocation_mode
                        base_df.at[idx, 'reallocation_status'] = reallocation_status

                    export_cn = base_df.rename(columns=DB_TO_CN_MAP)
                    employee_no_map = dict(zip(all_emps['emp_id'], all_emps['employee_no']))
                    export_cn['工号'] = export_cn['工号'].map(employee_no_map).fillna('待分配')

                    if clear_nums:
                        for cn_col in NUMERIC_COLS:
                            if cn_col in export_cn.columns: export_cn[cn_col] = 0.0

                    if not ss_df.empty:
                        ss_df['emp_id'] = ss_df['emp_id'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
                        ss_index_df = ss_df.set_index('emp_id')

                        for idx, row in export_cn.iterrows():
                            eid = resolve_employee_reference(
                                employee_no=row.get('工号'),
                                id_card=row.get('身份证号'),
                                name=row.get('姓名'),
                                conn=conn,
                            )
                            if eid in ss_index_df.index:
                                ss_rec = ss_index_df.loc[eid]
                                if isinstance(ss_rec, pd.DataFrame):
                                    ss_rec = ss_rec.iloc[0]

                                export_cn.at[idx, '养老保险-企业'] = ss_rec.get('pension_comp', 0.0) if pd.notna(ss_rec.get('pension_comp')) else 0.0
                                export_cn.at[idx, '医疗保险-企业'] = ss_rec.get('medical_comp', 0.0) if pd.notna(ss_rec.get('medical_comp')) else 0.0
                                export_cn.at[idx, '失业保险-企业'] = ss_rec.get('unemp_comp', 0.0) if pd.notna(ss_rec.get('unemp_comp')) else 0.0
                                export_cn.at[idx, '工伤保险-企业'] = ss_rec.get('injury_comp', 0.0) if pd.notna(ss_rec.get('injury_comp')) else 0.0
                                export_cn.at[idx, '生育保险-企业'] = ss_rec.get('maternity_comp', 0.0) if pd.notna(ss_rec.get('maternity_comp')) else 0.0
                                export_cn.at[idx, '住房公积金-企业'] = ss_rec.get('fund_comp', 0.0) if pd.notna(ss_rec.get('fund_comp')) else 0.0
                                export_cn.at[idx, '企业年金-企业'] = ss_rec.get('annuity_comp', 0.0) if pd.notna(ss_rec.get('annuity_comp')) else 0.0

                                export_cn.at[idx, '养老保险-个人'] = ss_rec.get('pension_pers', 0.0) if pd.notna(ss_rec.get('pension_pers')) else 0.0

                                m_pers = ss_rec.get('medical_pers', 0.0) if pd.notna(ss_rec.get('medical_pers')) else 0.0
                                export_cn.at[idx, '医疗保险-个人(含大病)'] = m_pers

                                export_cn.at[idx, '失业保险-个人'] = ss_rec.get('unemp_pers', 0.0) if pd.notna(ss_rec.get('unemp_pers')) else 0.0
                                export_cn.at[idx, '住房公积金-个人'] = ss_rec.get('fund_pers', 0.0) if pd.notna(ss_rec.get('fund_pers')) else 0.0
                                export_cn.at[idx, '企业年金-个人'] = ss_rec.get('annuity_pers', 0.0) if pd.notna(ss_rec.get('annuity_pers')) else 0.0

                    ordered_cols = [c for c in LEDGER_MAP.keys() if c in export_cn.columns]
                    export_cn = export_cn[ordered_cols]
                    export_cn = sort_flat_ledger_df(export_cn)

                    ob_clean = io.BytesIO()
                    with pd.ExcelWriter(ob_clean, engine='openpyxl') as w:
                        export_cn.to_excel(w, index=False, sheet_name=f"{target_month}融合明细")
                        ws = w.sheets[f"{target_month}融合明细"]
                        ws.freeze_panes = 'A2'
                        for col_idx in range(1, ws.max_column + 1):
                            ws.column_dimensions[get_column_letter(col_idx)].width = 15

                    st.success(f"✅ 底表生成成功！已成功从社保模块抓取 {target_month} 的数据，(基本医疗+大病)已完美合并注入底表。")
                    st.download_button(f"📥 下载 {target_month} 融合社保底表", data=ob_clean.getvalue(), file_name=f"台账初始化_{target_month}.xlsx", type="secondary")
                else:
                    st.error("基准月没有数据，无法繁衍！")
            except Exception as e:
                st.error(f"生成底表崩溃: {e}")
            finally:
                if 'conn' in locals(): conn.close()

# ------------------------------------------------------------------------------
# Tab 3: 财务底表导入引擎
# ------------------------------------------------------------------------------
with tab3:
    st.subheader("📥 人工成本底表导入与财务核对")

    st.write("### 🧾 导入前：财务表自动补数与双重核对")
    st.info(
        "只上传人员台账底表即可自动计算工会经费、职工教育经费，分摊尾差固定由人力资源部主任承接。"
        "当月财务表和累计财务表都是可选项：上传哪张就核对哪张，累计差异只报警，不会反过来改本月。"
    )

    finance_upload_col1, finance_upload_col2, finance_upload_col3 = st.columns(3)
    with finance_upload_col1:
        precheck_ledger_file = st.file_uploader(
            "1. 人员人工成本台账底表",
            type=["xlsx", "csv"],
            key="finance_precheck_ledger",
        )
    with finance_upload_col2:
        precheck_monthly_file = st.file_uploader(
            "2. 财务当月实际人工成本表（可选）",
            type=["xlsx"],
            key="finance_precheck_monthly",
        )
    with finance_upload_col3:
        precheck_ytd_file = st.file_uploader(
            "3. 财务本年累计人工成本表（可选）",
            type=["xlsx"],
            key="finance_precheck_ytd",
        )

    all_precheck_files_ready = precheck_ledger_file is not None
    if st.button(
        "🔍 自动补数并生成核对结果",
        type="primary",
        disabled=not all_precheck_files_ready,
    ):
        try:
            draft_ledger = read_labor_ledger_workbook(
                precheck_ledger_file,
                file_name=precheck_ledger_file.name,
            )
            month_values = (
                draft_ledger['核算月份']
                .dropna()
                .astype(str)
                .str[:7]
                .str.replace('/', '-', regex=False)
                .unique()
                .tolist()
            )
            if len(month_values) != 1:
                raise ValueError(
                    f"人员台账必须且只能包含一个核算月份，当前识别到：{month_values}"
                )
            precheck_month = month_values[0]

            monthly_finance = (
                read_finance_account_workbook(precheck_monthly_file)
                if precheck_monthly_file is not None
                else None
            )
            ytd_finance = (
                read_finance_account_workbook(precheck_ytd_file)
                if precheck_ytd_file is not None
                else None
            )
            if monthly_finance is not None and ytd_finance is not None:
                monthly_wage = monthly_finance.loc[
                    monthly_finance['科目编号'].eq('6400010100'),
                    '本期借方发生额',
                ].sum()
                ytd_wage = ytd_finance.loc[
                    ytd_finance['科目编号'].eq('6400010100'),
                    '本期借方发生额',
                ].sum()
                if ytd_wage + 0.01 < monthly_wage:
                    raise ValueError(
                        "累计表的工资发生额小于当月表，两个财务文件可能传反了。"
                    )

            history_conn = _get_db_connection()
            history_start = f"{precheck_month[:4]}-01"
            historical_ledger = pd.read_sql_query(
                '''
                SELECT * FROM labor_cost_ledger
                WHERE cost_month >= ? AND cost_month < ?
                ORDER BY cost_month, emp_id
                ''',
                history_conn,
                params=[history_start, precheck_month],
            )
            hr_director_id, hr_director_name = resolve_hr_director_tail_carrier(
                history_conn,
                draft_ledger,
            )
            history_conn.close()

            precheck_result = prepare_finance_labor_precheck(
                draft_ledger,
                monthly_finance,
                ytd_finance_df=ytd_finance,
                historical_ledger_df=historical_ledger,
                tail_carrier_emp_id=hr_director_id,
            )
            st.session_state['finance_labor_precheck'] = {
                'cost_month': precheck_month,
                'result': precheck_result,
                'file_names': {
                    'ledger': precheck_ledger_file.name,
                    'monthly': precheck_monthly_file.name if precheck_monthly_file else None,
                    'ytd': precheck_ytd_file.name if precheck_ytd_file else None,
                },
                'tail_carrier': f'{hr_director_name}（{hr_director_id}）',
            }
            completed_tasks = ['自动补数']
            if monthly_finance is not None:
                completed_tasks.append('当月核对')
            if ytd_finance is not None:
                completed_tasks.append('累计核对')
            st.success(
                f"✅ {precheck_month} {'、'.join(completed_tasks)}已完成；"
                f"经费尾差承接人：{hr_director_name}（{hr_director_id}）。"
            )
        except Exception as error:
            st.session_state.pop('finance_labor_precheck', None)
            st.error(f"预核对失败：{error}")

    precheck_state = st.session_state.get('finance_labor_precheck')
    if precheck_state:
        precheck_result = precheck_state['result']
        monthly_reconciliation = precheck_result['monthly_reconciliation']
        ytd_reconciliation = precheck_result['ytd_reconciliation']

        monthly_issue_count = int(
            monthly_reconciliation['核对状态'].ne('一致').sum()
        ) if not monthly_reconciliation.empty else 0
        ytd_issue_count = int(
            ytd_reconciliation['核对状态'].ne('一致').sum()
        ) if not ytd_reconciliation.empty else 0
        formula_status = precheck_result['business_checks'].iloc[0]['核对状态']

        summary_col1, summary_col2, summary_col3, summary_col4 = st.columns(4)
        summary_col1.metric("自动处理项目", len(precheck_result['auto_actions']))
        summary_col2.metric(
            "当月待处理项",
            monthly_issue_count if not monthly_reconciliation.empty else "未核对",
        )
        summary_col3.metric(
            "累计历史差异项",
            ytd_issue_count if not ytd_reconciliation.empty else "未核对",
        )
        summary_col4.metric("女工劳保实发公式", formula_status)

        if not precheck_result['auto_actions'].empty:
            st.write("#### 自动处理明细")
            st.dataframe(
                precheck_result['auto_actions'],
                use_container_width=True,
                hide_index=True,
            )

        if not monthly_reconciliation.empty:
            st.write("#### 当月财务核对")
            st.dataframe(
                monthly_reconciliation,
                use_container_width=True,
                hide_index=True,
            )
        if not ytd_reconciliation.empty:
            st.write("#### 本年累计核对")
            st.caption(
                "累计表用于发现以前月份留下的差异。这里有差异时，系统不会把差额硬塞进本月。"
            )
            st.dataframe(
                ytd_reconciliation,
                use_container_width=True,
                hide_index=True,
            )

        if not precheck_result['pending_accounts'].empty:
            st.warning("财务表中还有待确认或未建立映射的非零费用科目，请先确认业务性质。")
            st.dataframe(
                precheck_result['pending_accounts'],
                use_container_width=True,
                hide_index=True,
            )

        processed_workbook = build_finance_precheck_workbook(precheck_result)
        st.download_button(
            "📥 下载自动处理后的台账及核对报告",
            data=processed_workbook,
            file_name=f"{precheck_state['cost_month']}_人工成本自动补数及核对.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        pending_confirmation_rows = (
            monthly_reconciliation[
                monthly_reconciliation['处理方式'].eq('待业务确认')
                & monthly_reconciliation['财务金额'].abs().ge(0.005)
            ]
            if not monthly_reconciliation.empty
            else pd.DataFrame()
        )
        pending_amounts_balanced = (
            True
            if pending_confirmation_rows.empty
            else bool(
                pending_confirmation_rows['差额（台账-财务）']
                .abs()
                .le(0.01)
                .all()
            )
        )
        pending_business_confirmed = pending_confirmation_rows.empty
        if not pending_confirmation_rows.empty:
            pending_business_confirmed = st.checkbox(
                "我已确认待确认科目的费用性质、台账字段和人员归属",
                disabled=not pending_amounts_balanced,
                key="finance_pending_business_confirmed",
            )
        unresolved_amount_rows = (
            monthly_reconciliation[
                monthly_reconciliation['差额（台账-财务）'].abs().gt(0.01)
            ]
            if not monthly_reconciliation.empty
            else pd.DataFrame()
        )
        can_import_prechecked = (
            unresolved_amount_rows.empty
            and formula_status == '一致'
            and pending_business_confirmed
        )
        if can_import_prechecked:
            if st.button("✅ 核对无误，直接导入人工成本台账", type="primary"):
                try:
                    imported_count = upsert_labor_cost_dataframe(
                        precheck_result['processed_ledger']
                    )
                    batch_id = save_finance_precheck_audit(
                        precheck_state['cost_month'],
                        precheck_state['file_names'],
                        precheck_result,
                        imported_count,
                    )
                    st.success(
                        f"✅ 已导入 {imported_count} 条台账，并保存核对批次 {batch_id}。"
                    )
                except Exception as error:
                    st.error(f"核对结果导入失败：{error}")
        else:
            st.warning(
                "当前月仍有未处理差异或待确认科目，系统暂不允许直接入账。"
                "请在下载的台账中补充业务明细后重新核对。"
            )

    st.divider()

    # ==========================================================
    # 小工具：刷新已导入台账的部门归属
    # ==========================================================
    st.write("### 🧭 小工具：按人员变动刷新已导入台账部门归属")

    st.info(
        "这个工具只刷新已导入台账的【部门归属快照】，不会改任何金额。"
        "适用于你已经导入人工成本后，才发现人员调动日期漏维护或维护晚了的情况。"
    )

    conn_refresh = _get_db_connection()

    try:
        refresh_months = pd.read_sql_query(
            """
            SELECT DISTINCT cost_month
            FROM labor_cost_ledger
            ORDER BY cost_month DESC
            """,
            conn_refresh
        )["cost_month"].tolist()
    except Exception:
        refresh_months = []

    if not refresh_months:
        st.caption("当前人工成本台账暂无月份数据，无法刷新部门归属。")
        conn_refresh.close()
    else:
        refresh_col_1, refresh_col_2 = st.columns([1, 2])

        with refresh_col_1:
            refresh_month = st.selectbox(
                "选择要刷新的人工成本月份",
                refresh_months,
                key="ledger_refresh_month"
            )

        with refresh_col_2:
            st.warning(
                "执行前请确认：你已经在人员模块补录了正确的调动记录，"
                "并且已经备份数据库。刷新后会同时更新该月份的部门编号和部门名称快照。"
            )

        # ------------------------------------------------------
        # 生成刷新预览
        # ------------------------------------------------------
        if st.button("🔍 预览该月份需要调整的部门归属", type="secondary"):
            try:
                effective_dept_map = get_effective_department_snapshot(
                    refresh_month, conn_refresh
                )

                ledger_df = pd.read_sql_query(
                    """
                    SELECT
                        record_id,
                        cost_month,
                        emp_id,
                        emp_name,
                        dept_id,
                        dept_name,
                        emp_status,
                        total_labor_cost
                    FROM labor_cost_ledger
                    WHERE cost_month = ?
                    ORDER BY dept_name ASC, emp_id ASC
                    """,
                    conn_refresh,
                    params=[refresh_month]
                )

                if ledger_df.empty:
                    st.warning("当前月份没有人工成本台账数据。")
                else:
                    preview_rows = []

                    for _, row in ledger_df.iterrows():
                        emp_id = str(row["emp_id"]).replace(".0", "").strip()
                        employee_no = get_employee_no(emp_id, conn_refresh)
                        old_dept_id = row.get("dept_id")
                        old_dept_name = str(row["dept_name"]).strip()
                        target_department = effective_dept_map.get(emp_id)
                        if not target_department:
                            continue
                        new_dept_id = int(target_department['dept_id'])
                        new_dept_name = target_department['dept_name']

                        # 用部门ID判断真正的组织调动。仅仅改了部门名称时保留历史名称快照。
                        if pd.isna(old_dept_id) or int(old_dept_id) != new_dept_id:
                            preview_rows.append({
                                "流水ID": row["record_id"],
                                "核算月份": row["cost_month"],
                                "工号": employee_no,
                                "姓名": row["emp_name"],
                                "当前台账部门": old_dept_name,
                                "当前部门ID": old_dept_id,
                                "应调整为部门": new_dept_name,
                                "应调整为部门ID": new_dept_id,
                                "人员状态": row["emp_status"],
                                "人工成本合计": row["total_labor_cost"],
                            })

                    if not preview_rows:
                        st.success("✅ 预览完成：当前月份没有发现需要调整部门归属的台账记录。")
                    else:
                        preview_df = pd.DataFrame(preview_rows)

                        # 把预览结果暂存到 session_state。
                        # 这样用户看完预览后，再点执行按钮，系统知道要更新哪些 record_id。
                        st.session_state["ledger_dept_refresh_preview"] = preview_df
                        st.session_state["ledger_dept_refresh_month"] = refresh_month

                        st.warning(
                            f"⚠️ 预览发现 {len(preview_df)} 条台账记录需要刷新部门归属。"
                            "请在下方【待执行的部门归属调整】区域核对明细。"
                        )

            except Exception as e:
                st.error(f"预览失败：{e}")

        # ------------------------------------------------------
        # 执行刷新
        # ------------------------------------------------------
        if (
            "ledger_dept_refresh_preview" in st.session_state
            and "ledger_dept_refresh_month" in st.session_state
            and st.session_state["ledger_dept_refresh_month"] == refresh_month
        ):
            st.write("#### 待执行的部门归属调整")
            st.dataframe(
                st.session_state["ledger_dept_refresh_preview"],
                use_container_width=True,
                hide_index=True
            )

            confirm_refresh = st.checkbox(
                "我确认已经备份数据库，并确认执行上述部门归属刷新",
                key="confirm_ledger_dept_refresh"
            )

            if st.button("✅ 执行刷新部门归属", type="primary", disabled=not confirm_refresh):
                try:
                    cursor = conn_refresh.cursor()
                    preview_df = st.session_state["ledger_dept_refresh_preview"]

                    update_count = 0

                    for _, row in preview_df.iterrows():
                        cursor.execute(
                            """
                            UPDATE labor_cost_ledger
                            SET dept_id = ?, dept_name = ?
                            WHERE record_id = ?
                            """,
                            (
                                int(row["应调整为部门ID"]),
                                row["应调整为部门"],
                                int(row["流水ID"])
                            )
                        )
                        update_count += 1

                    conn_refresh.commit()

                    st.success(
                        f"✅ 部门归属刷新完成！已更新 {update_count} 条人工成本台账记录。"
                        "本次只修改归属部门，没有修改任何金额。"
                    )

                    # 清除预览缓存，防止重复执行。
                    del st.session_state["ledger_dept_refresh_preview"]
                    del st.session_state["ledger_dept_refresh_month"]

                except Exception as e:
                    conn_refresh.rollback()
                    st.error(f"执行刷新失败：{e}")

        conn_refresh.close()

    st.divider()

    tc1, tc2 = st.columns(2)
    with tc1:
        st.write("### 历史台账兼容导入")
        st.write("此入口不读取财务控制表，适合补录旧月份。新月份请优先使用上方的自动补数与双重核对。系统依靠 **`核算月份`** 和 **`工号`** 确认数据。")
        template_df = pd.DataFrame(columns=list(LEDGER_MAP.keys()))
        tout = io.BytesIO()
        with pd.ExcelWriter(tout, engine='openpyxl') as w: template_df.to_excel(w, index=False)
        st.download_button("下载标准导入模板", data=tout.getvalue(), file_name="人工成本台账导入模板.xlsx")

    with tc2:
        up_file = st.file_uploader("上传已填写的台账 Excel", type=["xlsx", "csv"])
        if up_file and st.button("🚀 执行导入与数据库覆盖"):
            conn = None
            try:
                in_df = read_labor_ledger_workbook(up_file, file_name=up_file.name)

                conn = _get_db_connection()
                cursor = conn.cursor()
                success_count = 0

                for idx, row in in_df.iterrows():
                    e_name = str(row.get('姓名', '')).strip()
                    d_name = str(row.get('归属部门', '')).strip()
                    if e_name in ['【小计】', '【实际成本总计】'] or d_name in ['【在职及统筹部分】']:
                        continue

                    raw_month = str(row.get('核算月份', '')).strip()
                    raw_id = row.get('工号', '')
                    if pd.isna(raw_id): e_id = ""
                    elif isinstance(raw_id, float): e_id = str(int(raw_id))
                    else: e_id = str(raw_id).replace('.0', '').strip()

                    if not raw_month or raw_month == 'nan': continue
                    c_month = raw_month[:7].replace('/', '-') if len(raw_month) >= 7 else raw_month
                    internal_emp_id = resolve_employee_reference(
                        employee_no=e_id,
                        id_card=row.get('身份证号'),
                        name=e_name,
                        conn=conn,
                    )
                    if not internal_emp_id:
                        raise ValueError(
                            f'无法识别人员：{e_name or "未填写姓名"}'
                            f'（工号：{e_id or "待分配"}）'
                        )
                    arrangement = get_effective_arrangement(internal_emp_id, c_month, conn)
                    if not int(arrangement.get('labor_cost_included', 1)):
                        cursor.execute(
                            "DELETE FROM labor_cost_ledger WHERE cost_month = ? AND emp_id = ?",
                            (c_month, internal_emp_id),
                        )
                        continue

                    db_data = {}
                    for cn_col, db_col in LEDGER_MAP.items():
                        if db_col == 'cost_month':
                            db_data[db_col] = c_month
                            continue

                        val = row.get(cn_col, None)
                        if cn_col in NUMERIC_COLS:
                            try:
                                clean_val = str(val).replace(',', '').strip()
                                db_data[db_col] = float(clean_val) if pd.notna(val) and clean_val != '' else 0.0
                            except:
                                db_data[db_col] = 0.0
                        else:
                            db_data[db_col] = str(val).strip() if pd.notna(val) else ""

                    relation_type = arrangement.get('arrangement_type', 'normal')
                    db_data['emp_id'] = internal_emp_id
                    db_data['arrangement_id'] = arrangement.get('arrangement_id')
                    db_data['business_type_snapshot'] = relation_type
                    db_data['labor_cost_included_snapshot'] = 1
                    db_data['actual_work_unit_code'] = arrangement.get('actual_work_unit_code')
                    db_data['accounting_entity_code'] = arrangement.get('accounting_entity_code') or 'province_public'
                    db_data['ultimate_cost_bearer_code'] = arrangement.get('ultimate_cost_bearer_code') or 'province_public'
                    department = get_effective_department_snapshot(c_month, conn).get(internal_emp_id)
                    if department:
                        db_data['dept_id'] = department['dept_id']
                        db_data['dept_name'] = department['dept_name']
                    if relation_type == 'city_transfer':
                        db_data['reallocation_mode'] = 'annual_labor_cost_reallocation'
                        db_data['reallocation_status'] = 'pending'
                    else:
                        db_data['reallocation_mode'] = 'none'
                        db_data['reallocation_status'] = 'not_required'

                    # =============================================================
                    # 女工劳保费特殊口径
                    # =============================================================
                    #
                    # 正式系统名称统一为：
                    # 女工劳保费
                    #
                    # 数据库字段统一为：
                    # allowance_women
                    #
                    # 业务规则：
                    # 1. 女工劳保费随工资发放；
                    # 2. 继续在人工成本台账中单独记录；
                    # 3. 不计入人工成本口径工资应发；
                    # 4. 不计入其他人工成本；
                    # 5. 不计入人工成本合计；
                    # 6. 计入员工个人实际实发。

                    # -------------------------------------------------------------
                    # 一、读取女工劳保费
                    # -------------------------------------------------------------
                    women_labor_fee = float(
                        db_data.get('allowance_women', 0.0) or 0.0
                    )

                    # -------------------------------------------------------------
                    # 二、计算人工成本口径工资应发
                    # -------------------------------------------------------------
                    #
                    # 注意：
                    # 这里故意不放 allowance_women。
                    #
                    # 因此 labor_cost_ledger.gross_salary_total
                    # 表示省公司人工成本口径的工资应发，
                    # 不等同于薪酬模块工资条上的完整应发。
                    gross_cols = [
                        'base_salary',
                        'seniority_pay',
                        'comp_subsidy',
                        'perf_float_subsidy',
                        'telecom_subsidy',
                        'other_base_pay',
                        'intern_subsidy',
                        'grad_allowance',
                        'eval_perf_pay',
                        'commission_pay',
                        'other_month_perf',
                        'special_award',
                        'year_end_bonus',
                        'other_special_award'
                    ]

                    calc_gross = sum(
                        db_data.get(col, 0.0)
                        for col in gross_cols
                    )

                    # -------------------------------------------------------------
                    # 三、计算个人代扣
                    # -------------------------------------------------------------
                    deduct_cols = [
                        'pension_personal',
                        'medical_personal',
                        'unemployment_personal',
                        'provident_fund_personal',
                        'annuity_personal',
                        'tax_personal_month',
                        'tax_personal_bonus'
                    ]

                    total_personal_deduction = sum(
                        db_data.get(col, 0.0)
                        for col in deduct_cols
                    )

                    # -------------------------------------------------------------
                    # 四、计算个人实际实发
                    # -------------------------------------------------------------
                    #
                    # 女工劳保费虽然不计入人工成本，
                    # 但它确实随工资发给员工，
                    # 所以个人实发必须加上女工劳保费。
                    #
                    # 个人实发 =
                    # 人工成本口径工资应发
                    # + 女工劳保费
                    # - 五险两金个人扣款
                    # - 个税
                    calc_net = (
                            calc_gross
                            + women_labor_fee
                            - total_personal_deduction
                    )

                    # -------------------------------------------------------------
                    # 五、计算其他人工成本
                    # -------------------------------------------------------------
                    #
                    # 原代码把 allowance_women 放在这里，
                    # 所以女工劳保费会进入其他人工成本。
                    #
                    # 现在必须移除。
                    company_cost_cols = [
                        'pension_company',
                        'medical_company',
                        'unemployment_company',
                        'work_injury_company',
                        'maternity_company',
                        'provident_fund_company',
                        'annuity_company',
                        'meal_daily',
                        'meal_ot',
                        'welfare_condolence',
                        'welfare_single_child',
                        'welfare_health_check',
                        'welfare_entry_check',
                        'welfare_other',
                        'allowance_heat',

                        # 注意：
                        # allowance_women 不再计入其他人工成本。

                        'medical_supplement',
                        'union_funds',
                        'edu_funds',
                        'cost_adjustment'
                    ]

                    calc_other = sum(
                        db_data.get(col, 0.0)
                        for col in company_cost_cols
                    )

                    # -------------------------------------------------------------
                    # 六、特殊公共账目保护
                    # -------------------------------------------------------------
                    #
                    # 有些股票增值权、公共统筹记录，
                    # 可能只有个人实发，没有工资应发明细。
                    #
                    # 只有当：
                    # 1. 工资应发为 0；
                    # 2. 女工劳保费也为 0；
                    # 3. Excel 原实发不为 0；
                    #
                    # 才保留 Excel 原始实发。
                    if (
                            calc_gross == 0.0
                            and women_labor_fee == 0.0
                            and db_data.get('net_salary', 0.0) != 0.0
                    ):
                        # 人工成本口径工资应发保持为 0。
                        db_data['gross_salary_total'] = 0.0

                        # 保留 Excel 中原始个人实发。
                        db_data['other_cost_total'] = calc_other

                        # 人工成本合计：
                        # 工资应发 + 其他人工成本。
                        #
                        # 女工劳保费不进入。
                        db_data['total_labor_cost'] = (
                                db_data['gross_salary_total']
                                + db_data['other_cost_total']
                        )

                    else:
                        # 人工成本口径工资应发：
                        # 不含女工劳保费。
                        db_data['gross_salary_total'] = calc_gross

                        # 个人实际实发：
                        # 包含女工劳保费。
                        db_data['net_salary'] = calc_net

                        # 其他人工成本：
                        # 不含女工劳保费。
                        db_data['other_cost_total'] = calc_other

                        # 人工成本合计：
                        # 不含女工劳保费。
                        db_data['total_labor_cost'] = (
                                calc_gross
                                + calc_other
                        )

                    columns = list(db_data.keys())
                    placeholders = ",".join(["?"] * len(columns))
                    updates = ",".join([f"{col}=excluded.{col}" for col in columns if col not in ['cost_month', 'emp_id']])

                    sql = f"""
                        INSERT INTO labor_cost_ledger ({",".join(columns)})
                        VALUES ({placeholders})
                        ON CONFLICT(cost_month, emp_id) DO UPDATE SET
                        {updates}
                    """
                    cursor.execute(sql, tuple(db_data.values()))
                    success_count += 1

                conn.commit()
                # [修复点 2：UI 防跳跃] 彻底抛弃强制重新加载整个页面的 st.rerun()
                st.success(f"✅ 台账导入/覆盖完成！成功处理 {success_count} 条记录。")
            except Exception as e:
                if conn is not None:
                    conn.rollback()
                st.error(f"导入底层崩溃: {e}")
            finally:
                if conn is not None:
                    conn.close()
