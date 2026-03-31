# ==============================================================================
# 文件路径: pages/2_人工成本台账.py
# 功能描述: 人工成本台账管理中心 (UI 呈现层)
# 核心修正说明:
#   1. 彻底解耦，底层逻辑导入自 modules.core_labor_cost，保持页面纯净。
#   2. Tab 2 导出模块修改为“选择起始至结束月份”的范围判定逻辑。
# ==============================================================================

import streamlit as st
import pandas as pd
import io

# 导入 Excel 报表精装修模块
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 导入业务分离后的核心层模块
from modules.core_labor_cost import (
    LEDGER_MAP, DB_TO_CN_MAP, NUMERIC_COLS,
    _get_db_connection, cleanse_db_timestamps,
    sort_flat_ledger_df, add_subtotals_and_totals, get_ledger_data
)

st.set_page_config(page_title="人工成本台账", layout="wide")

# 每次刷新页面静默执行时间戳解毒
cleanse_db_timestamps()

# ==============================================================================
# UI 消息锁状态机
# ==============================================================================
if 'ledger_msg' in st.session_state:
    if st.session_state.ledger_msg_type == 'success': st.success(st.session_state.ledger_msg)
    else: st.error(st.session_state.ledger_msg)
    del st.session_state.ledger_msg, st.session_state.ledger_msg_type

def set_msg(msg, type='success'):
    st.session_state.ledger_msg = msg
    st.session_state.ledger_msg_type = type
    st.rerun()

# ==============================================================================
# Excel 财务级排版渲染引擎 (3D 高亮视觉版)
# ==============================================================================
def format_excel_sheet(worksheet, df_columns):
    worksheet.freeze_panes = 'A2'
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    subtotal_fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
    total_fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
    key_col_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    key_columns = ['工资应发合计', '其他人工成本合计', '人工成本合计']

    df_cols_list = list(df_columns)
    name_col_idx = df_cols_list.index('姓名') + 1 if '姓名' in df_cols_list else -1

    for row_idx in range(1, worksheet.max_row + 1):
        is_subtotal = False
        is_total = False
        if name_col_idx != -1 and row_idx > 1:
            cell_val = str(worksheet.cell(row=row_idx, column=name_col_idx).value or "")
            if "【小计】" == cell_val: is_subtotal = True
            elif "【实际成本总计】" == cell_val: is_total = True

        for col_idx, col_name in enumerate(df_columns, 1):
            col_letter = get_column_letter(col_idx)
            cell = worksheet[f"{col_letter}{row_idx}"]
            cell.border = thin_border

            # 表头渲染
            if row_idx == 1:
                worksheet.column_dimensions[col_letter].width = 8 if col_name == '序号' else (15 if col_name in NUMERIC_COLS else 12)
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                # 动态数字列千分位格式化
                if col_name in NUMERIC_COLS:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # 汇总行底色高亮隔离
                if is_total:
                    cell.fill = total_fill
                    cell.font = Font(bold=True, color="000000")
                elif is_subtotal:
                    cell.fill = subtotal_fill
                    cell.font = Font(bold=True, color="000000")
                elif col_name in key_columns:
                    cell.fill = key_col_fill
                    cell.font = Font(bold=True)

# ==============================================================================
# 页面主框架
# ==============================================================================
st.title("💰 人工成本台账管理中心")
st.caption("🔒 财务数据合规要求：台账一旦生成不可在系统内手动篡改。如需修正，请在 Excel 中修改后重新导入，系统将自动覆盖重置原账目。")

tab1, tab2, tab3 = st.tabs(["📊 台账多维看板", "📤 领导审阅导出 (范围框选)", "📥 财务底表导入"])

# ------------------------------------------------------------------------------
# Tab 1: 台账多维看板
# ------------------------------------------------------------------------------
with tab1:
    conn = _get_db_connection()
    available_months = pd.read_sql_query("SELECT DISTINCT cost_month FROM labor_cost_ledger ORDER BY cost_month DESC", conn)['cost_month'].tolist()
    available_depts = pd.read_sql_query("SELECT DISTINCT dept_name FROM labor_cost_ledger", conn)['dept_name'].tolist()
    conn.close()

    sc1, sc2, sc3 = st.columns([1, 1, 1])
    with sc1: f_month = st.selectbox("📅 核算月份筛选", ["全部月份"] + available_months)
    with sc2: f_dept = st.multiselect("🏢 归属部门筛选", options=available_depts)
    with sc3: q_search = st.text_input("🔍 搜姓名 / 工号")

    raw_df = get_ledger_data(month_filter=None if f_month == "全部月份" else f_month, dept_filter=f_dept if f_dept else None)

    if not raw_df.empty:
        if q_search:
            raw_df = raw_df[raw_df['emp_name'].str.contains(q_search, na=False) | raw_df['emp_id'].str.contains(q_search, na=False)]

    if not raw_df.empty:
        # 指标看板：严格剔除退休人员，还原企业真实的实际运营人工成本
        active_metric_df = raw_df[~( (raw_df['emp_status'] == '退休') | (raw_df['dept_name'].str.contains('离退休', na=False)) )]

        total_cost = active_metric_df['total_labor_cost'].sum()
        total_gross = active_metric_df['gross_salary_total'].sum()
        total_other = active_metric_df['other_cost_total'].sum()
        total_headcount = len(active_metric_df)

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("在职及统筹总成本 (元)", f"{total_cost:,.2f}")
        m2.metric("工资应发合计 (元)", f"{total_gross:,.2f}")
        m3.metric("其他人工成本合计 (元)", f"{total_other:,.2f}")
        m4.metric("在职及统筹核算人次", f"{total_headcount} 人次")

        disp_df = raw_df.rename(columns=DB_TO_CN_MAP)
        disp_cols = [col for col in LEDGER_MAP.keys() if col in disp_df.columns]

        disp_final = sort_flat_ledger_df(disp_df[disp_cols].copy())
        disp_final.insert(0, '序号', range(1, len(disp_final) + 1))
        st.dataframe(disp_final, use_container_width=True, hide_index=True)
    else:
        st.info("💡 当前筛选条件下暂无台账数据。")

# ------------------------------------------------------------------------------
# Tab 2: 领导审阅导出 (跨期范围框选引擎)
# ------------------------------------------------------------------------------
with tab2:
    st.subheader("📤 生成向领导汇报的标准台账")

    st.info("💡 操作提示：请选择导出的【起始】和【结束】月份，系统会自动提取该时间段内的所有明细，并可生成累计总账。")

    tc1, tc2 = st.columns(2)
    with tc1:
        # [核心修改] 替换为单选框划定范围，不再让用户痛苦地多选打勾
        start_month = st.selectbox("📅 导出起始月份", options=available_months if available_months else ["无数据"], index=len(available_months)-1 if available_months else 0)
    with tc2:
        end_month = st.selectbox("📅 导出结束月份", options=available_months if available_months else ["无数据"], index=0 if available_months else 0)

    need_summary = st.checkbox("📊 同时生成【选中范围的累计汇总】Sheet (勾选后，系统会自动把这几个月加起来算个总账)", value=True)

    if st.button("🚀 一键生成并下载 Excel 报表", type="primary"):
        if start_month == "无数据" or end_month == "无数据":
            st.warning("⚠️ 暂无可导出的数据！")
        else:
            # 智能对撞，防止操作失误将起止时间选反
            s_m, e_m = min(start_month, end_month), max(start_month, end_month)
            # 根据划定区间，自动补全该范围内的所有月份列表
            selected_months = [m for m in available_months if s_m <= m <= e_m]
            selected_months.sort()

            conn = _get_db_connection()
            placeholders = ",".join(["?"] * len(selected_months))
            query = f"SELECT * FROM labor_cost_ledger WHERE cost_month IN ({placeholders}) ORDER BY dept_name ASC"
            raw_export_df = pd.read_sql_query(query, conn, params=selected_months)
            conn.close()

            if not raw_export_df.empty:
                ob = io.BytesIO()
                with pd.ExcelWriter(ob, engine='openpyxl') as writer:

                    if need_summary:
                        db_num_cols = [LEDGER_MAP[c] for c in NUMERIC_COLS if c in LEDGER_MAP]
                        agg_dict = {col: 'sum' for col in db_num_cols}
                        agg_dict.update({'emp_name': 'first', 'dept_name': 'first', 'emp_status': 'last'})

                        summary_df = raw_export_df.groupby('emp_id').agg(agg_dict).reset_index()
                        summary_cn = summary_df.rename(columns=DB_TO_CN_MAP)
                        report_cols = [c for c in LEDGER_MAP.keys() if c in summary_cn.columns and c != '核算月份']
                        summary_cn = summary_cn[report_cols]

                        summary_final = add_subtotals_and_totals(summary_cn, NUMERIC_COLS)

                        sum_sheet_name = f"{len(selected_months)}个月累计汇总"
                        summary_final.to_excel(writer, index=False, sheet_name=sum_sheet_name)
                        format_excel_sheet(writer.sheets[sum_sheet_name], summary_final.columns)

                    for month in sorted(selected_months):
                        month_df = raw_export_df[raw_export_df['cost_month'] == month].copy()
                        if not month_df.empty:
                            month_cn = month_df.rename(columns=DB_TO_CN_MAP)
                            month_cols = [c for c in LEDGER_MAP.keys() if c in month_cn.columns]
                            month_cn = month_cn[month_cols]

                            month_final = add_subtotals_and_totals(month_cn, NUMERIC_COLS)
                            # 清洗工作表名称中的非法符号
                            safe_month = str(month).replace(':', '-').replace('/', '-').replace('\\', '-').replace('?', '').replace('*', '')
                            safe_sheet_name = f"{safe_month[:28]}明细"

                            month_final.to_excel(writer, index=False, sheet_name=safe_sheet_name)
                            format_excel_sheet(writer.sheets[safe_sheet_name], month_final.columns)

                file_name = f"人工成本台账汇报_{len(selected_months)}个月数据.xlsx"
                st.download_button("📥 点击下载财务报表", data=ob.getvalue(), file_name=file_name, type="secondary")
            else:
                st.warning("所选范围内无数据。")

    st.divider()

    # ==========================================================================
    # 下月数据初始化引擎 (纯净底表繁衍 + 自动抓取当月社保固化数据)
    # ==========================================================================
    st.subheader("🆕 生成新月份初始化底表 (融合社保版)")
    st.info("💡 痛点解决：系统将提取【基准月】人员架构，追加新入职员工。同时，系统会自动去社保模块抓取【目标生成月】已固化的五险两金真实扣款数据，强行注入到底表中！")

    col_t1, col_t2, col_t3 = st.columns(3)
    with col_t1: base_month = st.selectbox("参照基准月 (提取其人员框架)", options=available_months)
    with col_t2: target_month = st.text_input("目标生成月 (格式 YYYY-MM)", value="2026-02" if base_month == "2026-01" else "")
    with col_t3: st.write(""); clear_nums = st.checkbox("清空薪酬类变动金额 (但保留自动抓取的社保公积金)", value=True)

    if st.button("🚀 生成融合社保的录入底表", type="primary"):
        if not target_month:
            st.warning("请填写目标月份！")
        else:
            conn = _get_db_connection()
            base_df = pd.read_sql_query("SELECT * FROM labor_cost_ledger WHERE cost_month = ?", conn, params=[base_month])
            all_emps = pd.read_sql_query("SELECT emp_id, name, dept_id, status FROM employees", conn)
            active_emps = all_emps[all_emps['status'] == '在职']

            dept_df = pd.read_sql_query("SELECT dept_id, dept_name FROM departments", conn)
            dept_dict = dict(zip(dept_df['dept_id'], dept_df['dept_name']))

            ss_query = """
                SELECT 
                    emp_id, 
                    pension_comp, medical_comp, unemp_comp, injury_comp, maternity_comp, fund_comp, annuity_comp,
                    pension_pers, medical_pers, medical_serious_pers, unemp_pers, fund_pers, annuity_pers
                FROM ss_monthly_records
                WHERE cost_month = ?
            """
            ss_df = pd.read_sql_query(ss_query, conn, params=[target_month])
            conn.close()

            if not base_df.empty:
                emp_status_dict = dict(zip(all_emps['emp_id'], all_emps['status']))
                keep_mask = []
                for _, row in base_df.iterrows():
                    eid = str(row['emp_id'])
                    cost = row.get('total_labor_cost', 0.0)
                    if pd.isna(cost): cost = 0.0

                    curr_status = emp_status_dict.get(eid, row.get('emp_status', '在职'))
                    # 强行驱逐离职且无残余成本的幽灵账目
                    if '离职' in curr_status and cost == 0.0:
                        keep_mask.append(False)
                    else:
                        keep_mask.append(True)

                base_df = base_df[keep_mask]
                base_df['cost_month'] = target_month
                base_emp_ids = set(base_df['emp_id'].tolist())
                new_emps = active_emps[~active_emps['emp_id'].isin(base_emp_ids)]

                new_rows = []
                for _, r in new_emps.iterrows():
                    new_row = {col: 0.0 if DB_TO_CN_MAP.get(col) in NUMERIC_COLS else None for col in base_df.columns}
                    new_row['cost_month'] = target_month
                    new_row['emp_id'] = r['emp_id']
                    new_row['emp_name'] = r['name']
                    new_row['dept_name'] = dept_dict.get(r['dept_id'], '未分配部门')
                    new_row['emp_status'] = r['status']
                    new_rows.append(new_row)

                if new_rows:
                    base_df = pd.concat([base_df, pd.DataFrame(new_rows)], ignore_index=True)

                export_cn = base_df.rename(columns=DB_TO_CN_MAP)

                if clear_nums:
                    for cn_col in NUMERIC_COLS:
                        if cn_col in export_cn.columns: export_cn[cn_col] = 0.0

                # 执行底层社保金额贴合注入
                if not ss_df.empty:
                    ss_df['emp_id'] = ss_df['emp_id'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
                    ss_index_df = ss_df.set_index('emp_id')

                    for idx, row in export_cn.iterrows():
                        eid = str(row.get('工号', '')).replace('.0', '').strip()
                        if eid in ss_index_df.index:
                            ss_rec = ss_index_df.loc[eid]
                            if isinstance(ss_rec, pd.DataFrame):
                                ss_rec = ss_rec.iloc[0]

                            export_cn.at[idx, '养老保险-企业'] = ss_rec.get('pension_comp', 0.0) if pd.notna(ss_rec.get('pension_comp')) else 0.0
                            export_cn.at[idx, '医疗保险-企业'] = ss_rec.get('medical_comp', 0.0) if pd.notna(ss_rec.get('medical_comp')) else 0.0
                            export_cn.at[idx, '失业保险-企业'] = ss_rec.get('unemp_comp', 0.0) if pd.notna(ss_rec.get('unemp_comp')) else 0.0
                            export_cn.at[idx, '工伤保险-企业'] = ss_rec.get('injury_comp', 0.0) if pd.notna(ss_rec.get('injury_comp')) else 0.0
                            export_cn.at[idx, '生育保险-企业'] = ss_rec.get('maternity_comp', 0.0) if pd.notna(ss_rec.get('maternity_comp')) else 0.0
                            export_cn.at[idx, '住房公积金-企业'] = ss_rec.get('fund_comp', 0.0) if pd.notna(ss_rec.get('fund_comp')) else 0.0
                            export_cn.at[idx, '企业年金-企业'] = ss_rec.get('annuity_comp', 0.0) if pd.notna(ss_rec.get('annuity_comp')) else 0.0

                            m_pers = ss_rec.get('medical_pers', 0.0) if pd.notna(ss_rec.get('medical_pers')) else 0.0
                            m_ser = ss_rec.get('medical_serious_pers', 0.0) if pd.notna(ss_rec.get('medical_serious_pers')) else 0.0

                            export_cn.at[idx, '养老保险-个人'] = ss_rec.get('pension_pers', 0.0) if pd.notna(ss_rec.get('pension_pers')) else 0.0
                            export_cn.at[idx, '医疗保险-个人'] = m_pers + m_ser
                            export_cn.at[idx, '失业保险-个人'] = ss_rec.get('unemp_pers', 0.0) if pd.notna(ss_rec.get('unemp_pers')) else 0.0
                            export_cn.at[idx, '住房公积金-个人'] = ss_rec.get('fund_pers', 0.0) if pd.notna(ss_rec.get('fund_pers')) else 0.0
                            export_cn.at[idx, '企业年金-个人'] = ss_rec.get('annuity_pers', 0.0) if pd.notna(ss_rec.get('annuity_pers')) else 0.0

                ordered_cols = [c for c in LEDGER_MAP.keys() if c in export_cn.columns]
                export_cn = export_cn[ordered_cols]
                export_cn = sort_flat_ledger_df(export_cn)

                ob_clean = io.BytesIO()
                with pd.ExcelWriter(ob_clean, engine='openpyxl') as w:
                    export_cn.to_excel(w, index=False, sheet_name=f"{target_month}融合明细")
                    ws = w.sheets[f"{target_month}融合明细"]
                    ws.freeze_panes = 'A2'
                    for col_idx in range(1, ws.max_column + 1):
                        ws.column_dimensions[get_column_letter(col_idx)].width = 13

                st.success(f"✅ 底表生成成功！已成功从社保模块抓取 {target_month} 的真实社保扣款并注入。")
                st.download_button(f"📥 下载 {target_month} 融合社保底表", data=ob_clean.getvalue(), file_name=f"台账初始化_{target_month}.xlsx", type="secondary")
            else:
                st.error("基准月没有数据，无法繁衍！")

# ------------------------------------------------------------------------------
# Tab 3: 财务底表导入引擎
# ------------------------------------------------------------------------------
with tab3:
    st.subheader("📥 历史财务数据导入引擎")

    tc1, tc2 = st.columns(2)
    with tc1:
        st.write("请先下载标准模板。系统依靠 **`核算月份`** 和 **`工号`** 确认数据。如果发现错漏，直接在 Excel 修改后重新上传，系统将自动覆盖。")
        template_df = pd.DataFrame(columns=list(LEDGER_MAP.keys()))
        tout = io.BytesIO()
        with pd.ExcelWriter(tout, engine='openpyxl') as w: template_df.to_excel(w, index=False)
        st.download_button("下载标准导入模板", data=tout.getvalue(), file_name="人工成本台账导入模板.xlsx")

    with tc2:
        up_file = st.file_uploader("上传已填写的台账 Excel", type=["xlsx", "csv"])
        if up_file and st.button("🚀 执行导入与数据库覆盖"):
            try:
                if up_file.name.endswith('.csv'):
                    in_df = pd.read_csv(up_file)
                else:
                    xls_dict = pd.read_excel(up_file, sheet_name=None)
                    in_df = pd.concat(xls_dict.values(), ignore_index=True)

                conn = _get_db_connection()
                cursor = conn.cursor()
                success_count = 0

                for idx, row in in_df.iterrows():
                    e_name = str(row.get('姓名', '')).strip()
                    d_name = str(row.get('归属部门', '')).strip()
                    # 防止导入包含汇总行的脏数据
                    if e_name in ['【小计】', '【实际成本总计】'] or d_name in ['【在职及统筹部分】']:
                        continue

                    raw_month = str(row.get('核算月份', '')).strip()
                    raw_id = row.get('工号', '')
                    if pd.isna(raw_id):
                        e_id = ""
                    elif isinstance(raw_id, float):
                        e_id = str(int(raw_id))
                    else:
                        e_id = str(raw_id).replace('.0', '').strip()

                    if not raw_month or not e_id or raw_month == 'nan' or e_id == 'nan': continue
                    c_month = raw_month[:7].replace('/', '-') if len(raw_month) >= 7 else raw_month

                    db_data = {}
                    for cn_col, db_col in LEDGER_MAP.items():
                        if db_col == 'cost_month':
                            db_data[db_col] = c_month
                            continue

                        val = row.get(cn_col, None)
                        if cn_col in NUMERIC_COLS:
                            try:
                                clean_val = str(val).replace(',', '').strip()
                                db_data[db_col] = float(clean_val) if pd.notna(val) and clean_val != '' else 0.0
                            except:
                                db_data[db_col] = 0.0
                        else:
                            db_data[db_col] = str(val).strip() if pd.notna(val) else ""

                    # 后端强制自动重新核算各项合计，防止前端 Excel 公式断裂导致入库错误
                    gross_cols = ['base_salary', 'seniority_pay', 'comp_subsidy', 'perf_float_subsidy', 'telecom_subsidy', 'other_base_pay', 'intern_subsidy', 'grad_allowance', 'eval_perf_pay', 'commission_pay', 'other_month_perf', 'special_award', 'year_end_bonus', 'other_special_award']
                    db_data['gross_salary_total'] = sum(db_data.get(col, 0.0) for col in gross_cols)

                    deduct_cols = ['pension_personal', 'medical_personal', 'unemployment_personal', 'provident_fund_personal', 'annuity_personal', 'tax_personal_month', 'tax_personal_bonus']
                    db_data['net_salary'] = db_data['gross_salary_total'] - sum(db_data.get(col, 0.0) for col in deduct_cols)

                    company_cost_cols = ['pension_company', 'medical_company', 'unemployment_company', 'work_injury_company', 'maternity_company', 'provident_fund_company', 'annuity_company', 'meal_daily', 'meal_ot', 'welfare_condolence', 'welfare_single_child', 'welfare_health_check', 'welfare_entry_check', 'welfare_other', 'allowance_heat', 'allowance_women', 'medical_supplement', 'union_funds', 'edu_funds', 'cost_adjustment']
                    db_data['other_cost_total'] = sum(db_data.get(col, 0.0) for col in company_cost_cols)

                    db_data['total_labor_cost'] = db_data['gross_salary_total'] + db_data['other_cost_total']

                    columns = list(db_data.keys())
                    placeholders = ",".join(["?"] * len(columns))
                    updates = ",".join([f"{col}=excluded.{col}" for col in columns if col not in ['cost_month', 'emp_id']])

                    sql = f"""
                        INSERT INTO labor_cost_ledger ({",".join(columns)})
                        VALUES ({placeholders})
                        ON CONFLICT(cost_month, emp_id) DO UPDATE SET
                        {updates}
                    """
                    cursor.execute(sql, tuple(db_data.values()))
                    success_count += 1

                conn.commit()
                set_msg(f"台账导入/覆盖完成！成功处理 {success_count} 条记录。")
            except Exception as e:
                conn.rollback()
                st.error(f"导入底层崩溃: {e}")
            finally:
                if 'conn' in locals(): conn.close()