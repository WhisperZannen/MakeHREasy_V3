# ==============================================================================
# 文件路径: pages/1_人员与组织.py
# 功能描述: 人员与组织架构管理中枢 (V3.19 挂靠人员兼容与工号解耦版)
# 实现了什么具体逻辑:
#   1. [核心防御] 彻底修复 Pandas Series 布尔判定歧义导致的 ValueError。
#   2. [核心防御] 引入动态列名推导，根除 KeyError: ['部门'] not in index 崩溃。
#   3. [全量复原] 满血召回“历史变动流水”的侧边栏高级多维筛选与时段 Excel 导出。
#   4. [状态对齐] 全面接入“挂靠人员”状态，支持以社保编号作为虚拟工号录入。
# ==============================================================================

import streamlit as st
import pandas as pd
import io
from datetime import datetime, date
from dateutil.relativedelta import relativedelta

# 导入底层接口
from modules.core_dept import get_all_departments, add_department, update_department, soft_delete_department
from modules.core_position import get_all_positions, add_position, update_position
from modules.core_personnel import get_all_employees, add_employee, update_employee, update_employee_status, get_all_history, rollback_history

st.set_page_config(page_title="组织人事中枢", layout="wide")

# ==============================================================================
# 自动初始化“公共挂靠池”
# ==============================================================================
def init_virtual_pools():
    ok_d, d_list = get_all_departments(include_inactive=True)
    if ok_d:
        d_names = [d['dept_name'] for d in d_list]
        if "离退休公共池" not in d_names: add_department("离退休公共池", "其他", None, 9999)

    ok_p, p_list = get_all_positions(include_inactive=True)
    if ok_p:
        p_names = [p['pos_name'] for p in p_list]
        if "无岗位" not in p_names: add_position("无岗位", "通用序列", 9999)

init_virtual_pools()

# ==============================================================================
# 消息状态保持与状态同步锁
# ==============================================================================
if 'ui_msg' in st.session_state:
    if st.session_state.ui_msg_type == 'success': st.success(st.session_state.ui_msg)
    else: st.error(st.session_state.ui_msg)
    del st.session_state.ui_msg, st.session_state.ui_msg_type

def set_msg_and_rerun(msg, type='success'):
    st.session_state.ui_msg = msg
    st.session_state.ui_msg_type = type
    st.session_state.editor_key = str(datetime.now())
    st.rerun()

if 'editor_key' not in st.session_state:
    st.session_state.editor_key = "init_v318"

# ==============================================================================
# 全局中文化映射
# ==============================================================================
DEPT_COL_MAP = {'dept_id': '部门ID', 'dept_name': '部门名称', 'dept_category': '性质', 'parent_dept_id': '上级ID', 'sort_order': '权重', 'status': '状态'}
POS_COL_MAP = {'pos_id': '岗位ID', 'pos_name': '岗位名称', 'pos_category': '序列', 'sort_order': '权重', 'status': '状态'}
EMP_COL_MAP = {
    'emp_id': '工号/编号', 'name': '姓名', 'id_card': '身份证号', 'dept_id': '部门ID', 'pos_id': '岗位ID',
    'post_rank': '岗级', 'post_grade': '档次', 'join_company_date': '入职日期', 'status': '状态',
    'pos_name': '岗位', 'tech_grade': 'T级', 'dept_name': '部门'
}

def clean_str(val): return "" if pd.isna(val) or val is None or str(val).lower()=='nan' else str(val).strip()
def clean_date(d_val):
    try: return pd.to_datetime(d_val).strftime('%Y-%m-%d') if pd.notna(pd.to_datetime(d_val)) else None
    except: return None

def refresh_data():
    ok_d, d = get_all_departments(include_inactive=True)
    ok_p, p = get_all_positions(include_inactive=True)
    ok_e, e = get_all_employees(include_resigned=True)
    return (pd.DataFrame(d) if ok_d and d else pd.DataFrame(columns=list(DEPT_COL_MAP.keys())),
            pd.DataFrame(p) if ok_p and p else pd.DataFrame(columns=list(POS_COL_MAP.keys())),
            pd.DataFrame(e) if ok_e and e else pd.DataFrame(columns=list(EMP_COL_MAP.keys())))

df_depts, df_positions, df_emps = refresh_data()

def build_dept_tree(df, parent_id=None, level=0):
    tree = []
    if df.empty: return tree
    children = df[df['parent_dept_id'].isna() | (df['parent_dept_id'] == 0)] if (pd.isna(parent_id) or parent_id == 0) else df[df['parent_dept_id'] == parent_id]
    for _, row in children.iterrows():
        r_d = row.to_dict()
        r_d['层级展示名'] = "　　" * level + ("└─ " if level > 0 else "") + str(row['dept_name'])
        p_name = "无"
        if pd.notna(row['parent_dept_id']) and row['parent_dept_id'] != 0:
            pm = df[df['dept_id'] == row['parent_dept_id']]
            if not pm.empty: p_name = pm.iloc[0]['dept_name']
        r_d['上级名称'] = p_name
        tree.append(r_d)
        tree.extend(build_dept_tree(df, row['dept_id'], level + 1))
    return tree

# ==============================================================================
# 侧边栏导航
# ==============================================================================
st.sidebar.title("🎛️ 系统架构中枢")
current_page = st.sidebar.radio("请选择操作模块:", ["🏢 部门管理", "🎯 岗位字典", "👥 人员档案", "🕰️ 历史变动流水"])
st.title(current_page)

# ==============================================================================
# 模块 A: 🏢 部门管理
# ==============================================================================
if current_page == "🏢 部门管理":
    col_d1, col_d2 = st.columns([1, 2])
    with col_d1:
        st.subheader("📝 部门信息维护")
        edit_mode = st.radio("模式", ["新增部门", "修改现有部门"], horizontal=True)
        target_dept = None
        if edit_mode == "修改现有部门":
            if df_depts.empty: st.warning("无部门数据")
            else:
                d_dict = {r['dept_id']: r['dept_name'] for _, r in df_depts.iterrows()}
                sel_d_id = st.selectbox("选择要修改的部门", options=list(d_dict.keys()), format_func=lambda x: d_dict[x])
                target_dept = df_depts[df_depts['dept_id'] == sel_d_id].iloc[0]

        with st.form("dept_form"):
            d_name = st.text_input("部门名称*", value=target_dept['dept_name'] if target_dept is not None else "")
            cat_opts = ["公司领导", "管控", "生产", "其他"]
            def_cat = target_dept['dept_category'] if target_dept is not None else "管控"
            d_cat = st.selectbox("性质*", cat_opts, index=cat_opts.index(def_cat) if def_cat in cat_opts else 1)
            d_sort = st.number_input("权重*", value=int(target_dept['sort_order']) if target_dept is not None else 999)

            v_p = df_depts[df_depts['status'] == 1]
            if target_dept is not None: v_p = v_p[v_p['dept_id'] != target_dept['dept_id']]
            p_opts = {0: "无(顶级)"}
            for _, r in v_p.iterrows(): p_opts[r['dept_id']] = r['dept_name']

            def_p = int(target_dept['parent_dept_id']) if target_dept is not None and pd.notna(target_dept['parent_dept_id']) else 0
            d_parent = st.selectbox("上级部门", options=list(p_opts.keys()), format_func=lambda x: p_opts[x], index=list(p_opts.keys()).index(def_p) if def_p in p_opts else 0)

            def_stat = "正常" if target_dept is None or target_dept['status'] == 1 else "已撤销"
            d_status = st.selectbox("状态", ["正常", "已撤销"], index=["正常", "已撤销"].index(def_stat))

            if st.form_submit_button("保存部门"):
                if not d_name: st.error("名称必填")
                else:
                    s_val = 1 if d_status == "正常" else 0
                    p_val = int(d_parent) if d_parent != 0 else None
                    if edit_mode == "新增部门":
                        success, msg = add_department(d_name, d_cat, p_val, d_sort)
                    else:
                        success, msg = update_department(int(target_dept['dept_id']), d_name, d_cat, p_val, d_sort, s_val)
                    if success: set_msg_and_rerun(msg)
                    else: st.error(msg)

        with st.expander("📥 批量导入部门"):
            tmp = pd.DataFrame(columns=['部门名称', '性质', '上级名称', '权重'])
            out = io.BytesIO()
            with pd.ExcelWriter(out, engine='openpyxl') as w: tmp.to_excel(w, index=False)
            st.download_button("下载部门模板", data=out.getvalue(), file_name="部门导入模板.xlsx")

            df_file = st.file_uploader("上传 Excel", type=["xlsx"], key="d_up")
            if df_file and st.button("开始执行部门导入"):
                in_df = pd.read_excel(df_file)
                for _, r in in_df.iterrows():
                    nm = clean_str(r.get('部门名称'))
                    if nm: add_department(nm, clean_str(r.get('性质')) or "其他", None, int(r.get('权重', 999)) if pd.notna(r.get('权重')) else 999)

                _, rd = get_all_departments(include_inactive=True); fdf = pd.DataFrame(rd)
                for _, r in in_df.iterrows():
                    nm = clean_str(r.get('部门名称')); pnm = clean_str(r.get('上级名称'))
                    if nm and pnm:
                        tc = fdf[fdf['dept_name'] == nm]; tp = fdf[fdf['dept_name'] == pnm]
                        if not tc.empty and not tp.empty:
                            update_department(
                                int(tc.iloc[0]['dept_id']), str(tc.iloc[0]['dept_name']), str(tc.iloc[0]['dept_category']),
                                int(tp.iloc[0]['dept_id']), int(tc.iloc[0]['sort_order']), int(tc.iloc[0]['status'])
                            )
                set_msg_and_rerun("部门架构生成完毕")

    with col_d2:
        st.subheader("📊 组织架构树")
        s1, s2 = st.columns([2, 1])
        with s1: d_s = st.text_input("🔍 搜索部门")
        with s2: st.write(""); show_i = st.checkbox("含已撤销")

        fdf = df_depts.copy()
        if not fdf.empty:
            if not show_i: fdf = fdf[fdf['status'] == 1]
            if d_s:
                m_ids = set()
                dh = fdf[fdf['dept_name'].str.contains(d_s, na=False)]['dept_id'].tolist()
                def trace(did, all_df):
                    m_ids.add(did); pid = all_df[all_df['dept_id']==did].iloc[0]['parent_dept_id']
                    if pd.notna(pid) and pid != 0: trace(pid, all_df)
                for h in dh: trace(h, df_depts)
                fdf = df_depts[df_depts['dept_id'].isin(m_ids)]
                if not show_i: fdf = fdf[fdf['status'] == 1]

            t_data = build_dept_tree(fdf)
            if t_data:
                tdf = pd.DataFrame(t_data)
                tdf['status'] = tdf['status'].apply(lambda x: "正常" if x == 1 else "已撤销")
                tdf = tdf.rename(columns=DEPT_COL_MAP)
                st.dataframe(tdf[['部门ID', '层级展示名', '性质', '权重', '状态']], use_container_width=True, hide_index=True)

# ==============================================================================
# 模块 B: 🎯 岗位字典
# ==============================================================================
elif current_page == "🎯 岗位字典":
    col_p1, col_p2 = st.columns([1, 2])
    with col_p1:
        st.subheader("📝 岗位维护")
        p_mode = st.radio("模式", ["新增岗位", "修改现有"], horizontal=True)
        t_pos = None
        if p_mode == "修改现有":
            if df_positions.empty: st.warning("无岗位")
            else:
                p_dict = {r['pos_id']: r['pos_name'] for _, r in df_positions.iterrows()}
                s_pid = st.selectbox("选择岗位", options=list(p_dict.keys()), format_func=lambda x: p_dict[x])
                t_pos = df_positions[df_positions['pos_id'] == s_pid].iloc[0]

        with st.form("pos_form"):
            p_name = st.text_input("岗位名称*", value=t_pos['pos_name'] if t_pos is not None else "")
            p_cat = st.text_input("岗位序列", value=t_pos['pos_category'] if t_pos is not None else "通用序列")
            p_sort = st.number_input("权重", value=int(t_pos['sort_order']) if t_pos is not None else 999)
            p_stat = "正常" if t_pos is None or t_pos['status'] == 1 else "停用"
            p_s_val = st.selectbox("状态", ["正常", "停用"], index=["正常", "停用"].index(p_stat))

            if st.form_submit_button("保存岗位"):
                if not p_name: st.error("必填")
                else:
                    sv = 1 if p_s_val == "正常" else 0
                    if p_mode == "新增岗位": ok, msg = add_position(p_name, p_cat, p_sort)
                    else: ok, msg = update_position(int(t_pos['pos_id']), p_name, p_cat, p_sort, sv)
                    if ok: set_msg_and_rerun(msg)
                    else: st.error(msg)

        with st.expander("📥 批量导入岗位字典"):
            tp = pd.DataFrame(columns=['岗位名称', '序列', '权重'])
            op = io.BytesIO()
            with pd.ExcelWriter(op, engine='openpyxl') as w: tp.to_excel(w, index=False)
            st.download_button("下载岗位模板", data=op.getvalue(), file_name="岗位导入模板.xlsx")

            pf = st.file_uploader("上传 Excel", type=["xlsx"], key="p_up")
            if pf and st.button("开始执行岗位导入"):
                idf = pd.read_excel(pf)
                sc = 0
                for _, r in idf.iterrows():
                    nm = clean_str(r.get('岗位名称'))
                    if nm:
                        ok, _ = add_position(nm, clean_str(r.get('序列')) or "通用序列", int(r.get('权重', 999)) if pd.notna(r.get('权重')) else 999)
                        if ok: sc += 1
                set_msg_and_rerun(f"导入完成，新增 {sc} 个岗位")

    with col_p2:
        st.subheader("📊 岗位清单")
        if not df_positions.empty:
            sh_p = st.checkbox("含已停用")
            fp = df_positions.copy()
            if not sh_p: fp = fp[fp['status']==1]
            fp['status'] = fp['status'].apply(lambda x: "正常" if x==1 else "停用")
            fp = fp.rename(columns=POS_COL_MAP)
            st.dataframe(fp[['岗位ID', '岗位名称', '序列', '权重', '状态']], use_container_width=True, hide_index=True)

# ==============================================================================
# 模块 C: 👥 人员档案
# ==============================================================================
elif current_page == "👥 人员档案":

    with st.expander("⏳ 实习生转正预警看板"):
        if not df_emps.empty and 'pos_name' in df_emps.columns:
            interns_df = df_emps[(df_emps['pos_name'] == '实习岗') & (df_emps['status'] == '在职')].copy()
            if not interns_df.empty:
                intern_list = []
                for _, intern in interns_df.iterrows():
                    start_p = intern['intern_start_date'] if pd.notna(intern['intern_start_date']) else intern['join_company_date']
                    start_d = pd.to_datetime(start_p)
                    if start_d:
                        months = 3 if str(intern.get('education_level')) == '硕士' else 6
                        expected = start_d + relativedelta(months=months)
                        days_left = (expected.date() - date.today()).days
                        st_str = f"🔴 已超期 {-days_left} 天" if days_left < 0 else (f"🟡 临近 ({days_left} 天)" if days_left <= 15 else f"🟢 正常 ({days_left} 天)")
                        intern_list.append({'工号': intern['emp_id'], '姓名': intern['name'], '部门': intern['dept_name'], '预计转正': expected.strftime('%Y-%m-%d'), '当前状态': st_str})
                if intern_list: st.dataframe(pd.DataFrame(intern_list), use_container_width=True, hide_index=True)
            else: st.success("🎉 无待转正人员")

    st.subheader("🔍 人员检索与档案控制")
    c1, c2, c3, c4 = st.columns(4)
    with c1: q_name = st.text_input("工号/姓名/社保编号")
    with c2: q_dept = st.multiselect("部门", options=df_depts['dept_name'].tolist())
    with c3: q_status = st.multiselect("状态", options=["在职", "离职", "退休","挂靠人员"], default=["在职", "挂靠人员"])

    f_df = df_emps.copy()
    if q_name: f_df = f_df[f_df['name'].str.contains(q_name) | f_df['emp_id'].str.contains(q_name)]
    if q_dept: f_df = f_df[f_df['dept_name'].isin(q_dept)]
    if q_status: f_df = f_df[f_df['status'].isin(q_status)]

    # ==========================================================================
    # [核心修复 1] 注入与财务台账同等级别的绝对排序基因
    # ==========================================================================
    if not f_df.empty:
        # 提取部门与岗位权重
        dept_weight_map = {r['dept_name']: r['sort_order'] for _, r in df_depts.iterrows()} if not df_depts.empty else {}
        pos_weight_map = {r['pos_name']: r['sort_order'] for _, r in df_positions.iterrows()} if not df_positions.empty else {}

        def get_emp_sort_keys(row):
            d_name = str(row.get('dept_name', ''))
            p_name = str(row.get('pos_name', ''))
            status = str(row.get('status', ''))

            # 1. 部门大权重
            if status == '退休' or '离退休' in d_name: d_weight = 9999
            elif status == '挂靠人员': d_weight = 9000  # [新增] 挂靠人员排在在职之后，离退休之前
            elif status == '公共账目' or '统筹' in d_name or '公共' in d_name: d_weight = 9998
            else: d_weight = dept_weight_map.get(d_name, 999)

            # 2. 岗位大权重 (图里总经理1，副总2)
            p_weight = pos_weight_map.get(p_name, 999)

            # 3. 个人锚点小权重 (带小数点的 21.2)
            try: r_weight = float(row.get('post_rank', 9999.0)) if pd.notna(row.get('post_rank')) else 9999.0
            except: r_weight = 9999.0

            return (d_weight, p_weight, r_weight, str(row.get('emp_id', '')))

        # 执行强力排序
        f_df['__sort_tuple__'] = f_df.apply(get_emp_sort_keys, axis=1)
        f_df[['__dw__', '__pw__', '__rw__', '__id__']] = pd.DataFrame(f_df['__sort_tuple__'].tolist(), index=f_df.index)
        f_df = f_df.sort_values(by=['__dw__', '__pw__', '__rw__', '__id__'], ascending=[True, True, True, True])
        f_df = f_df.drop(columns=['__sort_tuple__', '__dw__', '__pw__', '__rw__', '__id__'])

    # ==========================================================================
    # [核心修复 2] 满血版全量美化导出引擎
    # ==========================================================================
    with c4:
        st.write("")
        if not f_df.empty:
            import sqlite3, os
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            def format_roster_sheet(worksheet, df_columns):
                worksheet.freeze_panes = 'A2'
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                for col_idx, col_name in enumerate(df_columns, 1):
                    col_letter = get_column_letter(col_idx)
                    # 身份证、学校名字长一点，其它适中
                    worksheet.column_dimensions[col_letter].width = 20 if col_name in ['身份证号', '毕业院校', '专业'] else 13
                    for row_idx in range(1, worksheet.max_row + 1):
                        cell = worksheet[f"{col_letter}{row_idx}"]
                        cell.border = thin_border
                        if row_idx == 1:
                            cell.font = Font(bold=True)
                            cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')

            # 击穿底层：拉取扩展表数据实现“绝对全量”
            db_path = os.path.join('database', 'hr_core.db')
            conn = sqlite3.connect(db_path)
            profiles_df = pd.read_sql_query("SELECT * FROM employee_profiles", conn)
            conn.close()

            # 拼接并重命名
            out_df = pd.merge(f_df, profiles_df, on='emp_id', how='left')
            out_df = out_df.rename(columns={
                'emp_id': '工号/编号', 'name': '姓名', 'id_card': '身份证号', 'dept_name': '部门', 'pos_name': '岗位',
                'post_rank': '岗级', 'post_grade': '档次', 'tech_grade': 'T级', 'join_company_date': '入职日期',
                'status': '状态', 'education_level': '学历', 'degree': '学位', 'school_name': '毕业院校',
                'major': '专业', 'graduation_date': '毕业日期', 'first_job_date': '参加工作日期'
            })

            # 整理老板想看的专业顺序
            eo = ['工号/编号', '姓名', '部门', '岗位', '岗级', '档次', 'T级', '状态', '入职日期', '参加工作日期', '身份证号', '学历', '学位', '毕业院校', '专业', '毕业日期']
            out_df = out_df[[c for c in eo if c in out_df.columns]]

            # 执行导出与渲染
            ob = io.BytesIO()
            with pd.ExcelWriter(ob, engine='openpyxl') as w:
                out_df.to_excel(w, index=False, sheet_name='员工花名册')
                format_roster_sheet(w.sheets['员工花名册'], out_df.columns)

            st.download_button("📥 导出全量美化名单", data=ob.getvalue(), file_name=f"人员档案_{date.today()}.xlsx", type="primary")
        else:
            st.button("📥 导出全量美化名单", disabled=True)

    # 渲染到前端界面的展示列
    disp = f_df.rename(columns=EMP_COL_MAP)
    # [核心修复：补回防爆底线] 确保无论有没有选中人，变量都存在
    t_emp_sel = None
    if not disp.empty:
        target_cols = ['工号/编号', '姓名', '部门', '岗位', 'T级', '岗级', '档次', '状态']
        ui_cols = [c for c in target_cols if c in disp.columns]

        disp_ui = disp[ui_cols].copy()
        disp_ui.insert(0, "✅勾选修改", False)
        edited_df = st.data_editor(disp_ui, hide_index=True, use_container_width=True, key=st.session_state.editor_key, column_config={"✅勾选修改": st.column_config.CheckboxColumn(required=True)})
        selected_rows = edited_df[edited_df["✅勾选修改"] == True]
        if not selected_rows.empty:
            sel_id = selected_rows.iloc[0]['工号/编号']

            # 拉取完整底层数据用于回显
            import sqlite3, os
            db_path = os.path.join('database', 'hr_core.db')
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            full_emp_df = pd.read_sql_query("""
                SELECT e.*, p.* FROM employees e 
                LEFT JOIN employee_profiles p ON e.emp_id = p.emp_id 
                WHERE e.emp_id = ?
            """, conn, params=[sel_id])
            conn.close()

            t_emp_sel = full_emp_df.iloc[0].to_dict() if not full_emp_df.empty else df_emps[df_emps['emp_id'] == sel_id].iloc[0]
            st.info(f"已锁定: {t_emp_sel['name']}")

    st.divider()

    with st.expander("📥 批量智能导入 (支持离退免检与存在即更新)"):
        st.info("💡 若状态填为【离职/退休】，部门岗位可留空，系统将自动编入【离退休公共池】；若状态为【挂靠人员】，工号可填入社保编号。")
        t1, t2 = st.columns(2)
        with t1:
            icols = ['工号', '姓名', '状态', '所属部门', '岗位', '技术等级(T级)', '身份证号', '岗级', '档次', '入职日期', '参加工作日期', '学历', '学位', '毕业院校', '专业', '毕业日期']
            tmp = pd.DataFrame(columns=icols); tout = io.BytesIO()
            with pd.ExcelWriter(tout) as w: tmp.to_excel(w, index=False)
            st.download_button("下载人员模板", data=tout.getvalue(), file_name="人员导入模板.xlsx")
        with t2:
            uf = st.file_uploader("上传 Excel", type=["xlsx"], key="emp_up")
            if uf and st.button("开始执行智能导入"):
                idf = pd.read_excel(uf); sc = 0; uc = 0; errs = []
                ex_ids = df_emps['emp_id'].tolist() if not df_emps.empty else []
                _, c_d = get_all_departments(True); d_df = pd.DataFrame(c_d)
                _, c_p = get_all_positions(True); p_df = pd.DataFrame(c_p)
                for idx, row in idf.iterrows():
                    eid = clean_str(row.get('工号'))
                    if not eid: continue
                    status_val = clean_str(row.get('状态')) or '在职'
                    if status_val in ['离职', '退休']:
                        idn = "离退休公共池"; ipn = "无岗位"; rank_val = 0; grade_val = "-"
                    else:
                        idn = clean_str(row.get('所属部门')) or clean_str(row.get('部门'))
                        ipn = clean_str(row.get('岗位')) or clean_str(row.get('岗位名称'))
                        rank_val = int(row.get('岗级', 11)) if pd.notna(row.get('岗级')) else 11
                        grade_val = clean_str(row.get('档次')) or 'E'

                    td = None
                    if idn:
                        md = d_df[d_df['dept_name']==idn]
                        if not md.empty: td = int(md.iloc[0]['dept_id'])
                        else:
                            add_department(idn, "其他", None, 999); _, rd = get_all_departments(True); d_df = pd.DataFrame(rd)
                            td = int(d_df[d_df['dept_name']==idn].iloc[0]['dept_id'])
                    tp = None
                    if ipn:
                        mp = p_df[p_df['pos_name']==ipn]
                        if not mp.empty: tp = int(mp.iloc[0]['pos_id'])
                        else:
                            add_position(ipn, "通用", 999); _, rp = get_all_positions(True); p_df = pd.DataFrame(rp)
                            tp = int(p_df[p_df['pos_name']==ipn].iloc[0]['pos_id'])

                    idc_raw = clean_str(row.get('身份证号'))
                    ed = {'emp_id': eid, 'name': clean_str(row.get('姓名')), 'id_card': idc_raw if idc_raw else None, 'dept_id': td, 'post_rank': rank_val, 'post_grade': grade_val, 'status': status_val, 'join_company_date': clean_date(row.get('入职日期'))}
                    pd_info = {
                        'pos_id': tp,
                        'tech_grade': clean_str(row.get('技术等级(T级)')),
                        'title_order': 999,
                        'education_level': clean_str(row.get('学历')),
                        'degree': clean_str(row.get('学位')),
                        'school_name': clean_str(row.get('毕业院校')),
                        'major': clean_str(row.get('专业')),
                        'graduation_date': clean_date(row.get('毕业日期')),
                        'first_job_date': clean_date(row.get('参加工作日期'))
                    }

                    if eid in ex_ids:
                        ok, msg = update_employee(eid, ed, pd_info, reason="Excel批量覆盖更新")
                        if ok: uc += 1
                        else: errs.append(f"行{idx+2}更新失败: {msg}")
                    else:
                        ok, msg = add_employee(ed, pd_info, "初始批量导入")
                        if ok: sc += 1
                        else: errs.append(f"行{idx+2}新增失败: {msg}")
                if errs: st.error("部分记录异常:\n" + "\n".join(errs))
                set_msg_and_rerun(f"完成！新增 {sc} 人, 更新 {uc} 人")

    # [核心修复 2026-03-27] 极度严谨的类型推导，彻底根绝 ValueError 歧义死机
    with st.expander("📝 单条维护区", expanded=True):
        if t_emp_sel is None: st.warning("💡 新增模式。如果录入挂靠人员，工号处可直接填写其社保编号。")
        with st.form("emp_form", clear_on_submit=True):
            f1, f2, f3 = st.columns(3)
            with f1:
                fid = st.text_input("工号/社保编号*", value=str(t_emp_sel.get('emp_id', "")) if t_emp_sel is not None else "", help="挂靠人员请直接填写社保编号", disabled=(t_emp_sel is not None))
                fname = st.text_input("姓名*", value=str(t_emp_sel.get('name', "")) if t_emp_sel is not None else "")
                fidc = st.text_input("身份证", value=str(t_emp_sel.get('id_card', "")) if t_emp_sel is not None and pd.notna(t_emp_sel.get('id_card')) else "")
            with f2:
                vd = df_depts[df_depts['status']==1]; dm = {r['dept_id']: r['dept_name'] for _, r in vd.iterrows()}
                def_d = t_emp_sel.get('dept_id') if t_emp_sel is not None else None
                fdept = st.selectbox("部门*", options=list(dm.keys()), format_func=lambda x: dm[x], index=list(dm.keys()).index(def_d) if def_d in dm else 0) if dm else None

                # 强行开启小数支持，步长为 0.1
                frank = st.number_input("岗级*", 0.0, 28.0,
                                        float(t_emp_sel.get('post_rank', 11.0)) if t_emp_sel is not None and pd.notna(
                                            t_emp_sel.get('post_rank')) else 11.0, step=0.1, format="%.1f")

                g_opts = ["-","A","B","C","D","E","F","G","H","I","J"]
                cur_g = str(t_emp_sel.get('post_grade', 'E')) if t_emp_sel is not None and pd.notna(t_emp_sel.get('post_grade')) else 'E'
                fgrade = st.selectbox("档次*", g_opts, index=g_opts.index(cur_g) if cur_g in g_opts else 5)
            with f3:
                vp = df_positions[df_positions['status']==1]; pm = {r['pos_id']: r['pos_name'] for _, r in vp.iterrows()}
                def_p = t_emp_sel.get('pos_id') if t_emp_sel is not None else None
                fpos = st.selectbox("岗位*", options=list(pm.keys()), format_func=lambda x: pm[x], index=list(pm.keys()).index(def_p) if def_p in pm else 0) if pm else None

                ftg = st.text_input("T级", value=str(t_emp_sel.get('tech_grade', "")) if t_emp_sel is not None and pd.notna(t_emp_sel.get('tech_grade')) else "")

                fjoin_val = pd.to_datetime(t_emp_sel.get('join_company_date')) if t_emp_sel is not None and pd.notna(
                    t_emp_sel.get('join_company_date')) else date.today()
                # 强行突破 10 年限制
                fjoin = st.date_input("入职日", value=fjoin_val, min_value=date(1950, 1, 1))

                # 👇 新增：附加学籍与工作时间区 👇
                st.write("**--- 附加学籍与工作时间 ---**")
                fp1, fp2, fp3 = st.columns(3)
                with fp1:
                    e_opts = ["", "中专", "大专", "本科", "硕士", "博士"]
                    cur_edu = str(t_emp_sel.get('education_level', '')) if t_emp_sel is not None and pd.notna(
                        t_emp_sel.get('education_level')) else ""
                    f_edu = st.selectbox("学历", e_opts, index=e_opts.index(cur_edu) if cur_edu in e_opts else 0)
                    f_degree = st.text_input("学位", value=str(
                        t_emp_sel.get('degree', "")) if t_emp_sel is not None and pd.notna(
                        t_emp_sel.get('degree')) else "")
                with fp2:
                    f_school = st.text_input("毕业院校", value=str(
                        t_emp_sel.get('school_name', "")) if t_emp_sel is not None and pd.notna(
                        t_emp_sel.get('school_name')) else "")
                    f_major = st.text_input("所学专业",
                                            value=str(t_emp_sel.get('major', "")) if t_emp_sel is not None and pd.notna(
                                                t_emp_sel.get('major')) else "")
                with fp3:
                    # [核心修复 2]：强制转换为标准 date 对象，防止 Streamlit 时间轴崩溃
                    grad_raw = t_emp_sel.get('graduation_date') if t_emp_sel is not None else None
                    f_grad_date_val = pd.to_datetime(grad_raw).date() if pd.notna(grad_raw) and str(grad_raw).strip() != '' else None
                    f_grad_date = st.date_input("毕业日期", value=f_grad_date_val, min_value=date(1950, 1, 1))

                    first_work_raw = t_emp_sel.get('first_job_date') if t_emp_sel is not None else None
                    f_first_work_val = pd.to_datetime(first_work_raw).date() if pd.notna(first_work_raw) and str(first_work_raw).strip() != '' else None
                    f_first_work = st.date_input("参加工作时间*", value=f_first_work_val, min_value=date(1950, 1, 1))
                # 👆 新增结束 👆

            st.write("**--- 状态与快照控制 ---**")
            cs1, cs2, cs3 = st.columns(3)
            with cs1:
                cur_s = str(t_emp_sel.get('status', '在职')) if t_emp_sel is not None and pd.notna(t_emp_sel.get('status')) else "在职"
                s_opts = ["在职", "挂靠人员", "离职", "退休"]
                fst = st.selectbox("当前状态", s_opts, index=s_opts.index(cur_s) if cur_s in s_opts else 0)
            with cs2: fcd = st.date_input("生效日期*", value=date.today())
            with cs3: frsn = st.text_input("变动说明*", placeholder="必填")

            if st.form_submit_button("保存并生成快照"):
                if not fid or not fname or fdept is None: st.error("核心信息缺失")
                elif t_emp_sel is not None and not frsn: st.error("必填说明")
                else:
                    idc_val = fidc.strip() if fidc.strip() else None
                    # [核心修复 3]：将 int(frank) 强转改为 float(frank)，彻底保住 2.1 这种精细化锚点
                    ed = {'emp_id': fid, 'name': fname, 'id_card': idc_val, 'dept_id': int(fdept),
                          'post_rank': float(frank), 'post_grade': fgrade, 'status': fst,
                          'join_company_date': fjoin.strftime('%Y-%m-%d')}
                    # 👇 修改：将找回的字段存入扩展表字典 👇
                    pd_i = {
                        'pos_id': int(fpos) if fpos else None,
                        'tech_grade': ftg,
                        'title_order': 999,
                        'education_level': f_edu,
                        'degree': f_degree,
                        'school_name': f_school,
                        'major': f_major,
                        'graduation_date': f_grad_date.strftime('%Y-%m-%d') if f_grad_date else None,
                        'first_job_date': f_first_work.strftime('%Y-%m-%d') if f_first_work else None
                    }
                    # 👆 修改结束 👆
                    ad_str = fcd.strftime('%Y-%m-%d %H:%M:%S')
                    if t_emp_sel is not None: ok, msg = update_employee(fid, ed, pd_i, reason=frsn, change_date=ad_str)
                    else: ok, msg = add_employee(ed, pd_i, reason=frsn or "手工录入", change_date=ad_str)
                    if ok: set_msg_and_rerun(msg)
                    else: st.error(msg)

# ==============================================================================
# 模块 D: 🕰️ 历史变动流水 (全量归位侧边栏审计)
# ==============================================================================
elif current_page == "🕰️ 历史变动流水":
    st.subheader("🕰️ 全生命周期时空审计")
    ok, h_list = get_all_history()
    if ok and h_list:
        hdf = pd.DataFrame(h_list); hdf['dt_obj'] = pd.to_datetime(hdf['change_date'])
        hdf = hdf.sort_values(by='dt_obj', ascending=False)

        # [满血复原] 高级过滤与审计侧边栏
        with st.sidebar.expander("🔍 高级筛选与审计导出", expanded=True):
            hs = st.text_input("搜姓名/工号")
            ht = st.multiselect("变动类型", options=sorted(hdf['change_type'].unique().tolist()))
            dept_opts = sorted([str(d) for d in hdf['old_dept_name'].unique().tolist() if d])
            hd = st.multiselect("原归属部门", options=dept_opts)
            d_range = st.date_input("生效时段", value=(date.today() - relativedelta(months=3), date.today()))

        f_h = hdf.copy()
        if hs: f_h = f_h[f_h['emp_name'].str.contains(hs, na=False) | f_h['emp_id'].str.contains(hs, na=False)]
        if ht: f_h = f_h[f_h['change_type'].isin(ht)]
        if hd: f_h = f_h[f_h['old_dept_name'].isin(hd)]
        if len(d_range) == 2:
            # 1. 剔除因没填入职时间导致的空时间(NaT)，防止引擎比对时死机
            f_h = f_h.dropna(subset=['dt_obj'])

            # 2. 将你选择的日期强转为 Pandas 的时间戳，并包裹一整天的时间范围
            start_dt = pd.to_datetime(d_range[0])
            end_dt = pd.to_datetime(d_range[1]) + pd.Timedelta(days=1, seconds=-1)

            # 3. 同类型安全比对
            f_h = f_h[(f_h['dt_obj'] >= start_dt) & (f_h['dt_obj'] <= end_dt)]

        with st.sidebar:
            if not f_h.empty:
                audit_cols = {'change_date': '生效日期', 'emp_name': '姓名', 'emp_id': '工号', 'change_type': '变动类型', 'old_dept_name': '原部门', 'new_dept_name': '新部门', 'old_pos_name': '原岗位', 'new_pos_name': '新岗位', 'old_tech_grade': '原T级', 'new_tech_grade': '新T级', 'old_post_rank': '原岗级', 'new_post_rank': '新岗级', 'old_post_grade': '原档次', 'new_post_grade': '新档次', 'change_reason': '说明'}
                export_final = f_h[list(audit_cols.keys())].rename(columns=audit_cols)
                ob = io.BytesIO()
                with pd.ExcelWriter(ob, engine='openpyxl') as w: export_final.to_excel(w, index=False)
                st.download_button("📥 导出筛选流水", data=ob.getvalue(), file_name=f"变动审计_{date.today()}.xlsx", type="primary", use_container_width=True)

        if not f_h.empty:
            for _, row in f_h.iterrows():
                with st.container(border=True):
                    h1, h2 = st.columns([1, 5])
                    with h1:
                        st.write(f"**{row['emp_name']}**")
                        st.caption(f"工号: {row['emp_id']}")
                        type_str = row['change_type']
                        if '离职' in type_str or '退休' in type_str or '变为' in type_str: st.error(f"🛑 {type_str}")
                        elif '实习转正' in type_str: st.warning(f"🔥 {type_str}")
                        else: st.info(type_str)
                    with h2:
                        c1, c2, c3, c4 = st.columns(4)
                        with c1: st.caption("🏢 部门变动"); st.write(f"{row['old_dept_name'] or '-'}\n➡️\n{row['new_dept_name'] or '-'}")
                        with c2: st.caption("💼 岗位变动"); st.write(f"{row['old_pos_name'] or '-'}\n➡️\n{row['new_pos_name'] or '-'}")
                        with c3: st.caption("🏅 T级变动"); st.write(f"{row['old_tech_grade'] or '-'}\n➡️\n{row['new_tech_grade'] or '-'}")
                        with c4: st.caption("💰 待遇对比"); st.write(f"{row['old_post_rank'] or '-'}岗{row['old_post_grade'] or '-'}\n➡️\n{row['new_post_rank'] or '-'}岗{row['new_post_grade'] or '-'}")
                        st.divider()
                        st.caption(f"🕰️ 生效: {row['change_date']} | 说明: {row['change_reason']}")
                        if st.button("⏪ 撤销", key=f"rb_{row['change_id']}"):
                            ro, rm = rollback_history(row['change_id'])
                            if ro: set_msg_and_rerun(rm)
                            else: st.error(rm)
        else: st.warning("暂无匹配流水")
    else: st.info("尚未产生记录")