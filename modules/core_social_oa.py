"""OA 社保数据接口导出。

严格复刻现有 OA 模板的两张工作表：
1. Sheet0：中文业务数据；
2. Columns：OA 内部字段编码，导出时隐藏。

数据只读取已经固化的 social_monthly_items，不重新计算社保金额。
"""

import io
import os
import re
import sqlite3
import zipfile
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from modules.core_arrangements import resolve_social_route, social_export_included


class OAExportValidationError(ValueError):
    """OA 接口数据不完整，禁止生成可能漏人的上传文件。"""


COMMON_HEADERS = [
    '行号', '人员编码', '人员姓名', '社保账号', '基金基数', '基金基数（单位）',
    '单位缴纳比例', '单位缴交额', '个人缴纳比例', '个人缴交额',
    '单位补缴额', '个人补缴额', '个人缴交合计', '单位缴交合计',
]
COMMON_COLUMNS = [
    'other_rownum', 'bd_psndoc_code', 'bd_psndoc_name', 'bm_data_bmaccountno',
    'bm_data_f_1', 'bm_data_f_2', 'bm_data_f_3', 'bm_data_f_5',
    'bm_data_f_8', 'bm_data_f_10', 'bm_data_f_11', 'bm_data_f_13',
    'bm_data_f_30', 'bm_data_f_31',
]
MEDICAL_HEADERS = [
    '行号', '人员编码', '人员姓名', '社保账号', '基金基数', '基金基数（单位）',
    '单位缴纳比例', '单位缴纳固定值', '单位缴交额', '个人缴纳比例',
    '个人缴纳固定值', '个人缴交额', '单位补缴额', '个人补缴额',
    '个人缴交合计', '单位缴交合计',
]
MEDICAL_COLUMNS = [
    'other_rownum', 'bd_psndoc_code', 'bd_psndoc_name', 'bm_data_bmaccountno',
    'bm_data_f_1', 'bm_data_f_2', 'bm_data_f_3', 'bm_data_f_4',
    'bm_data_f_5', 'bm_data_f_8', 'bm_data_f_9', 'bm_data_f_10',
    'bm_data_f_11', 'bm_data_f_13', 'bm_data_f_30', 'bm_data_f_31',
]
FUND_HEADERS = [
    '人员编码', '人员姓名', '基金基数', '基金基数（单位）',
    '个人补缴额', '单位补缴额',
]
FUND_COLUMNS = [
    'bd_psndoc_code', 'bd_psndoc_name', 'bm_data_f_1',
    'bm_data_f_2', 'bm_data_f_13', 'bm_data_f_11',
]
ANNUITY_HEADERS = [
    '行号', '人员编码', '人员姓名', '社保账号', '基金基数',
    '单位缴纳比例', '单位缴交额', '个人缴纳比例', '个人缴交额',
    '单位补缴额', '个人补缴额', '个人缴交合计', '单位缴交合计',
]
ANNUITY_COLUMNS = [
    'other_rownum', 'bd_psndoc_code', 'bd_psndoc_name', 'bm_data_bmaccountno',
    'bm_data_f_1', 'bm_data_f_3', 'bm_data_f_5', 'bm_data_f_8',
    'bm_data_f_10', 'bm_data_f_11', 'bm_data_f_13',
    'bm_data_f_30', 'bm_data_f_31',
]

OA_EXPORT_DEFINITIONS = OrderedDict([
    ('injury', {'display_name': '工伤保险', 'headers': COMMON_HEADERS, 'columns': COMMON_COLUMNS}),
    ('pension', {'display_name': '基本养老保险', 'headers': COMMON_HEADERS, 'columns': COMMON_COLUMNS}),
    ('medical', {'display_name': '基本医疗保险', 'headers': MEDICAL_HEADERS, 'columns': MEDICAL_COLUMNS}),
    ('unemp', {'display_name': '失业保险', 'headers': COMMON_HEADERS, 'columns': COMMON_COLUMNS}),
    ('maternity', {'display_name': '生育保险', 'headers': COMMON_HEADERS, 'columns': COMMON_COLUMNS}),
    ('fund', {'display_name': '住房公积金', 'headers': FUND_HEADERS, 'columns': FUND_COLUMNS}),
    ('annuity', {'display_name': '企业年金', 'headers': ANNUITY_HEADERS, 'columns': ANNUITY_COLUMNS}),
])

RETRO_TYPE_TO_ITEM = {
    '养老保险': 'pension',
    '基本养老保险': 'pension',
    '医疗保险': 'medical',
    '基本医疗保险': 'medical',
    '大病医疗': 'medical',
    '失业保险': 'unemp',
    '工伤保险': 'injury',
    '生育保险': 'maternity',
    '住房公积金': 'fund',
    '企业年金': 'annuity',
}


def _get_db_connection():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(current_dir)
    db_path = os.environ.get(
        'MAKE_HR_DB_PATH', os.path.join(project_root, 'database', 'hr_core.db')
    )
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def _validate_month(month):
    value = str(month or '').strip()
    if not re.fullmatch(r'\d{4}-(0[1-9]|1[0-2])', value):
        raise OAExportValidationError("月份必须使用 YYYY-MM 格式")
    return value


def _text_account(value, fallback='0'):
    text = str(value or '').strip()
    if text and text.lower() not in {'none', 'nan'}:
        return text
    fallback_text = str(fallback or '').strip()
    return fallback_text or '0'


def _money(value):
    return round(float(value or 0), 2)


def _rate(amount, base):
    base_value = float(base or 0)
    if not base_value:
        return 0
    return round(float(amount or 0) / base_value, 10)


def _load_policy_rules(conn, month, entity):
    entity_name = str(entity or '省公众')
    row = conn.execute("""
        SELECT *
        FROM ss_policy_versions
        WHERE manage_entity=? AND effective_from_month<=?
        ORDER BY effective_from_month DESC
        LIMIT 1
    """, (entity_name, month)).fetchone()
    if row:
        return dict(row)
    row = conn.execute("""
        SELECT *
        FROM ss_policy_rules
        WHERE manage_entity=? AND rule_year<=?
        ORDER BY rule_year DESC
        LIMIT 1
    """, (entity_name, month[:4])).fetchone()
    return dict(row) if row else {}


def _resolve_interface_base(item, stored_base, company_amount, personal_amount, rules):
    """兼容旧月度明细曾把同一基数回填到所有险种的历史数据。"""
    base = float(stored_base or 0)
    if item in {'pension', 'medical', 'unemp', 'injury', 'maternity'}:
        upper = float(rules.get(f'{item}_upper') or 0)
        lower = float(rules.get(f'{item}_lower') or 0)
        if upper and base > upper:
            base = upper
        if lower and 0 < base < lower:
            base = lower
        return base

    company_rate = float(rules.get(f'{item}_comp_rate') or 0)
    personal_rate = float(rules.get(f'{item}_pers_rate') or 0)
    # 公积金 OA 模板展示的是由实际单边缴交额反推的执行基数；
    # 年金也用实际8%单位缴交额反推，避免旧月度回填了错误的公共基数。
    if item in {'fund', 'annuity'}:
        if company_rate and abs(float(company_amount or 0)) > 0:
            return round(float(company_amount) / company_rate, 2)
        if personal_rate and abs(float(personal_amount or 0)) > 0:
            return round(float(personal_amount) / personal_rate, 2)
    return base


def get_oa_export_settings(conn=None):
    owns_connection = conn is None
    conn = conn or _get_db_connection()
    try:
        rows = conn.execute("""
            SELECT insurance_item, display_name, file_name, sort_order
            FROM social_oa_export_settings
            ORDER BY sort_order, insurance_item
        """).fetchall()
        return [dict(row) for row in rows]
    finally:
        if owns_connection:
            conn.close()


def update_oa_export_filenames(file_names):
    """保存 7 个 OA 文件名；禁止目录、重名和非 xlsx 后缀。"""
    expected = set(OA_EXPORT_DEFINITIONS)
    supplied = set(file_names)
    if supplied != expected:
        missing = sorted(expected - supplied)
        extra = sorted(supplied - expected)
        raise OAExportValidationError(
            f"文件名配置不完整，缺少：{missing or '无'}；多余：{extra or '无'}"
        )

    normalized = {}
    for item, raw_name in file_names.items():
        name = str(raw_name or '').strip()
        if not name or name != os.path.basename(name) or '/' in name or '\\' in name:
            raise OAExportValidationError(f"{OA_EXPORT_DEFINITIONS[item]['display_name']}文件名不合法")
        if not name.lower().endswith('.xlsx'):
            raise OAExportValidationError(f"{name} 必须以 .xlsx 结尾")
        normalized[item] = name
    if len(set(normalized.values())) != len(normalized):
        raise OAExportValidationError("7 个 OA 文件名不能重复")

    conn = _get_db_connection()
    try:
        conn.executemany("""
            UPDATE social_oa_export_settings
            SET file_name=?, updated_at=CURRENT_TIMESTAMP
            WHERE insurance_item=?
        """, [(name, item) for item, name in normalized.items()])
        conn.commit()
    finally:
        conn.close()
    return True, "OA接口文件名已保存"


def _is_proxy_employee(conn, emp_id, month):
    month_start = f"{month}-01"
    return conn.execute("""
        SELECT 1
        FROM employee_arrangements
        WHERE emp_id=?
          AND arrangement_type='proxy_social'
          AND start_date <= date(?, '+1 month', '-1 day')
          AND date(COALESCE(actual_end_date, planned_end_date, '9999-12-31')) >= date(?)
        LIMIT 1
    """, (emp_id, month_start, month_start)).fetchone() is not None


def _load_monthly_rows(conn, month):
    rows = conn.execute("""
        SELECT i.emp_id, i.insurance_item, i.base_amount,
               i.company_amount, i.personal_amount,
               i.calculation_policy_entity,
               COALESCE(i.business_type_snapshot, 'normal') AS business_type_snapshot,
               i.amount_source, i.settlement_mode,
               i.payment_export_included, i.oa_export_included,
               e.employee_no, e.name,
               COALESCE(NULLIF(TRIM(p.oa_social_account_no), ''), '0') AS social_account,
               COALESCE(
                   NULLIF(TRIM(p.oa_annuity_account_no), ''),
                   NULLIF(TRIM(p.oa_social_account_no), ''),
                   '0'
               ) AS annuity_account
        FROM social_monthly_items i
        JOIN employees e ON e.emp_id=i.emp_id
        LEFT JOIN employee_profiles p ON p.emp_id=i.emp_id
        WHERE i.cost_month=?
        ORDER BY
            CASE WHEN e.employee_no IS NULL OR TRIM(e.employee_no)='' THEN 1 ELSE 0 END,
            e.employee_no, e.name, i.insurance_item
    """, (month,)).fetchall()
    return [dict(row) for row in rows]


def _load_retro_rows(conn, month):
    rows = conn.execute("""
        SELECT r.emp_id, r.retro_type,
               SUM(COALESCE(r.total_comp_retro, 0)) AS company_retro,
               SUM(COALESCE(r.total_pers_retro, 0)) AS personal_retro,
               e.employee_no, e.name,
               COALESCE(NULLIF(TRIM(p.oa_social_account_no), ''), '0') AS social_account,
               COALESCE(
                   NULLIF(TRIM(p.oa_annuity_account_no), ''),
                   NULLIF(TRIM(p.oa_social_account_no), ''),
                   '0'
               ) AS annuity_account
        FROM ss_retroactive_records r
        JOIN employees e ON e.emp_id=r.emp_id
        LEFT JOIN employee_profiles p ON p.emp_id=r.emp_id
        WHERE r.process_month=?
          AND COALESCE(r.status, '') NOT IN ('已撤销', '作废')
        GROUP BY r.emp_id, r.retro_type, e.employee_no, e.name,
                 p.oa_social_account_no, p.oa_annuity_account_no
        ORDER BY e.employee_no, e.name
    """, (month,)).fetchall()
    return [dict(row) for row in rows]


def build_oa_export_datasets(month, conn=None):
    """返回 7 个 OA 数据集和校验摘要，不生成 Excel。"""
    month = _validate_month(month)
    owns_connection = conn is None
    conn = conn or _get_db_connection()
    try:
        monthly_rows = _load_monthly_rows(conn, month)
        if not monthly_rows:
            raise OAExportValidationError(f"{month} 尚未固化社保明细")

        serious_by_emp = {
            row['emp_id']: _money(row['personal_amount'])
            for row in monthly_rows
            if row['insurance_item'] == 'medical_serious'
            and social_export_included(row, 'oa')
        }
        rows_by_item = {item: OrderedDict() for item in OA_EXPORT_DEFINITIONS}
        for row in monthly_rows:
            item = row['insurance_item']
            if item not in rows_by_item or not social_export_included(row, 'oa'):
                continue
            rows_by_item[item][row['emp_id']] = {
                **row,
                'company_retro': 0.0,
                'personal_retro': 0.0,
            }

        for retro in _load_retro_rows(conn, month):
            item = RETRO_TYPE_TO_ITEM.get(str(retro['retro_type'] or '').strip())
            if not item:
                continue
            retro_route = resolve_social_route(
                retro['emp_id'], item, month, conn=conn
            )
            if not social_export_included(retro_route, 'oa'):
                continue
            target = rows_by_item[item].get(retro['emp_id'])
            if target is None:
                target = {
                    'emp_id': retro['emp_id'],
                    'insurance_item': item,
                    'base_amount': 0.0,
                    'company_amount': 0.0,
                    'personal_amount': 0.0,
                    'calculation_policy_entity': '省公众',
                    'business_type_snapshot': 'normal',
                    'employee_no': retro['employee_no'],
                    'name': retro['name'],
                    'social_account': retro['social_account'],
                    'annuity_account': retro['annuity_account'],
                    'company_retro': 0.0,
                    'personal_retro': 0.0,
                }
                rows_by_item[item][retro['emp_id']] = target
            target['company_retro'] += _money(retro['company_retro'])
            target['personal_retro'] += _money(retro['personal_retro'])

        result = OrderedDict()
        missing_employee_numbers = OrderedDict()
        for item, definition in OA_EXPORT_DEFINITIONS.items():
            export_rows = []
            for source in rows_by_item[item].values():
                base = float(source.get('base_amount') or 0)
                company = _money(source.get('company_amount'))
                personal = _money(source.get('personal_amount'))
                rules = _load_policy_rules(
                    conn, month, source.get('calculation_policy_entity')
                )
                base = _resolve_interface_base(
                    item, base, company, personal, rules
                )
                company_retro = _money(source.get('company_retro'))
                personal_retro = _money(source.get('personal_retro'))
                serious = serious_by_emp.get(source['emp_id'], 0.0) if item == 'medical' else 0.0
                if not any(abs(value) > 0 for value in (
                    base, company, personal, company_retro, personal_retro, serious
                )):
                    continue
                employee_no = str(source.get('employee_no') or '').strip()
                if not employee_no:
                    missing_employee_numbers[source['emp_id']] = source['name']
                export_rows.append({
                    'emp_id': source['emp_id'],
                    'employee_no': employee_no,
                    'name': str(source.get('name') or ''),
                    'account': _text_account(
                        source.get('annuity_account')
                        if item == 'annuity' else source.get('social_account')
                    ),
                    'base': base,
                    'company_rate': float(
                        rules.get(f'{item}_comp_rate')
                        if rules.get(f'{item}_comp_rate') is not None
                        else _rate(company, base)
                    ),
                    'company_amount': company,
                    'personal_rate': float(
                        rules.get(f'{item}_pers_rate')
                        if rules.get(f'{item}_pers_rate') is not None
                        else _rate(personal, base)
                    ),
                    'personal_amount': personal,
                    'personal_fixed': _money(serious),
                    'company_retro': company_retro,
                    'personal_retro': personal_retro,
                    'company_total': _money(company + company_retro),
                    'personal_total': _money(personal + serious + personal_retro),
                })

            result[item] = {
                'definition': definition,
                'rows': export_rows,
                'row_count': len(export_rows),
                'company_total': _money(sum(row['company_total'] for row in export_rows)),
                'personal_total': _money(sum(row['personal_total'] for row in export_rows)),
                'zero_account_count': (
                    None if item == 'fund'
                    else sum(row['account'] == '0' for row in export_rows)
                ),
            }

        return {
            'month': month,
            'items': result,
            'missing_employee_numbers': list(missing_employee_numbers.values()),
            'excluded_proxy_people': len({
                row['emp_id'] for row in monthly_rows
                if row['business_type_snapshot'] == 'proxy_social'
            }),
            'excluded_down_people': len({
                row['emp_id'] for row in monthly_rows
                if row['business_type_snapshot'] == 'down_secondment'
            }),
        }
    finally:
        if owns_connection:
            conn.close()


def get_oa_export_summary(month):
    data = build_oa_export_datasets(month)
    return {
        'month': data['month'],
        'missing_employee_numbers': data['missing_employee_numbers'],
        'excluded_proxy_people': data['excluded_proxy_people'],
        'excluded_down_people': data['excluded_down_people'],
        'items': [
            {
                'insurance_item': item,
                'display_name': value['definition']['display_name'],
                'row_count': value['row_count'],
                'company_total': value['company_total'],
                'personal_total': value['personal_total'],
                'zero_account_count': value['zero_account_count'],
            }
            for item, value in data['items'].items()
        ],
    }


def _append_interface_row(item, row_number, row):
    if item == 'fund':
        return [
            row['employee_no'], row['name'], row['base'], row['base'],
            row['personal_retro'], row['company_retro'],
        ]
    if item == 'medical':
        return [
            row_number, row['employee_no'], row['name'], row['account'],
            row['base'], row['base'], row['company_rate'], 0,
            row['company_amount'], row['personal_rate'], row['personal_fixed'],
            row['personal_amount'], row['company_retro'], row['personal_retro'],
            row['personal_total'], row['company_total'],
        ]
    if item == 'annuity':
        return [
            row_number, row['employee_no'], row['name'], row['account'],
            row['base'], row['company_rate'], row['company_amount'],
            row['personal_rate'], row['personal_amount'], row['company_retro'],
            row['personal_retro'], row['personal_total'], row['company_total'],
        ]
    return [
        row_number, row['employee_no'], row['name'], row['account'],
        row['base'], row['base'], row['company_rate'], row['company_amount'],
        row['personal_rate'], row['personal_amount'], row['company_retro'],
        row['personal_retro'], row['personal_total'], row['company_total'],
    ]


def build_oa_workbook(item, rows):
    if item not in OA_EXPORT_DEFINITIONS:
        raise OAExportValidationError(f"不支持的险种：{item}")
    definition = OA_EXPORT_DEFINITIONS[item]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Sheet0'
    sheet.append(definition['headers'])
    for index, row in enumerate(rows, start=1):
        sheet.append(_append_interface_row(item, index, row))

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    sheet.freeze_panes = 'A2'

    text_headers = {'人员编码', '人员姓名', '社保账号'}
    for column_index, header in enumerate(definition['headers'], start=1):
        sheet.column_dimensions[sheet.cell(1, column_index).column_letter].width = max(
            12, min(18, len(header) * 2 + 2)
        )
        if header in text_headers:
            for row_index in range(2, sheet.max_row + 1):
                sheet.cell(row_index, column_index).number_format = '@'
        elif '比例' in header:
            for row_index in range(2, sheet.max_row + 1):
                sheet.cell(row_index, column_index).number_format = '0.0000'
        elif header not in {'行号'}:
            for row_index in range(2, sheet.max_row + 1):
                sheet.cell(row_index, column_index).number_format = '0.00'

    columns_sheet = workbook.create_sheet('Columns')
    columns_sheet.append(definition['columns'])
    columns_sheet.sheet_state = 'hidden'

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_oa_export_package(month):
    """生成 7 个独立文件和一个 ZIP；ZIP 内文件名保持 OA 要求。"""
    data = build_oa_export_datasets(month)
    missing = data['missing_employee_numbers']
    if missing:
        shown = '、'.join(missing[:10])
        suffix = f"等{len(missing)}人" if len(missing) > 10 else ''
        raise OAExportValidationError(
            f"以下人员尚无工号/人力编码，OA无法匹配，已阻止导出：{shown}{suffix}"
        )

    settings = {
        row['insurance_item']: row['file_name']
        for row in get_oa_export_settings()
    }
    files = []
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as archive:
        for item, value in data['items'].items():
            file_name = settings[item]
            content = build_oa_workbook(item, value['rows'])
            archive.writestr(file_name, content)
            files.append({
                'insurance_item': item,
                'display_name': value['definition']['display_name'],
                'file_name': file_name,
                'content': content,
                'row_count': value['row_count'],
            })
    return {
        'month': data['month'],
        'files': files,
        'zip_file_name': f"{data['month']}_OA社保数据接口.zip",
        'zip_content': zip_buffer.getvalue(),
        'excluded_proxy_people': data['excluded_proxy_people'],
        'excluded_down_people': data['excluded_down_people'],
    }
